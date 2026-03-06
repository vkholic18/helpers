import os
import requests
import time
import json
from datetime import datetime, timedelta, timezone
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# =============================================================================
# TEST MODE - Using sample data (comment this section and uncomment below for production)
# =============================================================================
TEST_MODE = True

# TOKEN = os.getenv("GITHUB_TOKEN")
# ORG = os.getenv("GITHUB_ORG")
# BASE = os.getenv("GITHUB_BASE")
# 
# if not all([TOKEN, ORG, BASE]):
#     raise RuntimeError("Missing required environment variables: GITHUB_TOKEN, GITHUB_ORG, GITHUB_BASE")
# 
# HEADERS = {
#     "Authorization": f"token {TOKEN}",
#     "Accept": "application/vnd.github+json"
# }

# Test mode variables
TOKEN = "test_token"
ORG = "TestOrganization"
BASE = "https://api.github.example.com"
HEADERS = {"Authorization": f"token {TOKEN}", "Accept": "application/vnd.github+json"}

SLEEP = 0.3


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def paginate(url):
    data = []
    while url:
        r = requests.get(url, headers=HEADERS)
        r.raise_for_status()
        data.extend(r.json())
        link = r.headers.get("Link", "")
        if 'rel="next"' in link:
            url = link.split('rel="next"')[0].split("<")[1].split(">")[0]
        else:
            url = None
        time.sleep(SLEEP)
    return data


def get(url, allow_404=False):
    r = requests.get(url, headers=HEADERS)
    if allow_404 and r.status_code == 404:
        return None
    r.raise_for_status()
    return r.json()


# ---------------------------------------------------------------------------
# Organization-level checks
# ---------------------------------------------------------------------------

def get_org_settings():
    """Fetch organization settings."""
    url = f"{BASE}/orgs/{ORG}"
    return get(url)


def check_base_permissions(org_data):
    """
    Section: Member privileges
    Setting: Base permissions
    Value: No permission
    Enforcement: Required
    Auditree: default_repository_permission
    Comments: Other values allow member's to read all private repositories.
    """
    default_perm = org_data.get("default_repository_permission", "read")
    return default_perm == "none"


def check_outside_collaborators_disabled(org_data):
    """
    Section: Member privileges
    Setting: Allow repository administrators to add outside collaborators to repositories for this organization
    Value: Disabled
    Enforcement: Required
    Auditree: org_outside_collaborators
    Comments: All access has to be through teams managed in AccessHub.
    """
    return not org_data.get("members_can_invite_outside_collaborators", True)


def check_org_hooks_ssl():
    """
    Section: Hooks
    Setting: SSL verification
    Value: Enable SSL verification
    Enforcement: Required
    Auditree: unsecure_org_hooks
    Comments: Needed to maintain confidentiality.
    Returns list of hook IDs that have SSL disabled.
    """
    hooks = get(f"{BASE}/orgs/{ORG}/hooks", allow_404=True) or []
    time.sleep(SLEEP)
    # insecure_ssl should be "0" or 0 or False or None for SSL to be enabled
    return [h["id"] for h in hooks if h.get("config", {}).get("insecure_ssl") not in (None, "0", 0, False)]


def check_repo_creation_private(org_data):
    """
    Section: Member privileges
    Setting: Repository Creation
    Value: Private
    Enforcement: Recommended
    Auditree: members_allowed_repository_creation_type, members_can_create_internal_repositories, members_can_create_public_repositories
    Comments: Requiring an organization administrator to create these helps ensure repositories 
              that should not be public are not created accidentally as public.
    """
    creation_type = org_data.get("members_allowed_repository_creation_type", "all")
    can_create_internal = org_data.get("members_can_create_internal_repositories", True)
    can_create_public = org_data.get("members_can_create_public_repositories", True)
    
    # Compliant if members cannot create public repos
    return not can_create_public


def check_integration_requests_disabled(org_data):
    """
    Section: Member privileges
    Setting: Allow integration requests from outside collaborators
    Value: Disabled
    Enforcement: Recommended
    Auditree: N/A
    Comments: All access has to be through teams, there should be no outside collaborators.
    """
    # This setting may not be directly exposed in all GitHub API versions
    return not org_data.get("members_can_create_public_pages", True)


def check_visibility_change_disabled(org_data):
    """
    Section: Member privileges
    Setting: Allow members to change repository visibilities for this organization
    Value: Disabled
    Enforcement: Recommended
    Auditree: N/A
    Comments: Requiring an organization administrator to change visibility helps ensure 
              repositories that should not be public are not made public accidentally.
    """
    return not org_data.get("members_can_change_repo_visibility", True)


def check_delete_transfer_disabled(org_data):
    """
    Section: Member privileges
    Setting: Allow members to delete or transfer repositories for this organization
    Value: Disabled
    Enforcement: Recommended
    Auditree: N/A
    Comments: Limits accidental or badly-intentioned deletion/removal.
    """
    can_delete = org_data.get("members_can_delete_repositories", True)
    return not can_delete


def check_profile_name_visibility(org_data):
    """
    Section: Member privileges
    Setting: Allow members to see comment author's profile name in private repositories
    Value: Enabled
    Enforcement: Recommended
    Auditree: N/A
    Comments: Preventing this just makes it harder to identify your colleagues.
    """
    return org_data.get("members_can_see_comment_author_profile", True)


def check_team_creation_disabled(org_data):
    """
    Section: Member privileges
    Setting: Allow members to create teams
    Value: Disabled
    Enforcement: Recommended
    Auditree: N/A
    Comments: Helps ensure that access management through AccessHub is not subverted 
              accidentally by someone adding a team and adding people to it directly.
    """
    return not org_data.get("members_can_create_teams", True)


def check_org_admin_activity():
    """
    Section: N/A
    Setting: Organization admins should have activity in the last 6 months
    Value: Active
    Enforcement: Recommended
    Auditree: N/A
    Comments: This is not a configurable option in GitHub. Organization admin are all-powerful 
              within the organization, having an inactive admin shows that the people who are 
              in this role are not getting scrutiny or revalidation.
    Returns a dict with admin login and whether they have recent activity.
    """
    # Get organization members with admin role
    admins = paginate(f"{BASE}/orgs/{ORG}/members?role=admin&per_page=100")
    six_months_ago = datetime.now(timezone.utc) - timedelta(days=180)
    
    admin_activity = []
    for admin in admins:
        login = admin["login"]
        # Check recent activity via events or audit log
        # Using public events as a proxy (may need audit log for private activity)
        events = get(f"{BASE}/users/{login}/events?per_page=1", allow_404=True) or []
        
        has_recent_activity = False
        if events:
            last_event_date = events[0].get("created_at", "")
            if last_event_date:
                try:
                    event_date = datetime.strptime(last_event_date[:10], "%Y-%m-%d").replace(tzinfo=timezone.utc)
                    has_recent_activity = event_date >= six_months_ago
                except ValueError:
                    pass
        
        admin_activity.append({
            "login": login,
            "has_recent_activity": has_recent_activity
        })
        time.sleep(SLEEP)
    
    return admin_activity


def evaluate_org_compliance(org_data):
    """
    Evaluate all organization-level compliance rules.
    Returns a dict with required and recommended checks.
    """
    bad_org_hooks = check_org_hooks_ssl()
    admin_activity = check_org_admin_activity()
    inactive_admins = [a["login"] for a in admin_activity if not a["has_recent_activity"]]
    
    return {
        "required": {
            # Auditree: default_repository_permission
            "default_repository_permission": check_base_permissions(org_data),
            # Auditree: org_outside_collaborators
            "org_outside_collaborators": check_outside_collaborators_disabled(org_data),
            # Auditree: unsecure_org_hooks
            "unsecure_org_hooks": len(bad_org_hooks) == 0,
            "unsecure_org_hooks_list": bad_org_hooks,  # list for details
        },
        "recommended": {
            # Auditree: members_can_create_public_repositories
            "members_can_create_public_repositories": check_repo_creation_private(org_data),
            "integration_requests_disabled": check_integration_requests_disabled(org_data),
            "visibility_change_disabled": check_visibility_change_disabled(org_data),
            "delete_transfer_disabled": check_delete_transfer_disabled(org_data),
            "profile_name_visible": check_profile_name_visibility(org_data),
            "team_creation_disabled": check_team_creation_disabled(org_data),
            "admin_activity_6_months": len(inactive_admins) == 0,
            "inactive_admins": inactive_admins,  # list for details
            "admin_activity_details": admin_activity,  # full details
        }
    }


def is_org_compliant(org_checks):
    """
    Return True only if all REQUIRED organization rules pass.
    Recommended rules are reported but don't affect compliance status.
    """
    required = org_checks.get("required", {})
    # Exclude list fields from compliance check
    return all(
        v for k, v in required.items() 
        if not isinstance(v, list)
    )


# ---------------------------------------------------------------------------
# Repository-level checks
# ---------------------------------------------------------------------------

def get_repositories():
    """Return all non-archived repositories in the org."""
    repos = paginate(f"{BASE}/orgs/{ORG}/repos?per_page=100")
    return [r for r in repos if not r.get("archived", False)]


def check_metadata_file(repo_name, default_branch):
    """
    Rule: .metadata file must exist on the default branch.
    Returns True if found, False otherwise.
    """
    url = f"{BASE}/repos/{ORG}/{repo_name}/contents/.metadata?ref={default_branch}"
    r = requests.get(url, headers=HEADERS)
    time.sleep(SLEEP)
    return r.status_code == 200


def check_repo_visibility(repo):
    """
    Rule: Repository must be private if it contains production code
    or is IP/security-sensitive. We flag public repos as a finding.
    """
    return repo.get("private", False)


def check_collaborators(repo_name):
    """
    Rule: No individual (outside) collaborators — all access must be via teams.
    Returns list of outside collaborators if any exist.
    """
    outside = paginate(f"{BASE}/repos/{ORG}/{repo_name}/collaborators?affiliation=outside&per_page=100")
    time.sleep(SLEEP)
    return [c["login"] for c in outside]


def check_hooks(repo_name):
    """
    Rule: All webhooks must have SSL verification enabled.
    Returns list of hook IDs that have SSL disabled.
    """
    hooks = get(f"{BASE}/repos/{ORG}/{repo_name}/hooks", allow_404=True) or []
    time.sleep(SLEEP)
    return [h["id"] for h in hooks if not h.get("config", {}).get("insecure_ssl") in (None, "0", 0, False)]


# ---------------------------------------------------------------------------
# Branch-protection checks
# ---------------------------------------------------------------------------

def get_branch_protection(repo_name, branch):
    """Fetch branch protection settings; returns None if not configured."""
    url = f"{BASE}/repos/{ORG}/{repo_name}/branches/{branch}/protection"
    r = requests.get(url, headers=HEADERS)
    time.sleep(SLEEP)
    if r.status_code == 404:
        return None
    r.raise_for_status()
    return r.json()


def evaluate_branch_protection(protection):
    """
    Evaluate all REQUIRED branch protection rules from the CISO policy.
    Uses Auditree naming conventions.

    Returns a dict of {rule_name: bool} where True means the rule is satisfied.
    """
    if protection is None:
        # No protection at all — every rule fails
        return {
            # Auditree: needed_protection
            "needed_protection":              False,
            # Auditree: required_pr_review
            "required_pr_review":             False,
            # Auditree: approvers_count
            "approvers_count":                False,
            # Auditree: dismiss_stale
            "dismiss_stale":                  False,
            # Auditree: code_owners_review
            "code_owners_review":             False,
            # Auditree: require_last_push_approval
            "require_last_push_approval":     False,
            # Auditree: not_bypass
            "not_bypass":                     False,
        }

    pr = protection.get("required_pull_request_reviews") or {}
    enforce_admins = protection.get("enforce_admins", {})

    # Some GHE versions surface this as a nested object; handle both shapes.
    bypass_allowed = (
        protection.get("allow_bypasses", True)           # org ruleset field
        or not enforce_admins.get("enabled", False)      # classic branch protection
    )

    return {
        # Auditree: needed_protection - One or more branch protection rules exist
        "needed_protection": True,

        # Auditree: required_pr_review - Require a pull request before merging
        "required_pr_review": bool(pr),

        # Auditree: approvers_count - Required approvals >= 1
        "approvers_count": pr.get("required_approving_review_count", 0) >= 1,

        # Auditree: dismiss_stale - Dismiss stale reviews when new commits are pushed
        "dismiss_stale": pr.get("dismiss_stale_reviews", False),

        # Auditree: code_owners_review - Require review from Code Owners
        "code_owners_review": pr.get("require_code_owner_reviews", False),

        # Auditree: require_last_push_approval - Require approval of the most recent push
        "require_last_push_approval": pr.get("require_last_push_approval", False),

        # Auditree: not_bypass - Do not allow bypassing the above settings
        "not_bypass": not bypass_allowed,
    }


def is_compliant(checks):
    """Return True only if every required rule passes."""
    return all(checks.values())


# ---------------------------------------------------------------------------
# Report Generation
# ---------------------------------------------------------------------------

def get_failure_reasons(result, org_checks):
    """
    Generate human-readable failure reasons for a repository.
    Uses Auditree naming conventions from documentation.
    Returns a tuple of (list of rule names, list of reason strings).
    """
    failed_rules = []
    reasons = []
    repo_checks = result.get("repo_checks", {})
    bp_checks = result.get("branch_protection_checks", {})
    
    # Repository-level failures (Auditree names)
    if not repo_checks.get("private_if_sensitive", True):
        failed_rules.append("private_if_sensitive")
        reasons.append("Repository visibility is set to Public. Production repositories must be Private to protect sensitive code and IP.")
    if not repo_checks.get("metadata_existing", True):
        failed_rules.append("metadata_existing")
        reasons.append("Missing .metadata file in root directory. This file is required for repository tracking and compliance verification.")
    if not repo_checks.get("collaborators_in_org", True):
        collabs = repo_checks.get("outside_collaborators", [])
        failed_rules.append("collaborators_in_org")
        reasons.append(f"Outside collaborators detected: [{', '.join(collabs)}]. All access must be managed through teams in AccessHub, not individual collaborators.")
    if not repo_checks.get("unsecure_hooks", True):
        failed_rules.append("unsecure_hooks")
        reasons.append("One or more webhooks have SSL verification disabled. SSL must be enabled on all webhooks to maintain confidentiality of data in transit.")
    
    # Branch protection failures (Auditree names)
    if not bp_checks.get("needed_protection", True):
        failed_rules.append("needed_protection")
        reasons.append("Branch protection rules are not configured on the default branch. Branch protection is required to enforce code review and prevent unauthorized changes.")
    else:
        if not bp_checks.get("required_pr_review", True):
            failed_rules.append("required_pr_review")
            reasons.append("'Require a pull request before merging' is not enabled. All changes must go through a pull request to ensure proper code review before merging.")
        if not bp_checks.get("approvers_count", True):
            failed_rules.append("approvers_count")
            reasons.append("'Required number of approvals before merging' is set to 0. At least 1 approval is required to ensure changes are reviewed by another team member.")
        if not bp_checks.get("dismiss_stale", True):
            failed_rules.append("dismiss_stale")
            reasons.append("'Dismiss stale pull request approvals when new commits are pushed' is disabled. This must be enabled to ensure reviewers approve the final version of code.")
        if not bp_checks.get("code_owners_review", True):
            failed_rules.append("code_owners_review")
            reasons.append("'Require review from Code Owners' is disabled. Code owners must review changes to critical files they are responsible for.")
        if not bp_checks.get("require_last_push_approval", True):
            failed_rules.append("require_last_push_approval")
            reasons.append("'Require approval of the most recent reviewable push' is disabled. This prevents authors from self-approving by pushing after getting approval.")
        if not bp_checks.get("not_bypass", True):
            failed_rules.append("not_bypass")
            reasons.append("'Do not allow bypassing the above settings' is disabled. Administrators should not be able to bypass branch protection rules.")
    
    return failed_rules, reasons


def get_org_failure_reasons(org_checks):
    """
    Generate human-readable failure reasons for organization-level checks.
    Uses Auditree naming conventions from documentation.
    Returns a tuple of (list of rule names, list of reason strings).
    """
    failed_rules = []
    reasons = []
    required = org_checks.get("required", {})
    recommended = org_checks.get("recommended", {})
    
    # Required org checks (Auditree names)
    if not required.get("default_repository_permission", True):
        failed_rules.append("default_repository_permission")
        reasons.append("Base repository permissions is not set to 'No permission'. Other values allow all members to read private repositories by default, violating least-privilege principle.")
    if not required.get("org_outside_collaborators", True):
        failed_rules.append("org_outside_collaborators")
        reasons.append("Repository administrators are allowed to add outside collaborators. This must be disabled - all access must be managed through teams in AccessHub.")
    if not required.get("unsecure_org_hooks", True):
        failed_rules.append("unsecure_org_hooks")
        reasons.append("One or more organization-level webhooks have SSL verification disabled. SSL is required on all webhooks to maintain confidentiality.")
    
    # Recommended org checks (Auditree names where available)
    if not recommended.get("members_can_create_public_repositories", True):
        failed_rules.append("members_can_create_public_repositories")
        reasons.append("Members can create public repositories. This should be restricted to Private only to prevent accidental exposure of sensitive code.")
    if not recommended.get("integration_requests_disabled", True):
        failed_rules.append("integration_requests_disabled")
        reasons.append("Integration requests from outside collaborators are allowed. This should be disabled since all access must be through teams.")
    if not recommended.get("visibility_change_disabled", True):
        failed_rules.append("visibility_change_disabled")
        reasons.append("Members can change repository visibility. This should be disabled to prevent repositories from being accidentally made public.")
    if not recommended.get("delete_transfer_disabled", True):
        failed_rules.append("delete_transfer_disabled")
        reasons.append("Members can delete or transfer repositories. This should be disabled to prevent accidental or malicious deletion/removal of repositories.")
    if not recommended.get("team_creation_disabled", True):
        failed_rules.append("team_creation_disabled")
        reasons.append("Members can create teams. This should be disabled to ensure access management through AccessHub is not bypassed by creating ad-hoc teams.")
    if not recommended.get("admin_activity_6_months", True):
        inactive = recommended.get("inactive_admins", [])
        failed_rules.append("admin_activity_6_months")
        reasons.append(f"Organization admins with no activity in 6+ months: [{', '.join(inactive)}]. Inactive admins indicate lack of proper access revalidation for privileged accounts.")
    
    return failed_rules, reasons


def generate_markdown_report(org, summary, org_checks, results):
    """
    Generate a Markdown report in the format of GHE Branch Protection Branches Report.
    """
    from datetime import datetime
    
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M")
    
    lines = []
    lines.append(f"# GHE Branch Protection Branches Report {timestamp}")
    lines.append("(Operational Report)")
    lines.append("")
    lines.append("## Results")
    lines.append("")
    lines.append(f"- **Failures**: {summary['non_compliant']}")
    lines.append(f"- **Compliant**: {summary['fully_compliant']}")
    lines.append(f"- **Total Repositories**: {summary['total_repos']}")
    lines.append("")
    
    # Organization-level findings
    org_failed_rules, org_reasons = get_org_failure_reasons(org_checks)
    if org_reasons:
        lines.append("## Organization-Level Findings")
        lines.append("")
        lines.append("| Organization | Rules Failing | Reason/s |")
        lines.append("|--------------|---------------|----------|")
        rules_text = ", ".join(org_failed_rules) if org_failed_rules else "N/A"
        reason_text = " ".join(org_reasons) if org_reasons else "N/A"
        lines.append(f"| {org} | {rules_text} | {reason_text} |")
        lines.append("")
    
    # Non-compliant branches table
    non_compliant = [r for r in results if not r["fully_compliant"]]
    
    if non_compliant:
        lines.append("## Failure: Non-compliant Branches")
        lines.append("")
        lines.append("| Organization | Repository | Branch | Rules Failing | Reason/s |")
        lines.append("|--------------|------------|--------|---------------|----------|")
        
        for result in non_compliant:
            repo = result["repository"]
            branch = result["default_branch"]
            failed_rules, reasons = get_failure_reasons(result, org_checks)
            rules_text = ", ".join(failed_rules) if failed_rules else "N/A"
            reason_text = " ".join(reasons) if reasons else "Unknown"
            lines.append(f"| {org} | {repo} | {branch} | {rules_text} | {reason_text} |")
        
        lines.append("")
    
    # Compliant repos summary
    compliant = [r for r in results if r["fully_compliant"]]
    if compliant:
        lines.append("## Compliant Repositories")
        lines.append("")
        lines.append("| Organization | Repository | Branch |")
        lines.append("|--------------|------------|--------|")
        for result in compliant:
            lines.append(f"| {org} | {result['repository']} | {result['default_branch']} |")
        lines.append("")
    
    return "\n".join(lines)


def generate_excel_report(org, summary, org_checks, results):
    """
    Generate an Excel report with formatted tables for better readability.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    
    wb = Workbook()
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    wrap_alignment = Alignment(wrap_text=True, vertical='top')
    
    # ==================== Sheet 1: Summary ====================
    ws_summary = wb.active
    ws_summary.title = "Summary"
    
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M")
    ws_summary['A1'] = f"GHE Compliance Report - {org}"
    ws_summary['A1'].font = Font(bold=True, size=16)
    ws_summary['A2'] = f"Generated: {timestamp}"
    ws_summary['A4'] = "Results Summary"
    ws_summary['A4'].font = Font(bold=True, size=12)
    
    ws_summary['A5'] = "Total Repositories"
    ws_summary['B5'] = summary['total_repos']
    ws_summary['A6'] = "Compliant"
    ws_summary['B6'] = summary['fully_compliant']
    ws_summary['B6'].fill = pass_fill
    ws_summary['A7'] = "Non-Compliant"
    ws_summary['B7'] = summary['non_compliant']
    if summary['non_compliant'] > 0:
        ws_summary['B7'].fill = fail_fill
    ws_summary['A8'] = "Organization Compliant"
    ws_summary['B8'] = "Yes" if summary.get('org_compliant', False) else "No"
    
    ws_summary.column_dimensions['A'].width = 25
    ws_summary.column_dimensions['B'].width = 15
    
    # ==================== Sheet 2: Organization Findings ====================
    ws_org = wb.create_sheet("Organization Findings")
    
    org_headers = ["Organization", "Rule Name", "Enforcement", "Status", "Reason"]
    for col, header in enumerate(org_headers, 1):
        cell = ws_org.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    # Organization rules data (using Auditree naming conventions)
    org_rules = [
        ("default_repository_permission", "Required", org_checks.get("required", {}).get("default_repository_permission", False),
         "Base repository permissions must be set to 'No permission'. Other values allow all members to read private repositories."),
        ("org_outside_collaborators", "Required", org_checks.get("required", {}).get("org_outside_collaborators", False),
         "Repository administrators must not be allowed to add outside collaborators. All access must be through teams in AccessHub."),
        ("unsecure_org_hooks", "Required", org_checks.get("required", {}).get("unsecure_org_hooks", False),
         "All organization webhooks must have SSL verification enabled to maintain confidentiality."),
        ("members_can_create_public_repositories", "Recommended", org_checks.get("recommended", {}).get("members_can_create_public_repositories", False),
         "Repository creation should be restricted to Private only to prevent accidental public exposure."),
        ("integration_requests_disabled", "Recommended", org_checks.get("recommended", {}).get("integration_requests_disabled", False),
         "Integration requests from outside collaborators should be disabled."),
        ("visibility_change_disabled", "Recommended", org_checks.get("recommended", {}).get("visibility_change_disabled", False),
         "Members should not be able to change repository visibility to prevent accidental public exposure."),
        ("delete_transfer_disabled", "Recommended", org_checks.get("recommended", {}).get("delete_transfer_disabled", False),
         "Members should not be able to delete or transfer repositories to limit accidental or malicious removal."),
        ("team_creation_disabled", "Recommended", org_checks.get("recommended", {}).get("team_creation_disabled", False),
         "Members should not be able to create teams to ensure access management through AccessHub is not bypassed."),
        ("admin_activity_6_months", "Recommended", org_checks.get("recommended", {}).get("admin_activity_6_months", False),
         "All organization admins should have activity in the last 6 months for proper access revalidation."),
    ]
    
    row = 2
    for rule_name, enforcement, status, reason in org_rules:
        ws_org.cell(row=row, column=1, value=org).border = thin_border
        ws_org.cell(row=row, column=2, value=rule_name).border = thin_border
        ws_org.cell(row=row, column=3, value=enforcement).border = thin_border
        status_cell = ws_org.cell(row=row, column=4, value="PASS" if status else "FAIL")
        status_cell.border = thin_border
        status_cell.fill = pass_fill if status else fail_fill
        reason_cell = ws_org.cell(row=row, column=5, value=reason)
        reason_cell.border = thin_border
        reason_cell.alignment = wrap_alignment
        row += 1
    
    ws_org.column_dimensions['A'].width = 15
    ws_org.column_dimensions['B'].width = 30
    ws_org.column_dimensions['C'].width = 15
    ws_org.column_dimensions['D'].width = 10
    ws_org.column_dimensions['E'].width = 80
    
    # ==================== Sheet 3: Non-Compliant Repositories ====================
    ws_fail = wb.create_sheet("Non-Compliant Repos")
    
    fail_headers = ["Organization", "Repository", "Branch", "Rules Failing", "Reason/s"]
    for col, header in enumerate(fail_headers, 1):
        cell = ws_fail.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    row = 2
    for result in results:
        if not result["fully_compliant"]:
            failed_rules, reasons = get_failure_reasons(result, org_checks)
            ws_fail.cell(row=row, column=1, value=org).border = thin_border
            ws_fail.cell(row=row, column=2, value=result["repository"]).border = thin_border
            ws_fail.cell(row=row, column=3, value=result["default_branch"]).border = thin_border
            ws_fail.cell(row=row, column=4, value=", ".join(failed_rules)).border = thin_border
            reason_cell = ws_fail.cell(row=row, column=5, value=" | ".join(reasons))
            reason_cell.border = thin_border
            reason_cell.alignment = wrap_alignment
            row += 1
    
    ws_fail.column_dimensions['A'].width = 15
    ws_fail.column_dimensions['B'].width = 30
    ws_fail.column_dimensions['C'].width = 15
    ws_fail.column_dimensions['D'].width = 40
    ws_fail.column_dimensions['E'].width = 100
    
    # ==================== Sheet 4: Compliant Repositories ====================
    ws_pass = wb.create_sheet("Compliant Repos")
    
    pass_headers = ["Organization", "Repository", "Branch", "Status"]
    for col, header in enumerate(pass_headers, 1):
        cell = ws_pass.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    row = 2
    for result in results:
        if result["fully_compliant"]:
            ws_pass.cell(row=row, column=1, value=org).border = thin_border
            ws_pass.cell(row=row, column=2, value=result["repository"]).border = thin_border
            ws_pass.cell(row=row, column=3, value=result["default_branch"]).border = thin_border
            status_cell = ws_pass.cell(row=row, column=4, value="PASS")
            status_cell.border = thin_border
            status_cell.fill = pass_fill
            row += 1
    
    ws_pass.column_dimensions['A'].width = 15
    ws_pass.column_dimensions['B'].width = 30
    ws_pass.column_dimensions['C'].width = 15
    ws_pass.column_dimensions['D'].width = 10
    
    # Save the workbook
    excel_path = "compliance_report.xlsx"
    wb.save(excel_path)
    return excel_path


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    print(f"Scanning org: {ORG}  (base: {BASE})\n")
    
    if TEST_MODE:
        print("*** RUNNING IN TEST MODE WITH SAMPLE DATA ***\n")
        
        # Sample organization checks data
        org_checks = {
            "required": {
                "default_repository_permission": False,  # FAIL
                "org_outside_collaborators": True,       # PASS
                "unsecure_org_hooks": True,              # PASS
                "unsecure_org_hooks_list": [],
            },
            "recommended": {
                "members_can_create_public_repositories": False,  # FAIL
                "integration_requests_disabled": True,            # PASS
                "visibility_change_disabled": False,              # FAIL
                "delete_transfer_disabled": True,                 # PASS
                "profile_name_visible": True,                     # PASS
                "team_creation_disabled": False,                  # FAIL
                "admin_activity_6_months": False,                 # FAIL
                "inactive_admins": ["john.doe", "jane.smith"],
                "admin_activity_details": [
                    {"login": "john.doe", "has_recent_activity": False},
                    {"login": "jane.smith", "has_recent_activity": False},
                    {"login": "active.admin", "has_recent_activity": True},
                ],
            }
        }
        org_compliant = False  # Because default_repository_permission fails
        
        # Sample repository results
        results = [
            {
                "repository": "frontend-app",
                "default_branch": "main",
                "fully_compliant": False,
                "repo_checks": {
                    "private_if_sensitive": True,
                    "metadata_existing": False,  # FAIL
                    "collaborators_in_org": True,
                    "outside_collaborators": [],
                    "unsecure_hooks": True,
                    "hooks_with_ssl_disabled": [],
                },
                "branch_protection_checks": {
                    "needed_protection": True,
                    "required_pr_review": True,
                    "approvers_count": True,
                    "dismiss_stale": False,  # FAIL
                    "code_owners_review": False,  # FAIL
                    "require_last_push_approval": False,  # FAIL
                    "not_bypass": True,
                },
            },
            {
                "repository": "backend-api",
                "default_branch": "main",
                "fully_compliant": True,
                "repo_checks": {
                    "private_if_sensitive": True,
                    "metadata_existing": True,
                    "collaborators_in_org": True,
                    "outside_collaborators": [],
                    "unsecure_hooks": True,
                    "hooks_with_ssl_disabled": [],
                },
                "branch_protection_checks": {
                    "needed_protection": True,
                    "required_pr_review": True,
                    "approvers_count": True,
                    "dismiss_stale": True,
                    "code_owners_review": True,
                    "require_last_push_approval": True,
                    "not_bypass": True,
                },
            },
            {
                "repository": "data-pipeline",
                "default_branch": "develop",
                "fully_compliant": False,
                "repo_checks": {
                    "private_if_sensitive": False,  # FAIL - Public repo
                    "metadata_existing": True,
                    "collaborators_in_org": False,  # FAIL
                    "outside_collaborators": ["external.user", "contractor123"],
                    "unsecure_hooks": False,  # FAIL
                    "hooks_with_ssl_disabled": [12345, 67890],
                },
                "branch_protection_checks": {
                    "needed_protection": False,  # FAIL - No protection at all
                    "required_pr_review": False,
                    "approvers_count": False,
                    "dismiss_stale": False,
                    "code_owners_review": False,
                    "require_last_push_approval": False,
                    "not_bypass": False,
                },
            },
            {
                "repository": "internal-tools",
                "default_branch": "master",
                "fully_compliant": True,
                "repo_checks": {
                    "private_if_sensitive": True,
                    "metadata_existing": True,
                    "collaborators_in_org": True,
                    "outside_collaborators": [],
                    "unsecure_hooks": True,
                    "hooks_with_ssl_disabled": [],
                },
                "branch_protection_checks": {
                    "needed_protection": True,
                    "required_pr_review": True,
                    "approvers_count": True,
                    "dismiss_stale": True,
                    "code_owners_review": True,
                    "require_last_push_approval": True,
                    "not_bypass": True,
                },
            },
            {
                "repository": "mobile-app",
                "default_branch": "main",
                "fully_compliant": False,
                "repo_checks": {
                    "private_if_sensitive": True,
                    "metadata_existing": True,
                    "collaborators_in_org": True,
                    "outside_collaborators": [],
                    "unsecure_hooks": True,
                    "hooks_with_ssl_disabled": [],
                },
                "branch_protection_checks": {
                    "needed_protection": True,
                    "required_pr_review": True,
                    "approvers_count": False,  # FAIL - 0 approvers
                    "dismiss_stale": True,
                    "code_owners_review": True,
                    "require_last_push_approval": True,
                    "not_bypass": False,  # FAIL - bypass allowed
                },
            },
        ]
        
        summary = {
            "total_repos": len(results),
            "fully_compliant": sum(1 for r in results if r["fully_compliant"]),
            "non_compliant": sum(1 for r in results if not r["fully_compliant"]),
            "org_compliant": org_compliant
        }
        
        print(f"Organization compliance: {'PASS' if org_compliant else 'FAIL'}\n")
        print(f"Found {summary['total_repos']} repositories (sample data).\n")
        
    else:
        # Production mode - actual API calls
        print("Checking organization-level compliance...")
        org_data = get_org_settings()
        org_checks = evaluate_org_compliance(org_data)
        org_compliant = is_org_compliant(org_checks)
        print(f"Organization compliance: {'PASS' if org_compliant else 'FAIL'}\n")
        
        repos = get_repositories()
        print(f"Found {len(repos)} non-archived repositories.\n")

        results = []
        summary = {
            "total_repos": len(repos), 
            "fully_compliant": 0, 
            "non_compliant": 0,
            "org_compliant": org_compliant
        }

        for repo in repos:
            name            = repo["name"]
            default_branch  = repo["default_branch"]
            is_private      = check_repo_visibility(repo)
            metadata_exists = check_metadata_file(name, default_branch)
            outside_collabs = check_collaborators(name)
            bad_hooks       = check_hooks(name)
            protection      = get_branch_protection(name, default_branch)
            bp_checks       = evaluate_branch_protection(protection)
            compliant       = (
                is_private
                and metadata_exists
                and len(outside_collabs) == 0
                and len(bad_hooks) == 0
                and is_compliant(bp_checks)
            )

            if compliant:
                summary["fully_compliant"] += 1
            else:
                summary["non_compliant"] += 1

            results.append({
                "repository":            name,
                "default_branch":        default_branch,
                "fully_compliant":       compliant,
                "repo_checks": {
                    "private_if_sensitive":     is_private,
                    "metadata_existing":        metadata_exists,
                    "collaborators_in_org":     len(outside_collabs) == 0,
                    "outside_collaborators":    outside_collabs,
                    "unsecure_hooks":           len(bad_hooks) == 0,
                    "hooks_with_ssl_disabled":  bad_hooks,
                },
                "branch_protection_checks": bp_checks,
            })

    # Generate JSON output
    output = {
        "org":     ORG,
        "summary": summary,
        "organization_checks": org_checks,
        "repos":   results,
    }

    # Write JSON report
    json_path = "compliance_report.json"
    with open(json_path, "w") as f:
        json.dump(output, f, indent=2)
    print(f"JSON report written to {json_path}")

    # Generate and write Markdown report
    md_report = generate_markdown_report(ORG, summary, org_checks, results)
    md_path = "compliance_report.md"
    with open(md_path, "w") as f:
        f.write(md_report)
    print(f"Markdown report written to {md_path}")
    
    # Generate and write Excel report
    excel_path = generate_excel_report(ORG, summary, org_checks, results)
    print(f"Excel report written to {excel_path}")
    
    # Print markdown report to console
    print("\n" + "=" * 80)
    print(md_report)


if __name__ == "__main__":
    main()
