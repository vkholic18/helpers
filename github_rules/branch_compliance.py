"""
================================================================================
BRANCH PROTECTION COMPLIANCE CHECKER
================================================================================

This script checks branch protection settings against IBM CISO policy requirements.
Only checks repos where .metadata has production_code="yes" and only checks
branches listed in production_branches array.

FILTERING LOGIC (based on .metadata file):
    - If production_code = "no" → SKIP repo
    - If production_code = "yes" but production_branches = [] → SKIP repo
    - If production_code = "yes" and production_branches = ["master"] → check only master

Can also APPLY compliant settings and ROLLBACK changes if needed.

HOW TO RUN:
    1. Set environment variables:
       - GITHUB_TOKEN: Your GitHub personal access token
       - GITHUB_ORG: Organization name to check
       - GITHUB_BASE: GitHub API base URL (e.g., https://api.github.example.com)
    
    2. Run in CHECK mode (report only - default):
       python branch_compliance.py --check
       python branch_compliance.py  # same as --check
    
    3. Run in APPLY mode (fix non-compliant settings):
       python branch_compliance.py --apply
       python branch_compliance.py --apply --dry-run  # preview changes without applying
    
    4. Run in ROLLBACK mode (revert to previous settings):
       python branch_compliance.py --rollback backup_2024-01-15_120000.json
    
    5. Output files will be generated:
       - branch_compliance_report.json
       - branch_compliance_report.md
       - branch_compliance_report.xlsx
       - backup_TIMESTAMP.json (when using --apply)

RULES CHECKED (Reference: IBM Cloud Policy 3.4.1, 3.1.1, 3.1.2):

REQUIRED RULES:
---------------
1. needed_protection
   - Setting: Branch protection rules exist
   - Required Value: Configured
   - How we check: GET /repos/{org}/{repo}/branches/{branch}/protection returns data
   - Why: Without protection, anyone with write access can push directly.

2. required_pr_review
   - Setting: Require a pull request before merging
   - Required Value: Enabled
   - How we check: protection.required_pull_request_reviews is not null
   - Why: Prevents direct pushes to production without code review.

3. approvers_count
   - Setting: Required number of approvals
   - Required Value: >= 1
   - How we check: protection.required_pull_request_reviews.required_approving_review_count >= 1
   - Why: Ensures someone other than author reviews code.

4. dismiss_stale
   - Setting: Dismiss stale pull request approvals when new commits are pushed
   - Required Value: Enabled
   - How we check: protection.required_pull_request_reviews.dismiss_stale_reviews = true
   - Why: Prevents adding unreviewed commits after approval.

5. code_owners_review
   - Setting: Require review from Code Owners
   - Required Value: Enabled
   - How we check: protection.required_pull_request_reviews.require_code_owner_reviews = true
   - Why: Limits approval to designated code owners.

6. require_last_push_approval
   - Setting: Require approval of the most recent reviewable push
   - Required Value: Enabled
   - How we check: protection.required_pull_request_reviews.require_last_push_approval = true
   - Why: Prevents self-approval by pushing after review.

7. not_bypass
   - Setting: Do not allow bypassing settings
   - Required Value: Enabled
   - How we check: protection.enforce_admins.enabled = true
   - Why: Even admins should not bypass protection on production branches.

8. codeowners_existing
   - Setting: CODEOWNERS file exists
   - Required Value: File exists in root, docs/, or .github/
   - How we check: GET /repos/{org}/{repo}/contents/CODEOWNERS (try multiple paths)
   - Why: Required for "Require review from Code Owners" to work.

RECOMMENDED RULES:
------------------
9. status_check
   - Setting: Require status checks to pass before merging
   - Recommended Value: Enabled with checks
   - How we check: protection.required_status_checks.checks.length > 0
   - Why: Automated tests improve code quality.

10. branch_uptodate
    - Setting: Require branches to be up to date before merging
    - Recommended Value: Enabled
    - How we check: protection.required_status_checks.strict = true
    - Why: Ensures PR is tested against latest base branch.

11. conversation_resolution
    - Setting: Require conversation resolution before merging
    - Recommended Value: Enabled
    - How we check: protection.required_conversation_resolution.enabled = true
    - Why: Ensures all review comments are addressed.

OPTIONAL RULES:
---------------
12. bypass_actors
    - Setting: Allowed bypass actors
    - Optional Value: Documented in .metadata if used
    - How we check: If bypass actors exist, verify documented in .metadata
    - Why: Bypass must be justified and documented.

Author: GitHub Compliance Team
Last Updated: 2024
================================================================================
"""

import os
import sys
import json
import time
import base64
import argparse
import requests
import urllib3
from datetime import datetime, timezone

# Suppress SSL warnings when using verify=False
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# Excel support
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side

# YAML support (optional)
try:
    import yaml
    YAML_AVAILABLE = True
except ImportError:
    YAML_AVAILABLE = False


# =============================================================================
# CONFIGURATION
# =============================================================================

GITHUB_TOKEN = os.environ.get("GITHUB_TOKEN")
GITHUB_ORG = os.environ.get("GITHUB_ORG")
GITHUB_BASE = os.environ.get("GITHUB_BASE", "https://api.github.com")

# API settings
SLEEP_INTERVAL = 0.3


# =============================================================================
# GITHUB API CLIENT
# =============================================================================

class GitHubAPIClient:
    """Simple GitHub API client with authentication and pagination support."""
    
    def __init__(self, base_url, token):
        self.base_url = base_url.rstrip("/")
        self.headers = {
            "Authorization": f"token {token}",
            "Accept": "application/vnd.github.v3+json"
        }
    
    def get(self, endpoint, allow_404=False):
        """Make a GET request to the GitHub API."""
        url = f"{self.base_url}{endpoint}"
        response = requests.get(url, headers=self.headers, verify=False)
        
        if response.status_code == 404 and allow_404:
            return None
        
        response.raise_for_status()
        return response.json()
    
    def paginate(self, endpoint):
        """Fetch all pages of a paginated API endpoint."""
        results = []
        url = f"{self.base_url}{endpoint}"
        
        while url:
            response = requests.get(url, headers=self.headers, verify=False)
            response.raise_for_status()
            data = response.json()
            
            if isinstance(data, list):
                results.extend(data)
            else:
                results.append(data)
            
            url = None
            link_header = response.headers.get("Link", "")
            for link in link_header.split(","):
                if 'rel="next"' in link:
                    url = link.split(";")[0].strip()[1:-1]
                    break
            
            time.sleep(SLEEP_INTERVAL)
        
        return results
    
    def put(self, endpoint, data):
        """
        Make a PUT request to the GitHub API.
        
        Used for updating branch protection settings.
        
        Args:
            endpoint: API endpoint path
            data: JSON data to send
        
        Returns:
            dict: Response JSON
        """
        url = f"{self.base_url}{endpoint}"
        response = requests.put(url, headers=self.headers, json=data, verify=False)
        response.raise_for_status()
        return response.json()
    
    def delete(self, endpoint):
        """
        Make a DELETE request to the GitHub API.
        
        Used for removing branch protection (for rollback to 'no protection' state).
        
        Args:
            endpoint: API endpoint path
        
        Returns:
            bool: True if successful
        """
        url = f"{self.base_url}{endpoint}"
        response = requests.delete(url, headers=self.headers, verify=False)
        response.raise_for_status()
        return True
    
    def post_admin(self, endpoint):
        """
        Make a POST request to enable admin enforcement.
        
        Args:
            endpoint: API endpoint path
        
        Returns:
            dict: Response JSON
        """
        url = f"{self.base_url}{endpoint}"
        response = requests.post(url, headers=self.headers, verify=False)
        response.raise_for_status()
        return response.json()


# =============================================================================
# ORGANIZATION QUALIFICATION CHECKER
# =============================================================================

class OrgQualificationChecker:
    """
    Determines if an organization requires compliance checks.
    
    QUALIFICATION RULE:
    -------------------
    An organization requires compliance checks if it contains AT LEAST ONE
    repository with:
    - production_code = yes, OR
    - ip_sensitive = yes, OR
    - security_sensitive = yes
    
    If an org qualifies:
    - ALL org-level settings must be checked
    - ALL repositories in the org must be checked (not just sensitive ones)
    - Branch protection must be checked on production_branches of production repos
    """
    
    def __init__(self, api_client, org_name):
        self.api = api_client
        self.org = org_name
        self.sensitive_repos = []
        self.all_repos = []
        self.qualified = False
    
    def fetch_metadata(self, repo_name, default_branch):
        """Fetch and parse .metadata file from repository."""
        url = f"/repos/{self.org}/{repo_name}/contents/.metadata?ref={default_branch}"
        response = self.api.get(url, allow_404=True)
        time.sleep(SLEEP_INTERVAL)
        
        if not response:
            return None
        
        try:
            content = base64.b64decode(response.get("content", "")).decode("utf-8")
            
            if YAML_AVAILABLE:
                try:
                    return yaml.safe_load(content)
                except:
                    pass
            
            try:
                return json.loads(content)
            except:
                pass
            
            return None
        except Exception:
            return None
    
    def is_repo_sensitive(self, metadata):
        """
        Check if a repository is sensitive based on its metadata.
        
        Returns True if any of:
        - production_code = yes
        - ip_sensitive = yes
        - security_sensitive = yes
        """
        if not metadata:
            return False
        
        production_code = str(metadata.get("production_code", "no")).lower() == "yes"
        ip_sensitive = str(metadata.get("ip_sensitive", "no")).lower() == "yes"
        security_sensitive = str(metadata.get("security_sensitive", "no")).lower() == "yes"
        
        return production_code or ip_sensitive or security_sensitive
    
    def check_qualification(self):
        """
        Check if the organization requires compliance checks.
        
        Scans all repositories in the org to find if any contain
        production code, IP-sensitive, or security-sensitive content.
        
        Returns:
            dict: Qualification result with details
        """
        print("\n" + "=" * 60)
        print("ORGANIZATION QUALIFICATION CHECK")
        print("=" * 60)
        print(f"\n  Scanning repositories in '{self.org}' for sensitive content...")
        
        # Fetch all repositories
        self.all_repos = self.api.paginate(f"/orgs/{self.org}/repos?per_page=100")
        time.sleep(SLEEP_INTERVAL)
        
        print(f"  Found {len(self.all_repos)} repositories")
        
        # Check each repo's metadata
        print("  Checking .metadata files for sensitive content markers...")
        
        for repo in self.all_repos:
            repo_name = repo["name"]
            default_branch = repo.get("default_branch", "master")
            
            metadata = self.fetch_metadata(repo_name, default_branch)
            
            if self.is_repo_sensitive(metadata):
                sensitivity_reasons = []
                if metadata:
                    if str(metadata.get("production_code", "no")).lower() == "yes":
                        sensitivity_reasons.append("production_code")
                    if str(metadata.get("ip_sensitive", "no")).lower() == "yes":
                        sensitivity_reasons.append("ip_sensitive")
                    if str(metadata.get("security_sensitive", "no")).lower() == "yes":
                        sensitivity_reasons.append("security_sensitive")
                
                self.sensitive_repos.append({
                    "repository": repo_name,
                    "reasons": sensitivity_reasons
                })
                print(f"    ✓ {repo_name}: SENSITIVE ({', '.join(sensitivity_reasons)})")
        
        self.qualified = len(self.sensitive_repos) > 0
        
        print("\n" + "-" * 40)
        print("QUALIFICATION RESULT")
        print("-" * 40)
        
        if self.qualified:
            print(f"  ✅ Organization QUALIFIES for compliance checks")
            print(f"     Found {len(self.sensitive_repos)} sensitive repositories:")
            for repo in self.sensitive_repos[:5]:  # Show first 5
                print(f"       - {repo['repository']}: {', '.join(repo['reasons'])}")
            if len(self.sensitive_repos) > 5:
                print(f"       ... and {len(self.sensitive_repos) - 5} more")
            print(f"\n     Branch protection will be checked on production_branches")
            print(f"     of repos with production_code=yes.")
        else:
            print(f"  ℹ️  Organization does NOT qualify for compliance checks")
            print(f"     No repositories found with production_code, ip_sensitive, or security_sensitive = yes")
            print(f"     Compliance checking is not required for this organization.")
        
        return {
            "qualified": self.qualified,
            "total_repos": len(self.all_repos),
            "sensitive_repos_count": len(self.sensitive_repos),
            "sensitive_repos": self.sensitive_repos
        }


# =============================================================================
# BRANCH PROTECTION COMPLIANCE CHECKER
# =============================================================================

class BranchComplianceChecker:
    """
    Checks branch protection settings against IBM CISO policy requirements.
    Only checks repositories with production code.
    """
    
    def __init__(self, api_client, org_name, target_repo=None):
        self.api = api_client
        self.org = org_name
        self.target_repo = target_repo
        self.results = []
    
    def get_repositories(self):
        """Fetch all non-archived repositories in the organization."""
        # If targeting a specific repo, fetch just that one
        if self.target_repo:
            print(f"  Fetching specific repository: '{self.target_repo}'...")
            repo_data = self.api.get(f"/repos/{self.org}/{self.target_repo}", allow_404=True)
            if not repo_data:
                print(f"    ERROR: Repository '{self.target_repo}' not found in {self.org}")
                return []
            if repo_data.get("archived", False):
                print(f"    ERROR: Repository '{self.target_repo}' is archived")
                return []
            print(f"    Found repository: {self.target_repo}")
            return [repo_data]
        
        # Otherwise fetch all repos
        print(f"  Fetching repositories for '{self.org}'...")
        repos = self.api.paginate(f"/orgs/{self.org}/repos?per_page=100")
        # Filter out archived repos - they can't be updated anyway
        active_repos = [r for r in repos if not r.get("archived", False)]
        print(f"    Found {len(active_repos)} active repositories")
        return active_repos
    
    def fetch_metadata(self, repo_name, default_branch):
        """
        Fetch and parse .metadata file from repository.
        
        Returns parsed metadata dict or None if not found.
        """
        url = f"/repos/{self.org}/{repo_name}/contents/.metadata?ref={default_branch}"
        response = self.api.get(url, allow_404=True)
        time.sleep(SLEEP_INTERVAL)
        
        if not response:
            return None
        
        try:
            content = base64.b64decode(response.get("content", "")).decode("utf-8")
            
            if YAML_AVAILABLE:
                try:
                    return yaml.safe_load(content)
                except:
                    pass
            
            try:
                return json.loads(content)
            except:
                pass
            
            return None
        except Exception:
            return None
    
    def is_production_repo(self, metadata):
        """
        Check if repository contains active production code.
        
        HOW WE DETERMINE PRODUCTION STATUS:
        ------------------------------------
        1. Check .metadata for production_code field
        2. If production_code = "yes", repo has production code
        3. Check production_code_end date if present
        4. If end date has passed, repo is no longer production
        5. If production_code = "no" or not set, not a production repo
        
        Returns:
            bool: True if repository contains active production code
        """
        if not metadata:
            return False
        
        has_production = str(metadata.get("production_code", "no")).lower() == "yes"
        
        if not has_production:
            return False
        
        # Check if production end date has passed
        production_end = metadata.get("production_code_end")
        if production_end:
            try:
                end_date = datetime.strptime(str(production_end), "%Y-%m-%d").replace(tzinfo=timezone.utc)
                if end_date < datetime.now(timezone.utc):
                    return False  # Production code is no longer active
            except ValueError:
                pass
        
        return True
    
    def get_all_branches(self, repo_name):
        """
        Fetch ALL branches from a repository.
        
        HOW WE FETCH ALL BRANCHES:
        --------------------------
        1. Call GET /repos/{org}/{repo}/branches?per_page=100
        2. Handle pagination to get all branches
        3. Return list of all branch names
        
        API Endpoint: GET /repos/{org}/{repo}/branches
        
        Returns:
            list: All branch names in the repository
        """
        branches_data = self.api.paginate(f"/repos/{self.org}/{repo_name}/branches?per_page=100")
        time.sleep(SLEEP_INTERVAL)
        
        branch_names = [b["name"] for b in branches_data]
        return branch_names
    
    def get_branch_protection(self, repo_name, branch):
        """
        Fetch branch protection settings.
        
        API Call: GET /repos/{org}/{repo}/branches/{branch}/protection
        
        Returns protection settings dict or None if not configured.
        
        Protection object structure:
        {
            "required_pull_request_reviews": {
                "dismiss_stale_reviews": bool,
                "require_code_owner_reviews": bool,
                "required_approving_review_count": int,
                "require_last_push_approval": bool
            },
            "required_status_checks": {
                "strict": bool,
                "checks": [{"context": str, "app_id": int}]
            },
            "enforce_admins": {
                "enabled": bool
            },
            "required_conversation_resolution": {
                "enabled": bool
            }
        }
        """
        url = f"/repos/{self.org}/{repo_name}/branches/{branch}/protection"
        protection = self.api.get(url, allow_404=True)
        time.sleep(SLEEP_INTERVAL)
        return protection
    
    # =========================================================================
    # REQUIRED RULES
    # =========================================================================
    
    def check_needed_protection(self, protection):
        """
        REQUIRED RULE: needed_protection
        
        HOW WE APPLY THIS RULE:
        -----------------------
        1. Attempt to fetch branch protection via API
        2. If API returns 404, branch protection is NOT configured
        3. If API returns data, protection IS configured
        4. Rule passes only if protection is configured
        
        API Endpoint: GET /repos/{org}/{repo}/branches/{branch}/protection
        Expected: Non-null response
        
        Why this matters:
        - Without branch protection, anyone with write access can push directly
        - Production branches MUST have protection to enforce code review
        - This is the foundation of all other branch protection rules
        """
        passed = protection is not None
        
        return {
            "rule": "needed_protection",
            "passed": passed,
            "current_value": "Configured" if passed else "Not configured",
            "expected_value": "Configured",
            "enforcement": "Required",
            "reason": (
                "Branch protection is NOT configured. Enable branch protection "
                "to enforce pull request reviews and prevent direct pushes."
                if not passed else
                "Branch protection is configured."
            )
        }
    
    def check_required_pr_review(self, protection):
        """
        REQUIRED RULE: required_pr_review
        
        HOW WE APPLY THIS RULE:
        -----------------------
        1. Access protection.required_pull_request_reviews object
        2. If null/missing, PR reviews are NOT required - FAIL
        3. If object exists, PR reviews ARE required - PASS
        
        Protection Field: required_pull_request_reviews
        Expected: Not null
        
        Why this matters:
        - Prevents direct pushes to production branches
        - All changes must go through a pull request
        - Enables code review workflow
        """
        if not protection:
            return {
                "rule": "required_pr_review",
                "passed": False,
                "current_value": "N/A (no protection)",
                "expected_value": "Enabled",
                "enforcement": "Required",
                "reason": "Cannot check - branch protection not configured."
            }
        
        pr_reviews = protection.get("required_pull_request_reviews")
        passed = pr_reviews is not None and bool(pr_reviews)
        
        return {
            "rule": "required_pr_review",
            "passed": passed,
            "current_value": "Enabled" if passed else "Disabled",
            "expected_value": "Enabled",
            "enforcement": "Required",
            "reason": (
                "'Require a pull request before merging' is NOT enabled. "
                "All changes must go through PR to ensure code review."
                if not passed else
                "Pull request reviews are required before merging."
            )
        }
    
    def check_approvers_count(self, protection):
        """
        REQUIRED RULE: approvers_count
        
        HOW WE APPLY THIS RULE:
        -----------------------
        1. Access protection.required_pull_request_reviews object
        2. Read required_approving_review_count field
        3. Must be >= 1 for someone other than author to approve
        4. If 0 or missing, self-merge is possible - FAIL
        
        Protection Field: required_pull_request_reviews.required_approving_review_count
        Expected: >= 1
        
        Why this matters:
        - Ensures at least one person reviews the code
        - Prevents self-merging without review
        - Enforces "four eyes" principle
        """
        if not protection:
            return {
                "rule": "approvers_count",
                "passed": False,
                "current_value": "N/A (no protection)",
                "expected_value": ">= 1",
                "enforcement": "Required",
                "reason": "Cannot check - branch protection not configured."
            }
        
        pr_reviews = protection.get("required_pull_request_reviews") or {}
        count = pr_reviews.get("required_approving_review_count", 0)
        passed = count >= 1
        
        return {
            "rule": "approvers_count",
            "passed": passed,
            "current_value": str(count),
            "expected_value": ">= 1",
            "enforcement": "Required",
            "reason": (
                f"Required approvers is {count}. Must be at least 1 to ensure "
                "someone other than the PR author reviews the code."
                if not passed else
                f"Requires {count} approving review(s) before merge."
            )
        }
    
    def check_dismiss_stale(self, protection):
        """
        REQUIRED RULE: dismiss_stale
        
        HOW WE APPLY THIS RULE:
        -----------------------
        1. Access protection.required_pull_request_reviews object
        2. Check dismiss_stale_reviews field
        3. Must be true to invalidate old approvals on new commits
        4. If false/missing, author can push after approval - FAIL
        
        Protection Field: required_pull_request_reviews.dismiss_stale_reviews
        Expected: true
        
        Why this matters:
        - Prevents "bait and switch" attacks
        - If author pushes new commits after approval, those commits are unreviewed
        - Dismissing stale reviews forces re-review of new changes
        """
        if not protection:
            return {
                "rule": "dismiss_stale",
                "passed": False,
                "current_value": "N/A (no protection)",
                "expected_value": "Enabled",
                "enforcement": "Required",
                "reason": "Cannot check - branch protection not configured."
            }
        
        pr_reviews = protection.get("required_pull_request_reviews") or {}
        passed = pr_reviews.get("dismiss_stale_reviews", False)
        
        return {
            "rule": "dismiss_stale",
            "passed": passed,
            "current_value": "Enabled" if passed else "Disabled",
            "expected_value": "Enabled",
            "enforcement": "Required",
            "reason": (
                "'Dismiss stale pull request approvals when new commits are pushed' "
                "is NOT enabled. This allows unreviewed code to be merged."
                if not passed else
                "Stale approvals are dismissed when new commits are pushed."
            )
        }
    
    def check_code_owners_review(self, protection):
        """
        REQUIRED RULE: code_owners_review
        
        HOW WE APPLY THIS RULE:
        -----------------------
        1. Access protection.required_pull_request_reviews object
        2. Check require_code_owner_reviews field
        3. Must be true to limit approval to CODEOWNERS
        4. If false, anyone with write access can approve - FAIL
        
        Protection Field: required_pull_request_reviews.require_code_owner_reviews
        Expected: true
        
        Why this matters:
        - Limits who can approve changes
        - CODEOWNERS file defines authoritative reviewers
        - Prevents unauthorized approvals
        """
        if not protection:
            return {
                "rule": "code_owners_review",
                "passed": False,
                "current_value": "N/A (no protection)",
                "expected_value": "Enabled",
                "enforcement": "Required",
                "reason": "Cannot check - branch protection not configured."
            }
        
        pr_reviews = protection.get("required_pull_request_reviews") or {}
        passed = pr_reviews.get("require_code_owner_reviews", False)
        
        return {
            "rule": "code_owners_review",
            "passed": passed,
            "current_value": "Enabled" if passed else "Disabled",
            "expected_value": "Enabled",
            "enforcement": "Required",
            "reason": (
                "'Require review from Code Owners' is NOT enabled. "
                "Code owners must review to ensure authorized approval."
                if not passed else
                "Code owners must approve before merge."
            )
        }
    
    def check_require_last_push_approval(self, protection):
        """
        REQUIRED RULE: require_last_push_approval
        
        HOW WE APPLY THIS RULE:
        -----------------------
        1. Access protection.required_pull_request_reviews object
        2. Check require_last_push_approval field
        3. Must be true to require approval after last push
        4. Prevents author from pushing and self-approving
        
        Protection Field: required_pull_request_reviews.require_last_push_approval
        Expected: true
        
        Why this matters:
        - Prevents self-approval by the last pusher
        - The person who pushed most recently cannot be the sole approver
        - Ensures independent review of final changes
        """
        if not protection:
            return {
                "rule": "require_last_push_approval",
                "passed": False,
                "current_value": "N/A (no protection)",
                "expected_value": "Enabled",
                "enforcement": "Required",
                "reason": "Cannot check - branch protection not configured."
            }
        
        pr_reviews = protection.get("required_pull_request_reviews") or {}
        passed = pr_reviews.get("require_last_push_approval", False)
        
        return {
            "rule": "require_last_push_approval",
            "passed": passed,
            "current_value": "Enabled" if passed else "Disabled",
            "expected_value": "Enabled",
            "enforcement": "Required",
            "reason": (
                "'Require approval of the most recent reviewable push' is NOT enabled. "
                "This prevents the last pusher from self-approving."
                if not passed else
                "Last push must be approved by someone other than pusher."
            )
        }
    
    def check_not_bypass(self, protection):
        """
        REQUIRED RULE: not_bypass
        
        HOW WE APPLY THIS RULE:
        -----------------------
        1. Access protection.enforce_admins object
        2. Check enabled field
        3. Must be true to prevent admin bypass
        4. If false, admins can push directly without protection
        
        Protection Field: enforce_admins.enabled
        Expected: true
        
        Why this matters:
        - Admins should not be able to bypass protection
        - Production code should be protected from ALL users
        - Admin bypass is a security loophole
        """
        if not protection:
            return {
                "rule": "not_bypass",
                "passed": False,
                "current_value": "N/A (no protection)",
                "expected_value": "Enabled (no bypass)",
                "enforcement": "Required",
                "reason": "Cannot check - branch protection not configured."
            }
        
        enforce_admins = protection.get("enforce_admins") or {}
        enforced = enforce_admins.get("enabled", False)
        passed = enforced
        
        return {
            "rule": "not_bypass",
            "passed": passed,
            "current_value": "No bypass" if enforced else "Bypass allowed",
            "expected_value": "No bypass",
            "enforcement": "Required",
            "reason": (
                "'Include administrators' (enforce for admins) is NOT enabled. "
                "Administrators can bypass protection and push directly."
                if not passed else
                "Administrators cannot bypass branch protection."
            )
        }
    
    def check_codeowners_existing(self, repo_name, default_branch):
        """
        REQUIRED RULE: codeowners_existing
        
        HOW WE APPLY THIS RULE:
        -----------------------
        1. Try to fetch CODEOWNERS from three locations:
           - /CODEOWNERS (root)
           - /docs/CODEOWNERS
           - /.github/CODEOWNERS
        2. If found in any location, rule passes
        3. If not found anywhere, rule fails
        
        API Endpoints tried:
        - GET /repos/{org}/{repo}/contents/CODEOWNERS?ref={branch}
        - GET /repos/{org}/{repo}/contents/docs/CODEOWNERS?ref={branch}
        - GET /repos/{org}/{repo}/contents/.github/CODEOWNERS?ref={branch}
        
        Why this matters:
        - Required for "Require review from Code Owners" to work
        - Defines who is authorized to approve changes
        - Without CODEOWNERS, code owner review requirement is meaningless
        """
        # Check all standard CODEOWNERS locations
        locations = [
            ("CODEOWNERS", f"/repos/{self.org}/{repo_name}/contents/CODEOWNERS?ref={default_branch}"),
            ("docs/CODEOWNERS", f"/repos/{self.org}/{repo_name}/contents/docs/CODEOWNERS?ref={default_branch}"),
            (".github/CODEOWNERS", f"/repos/{self.org}/{repo_name}/contents/.github/CODEOWNERS?ref={default_branch}"),
        ]
        
        found_location = None
        for loc_name, url in locations:
            result = self.api.get(url, allow_404=True)
            time.sleep(SLEEP_INTERVAL)
            if result:
                found_location = loc_name
                break
        
        passed = found_location is not None
        
        return {
            "rule": "codeowners_existing",
            "passed": passed,
            "current_value": f"Found at {found_location}" if passed else "Not found",
            "expected_value": "Exists in root, docs/, or .github/",
            "enforcement": "Required",
            "reason": (
                "CODEOWNERS file not found in repository. Create CODEOWNERS in "
                "root, docs/, or .github/ to define code owners."
                if not passed else
                f"CODEOWNERS file found at {found_location}."
            )
        }
    
    # =========================================================================
    # RECOMMENDED RULES
    # =========================================================================
    
    def check_status_check(self, protection):
        """
        RECOMMENDED RULE: status_check
        
        HOW WE APPLY THIS RULE:
        -----------------------
        1. Access protection.required_status_checks object
        2. Check if checks array has any entries
        3. At least one status check should be configured
        
        Protection Field: required_status_checks.checks
        Expected: At least 1 check configured
        
        Why this matters:
        - Status checks run automated tests/validation
        - Improves code quality and catches issues early
        - Examples: CI builds, linting, security scans
        """
        if not protection:
            return {
                "rule": "status_check",
                "passed": False,
                "current_value": "N/A (no protection)",
                "expected_value": "At least 1 check",
                "enforcement": "Recommended",
                "reason": "Cannot check - branch protection not configured."
            }
        
        status_checks = protection.get("required_status_checks") or {}
        checks = status_checks.get("checks", []) or status_checks.get("contexts", [])
        passed = len(checks) > 0
        
        return {
            "rule": "status_check",
            "passed": passed,
            "current_value": f"{len(checks)} checks configured",
            "expected_value": "At least 1 check",
            "enforcement": "Recommended",
            "status_checks": checks,
            "reason": (
                "No status checks configured. Add CI/CD checks to validate code."
                if not passed else
                f"{len(checks)} status check(s) configured."
            )
        }
    
    def check_branch_uptodate(self, protection):
        """
        RECOMMENDED RULE: branch_uptodate
        
        HOW WE APPLY THIS RULE:
        -----------------------
        1. Access protection.required_status_checks object
        2. Check strict field
        3. If true, PR must be up-to-date with base branch
        
        Protection Field: required_status_checks.strict
        Expected: true
        
        Why this matters:
        - Ensures PR is tested against latest base branch
        - Prevents "works on my branch" issues
        - Catches integration conflicts before merge
        """
        if not protection:
            return {
                "rule": "branch_uptodate",
                "passed": False,
                "current_value": "N/A (no protection)",
                "expected_value": "Enabled",
                "enforcement": "Recommended",
                "reason": "Cannot check - branch protection not configured."
            }
        
        status_checks = protection.get("required_status_checks") or {}
        passed = status_checks.get("strict", False)
        
        return {
            "rule": "branch_uptodate",
            "passed": passed,
            "current_value": "Enabled" if passed else "Disabled",
            "expected_value": "Enabled",
            "enforcement": "Recommended",
            "reason": (
                "'Require branches to be up to date before merging' is NOT enabled. "
                "PRs may be merged without testing against latest base branch."
                if not passed else
                "Branches must be up-to-date before merging."
            )
        }
    
    def check_conversation_resolution(self, protection):
        """
        RECOMMENDED RULE: conversation_resolution
        
        HOW WE APPLY THIS RULE:
        -----------------------
        1. Access protection.required_conversation_resolution object
        2. Check enabled field
        3. If true, all conversations must be resolved before merge
        
        Protection Field: required_conversation_resolution.enabled
        Expected: true
        
        Why this matters:
        - Ensures review comments are addressed
        - Prevents merging with outstanding issues
        - Improves code review quality
        """
        if not protection:
            return {
                "rule": "conversation_resolution",
                "passed": False,
                "current_value": "N/A (no protection)",
                "expected_value": "Enabled",
                "enforcement": "Recommended",
                "reason": "Cannot check - branch protection not configured."
            }
        
        conversation = protection.get("required_conversation_resolution") or {}
        passed = conversation.get("enabled", False)
        
        return {
            "rule": "conversation_resolution",
            "passed": passed,
            "current_value": "Enabled" if passed else "Disabled",
            "expected_value": "Enabled",
            "enforcement": "Recommended",
            "reason": (
                "'Require conversation resolution before merging' is NOT enabled. "
                "Review comments may be ignored when merging."
                if not passed else
                "All conversations must be resolved before merge."
            )
        }
    
    # =========================================================================
    # CHECK SINGLE BRANCH
    # =========================================================================
    
    def check_branch(self, repo_name, branch_name, default_branch):
        """
        Run all compliance checks on a single branch.
        
        Returns dict with all rule results for this branch.
        """
        protection = self.get_branch_protection(repo_name, branch_name)
        
        rules = []
        
        # Required rules
        rules.append(self.check_needed_protection(protection))
        rules.append(self.check_required_pr_review(protection))
        rules.append(self.check_approvers_count(protection))
        rules.append(self.check_dismiss_stale(protection))
        rules.append(self.check_code_owners_review(protection))
        rules.append(self.check_require_last_push_approval(protection))
        rules.append(self.check_not_bypass(protection))
        rules.append(self.check_codeowners_existing(repo_name, default_branch))
        
        # Recommended rules
        rules.append(self.check_status_check(protection))
        rules.append(self.check_branch_uptodate(protection))
        rules.append(self.check_conversation_resolution(protection))
        
        return {
            "branch": branch_name,
            "has_protection": protection is not None,
            "rules": rules
        }
    
    def check_repository(self, repo_data):
        """
        Check production branches in a repository for branch protection compliance.
        
        FILTERING BASED ON .metadata FILE:
        -----------------------------------
        1. Fetch .metadata file from repository
        2. If production_code = "no" → SKIP entire repo
        3. If production_code = "yes" but production_branches is empty → SKIP repo
        4. If production_code = "yes" and production_branches has values → check ONLY those branches
        
        Returns dict with repository info and branch results, or None if skipped.
        """
        repo_name = repo_data["name"]
        default_branch = repo_data.get("default_branch", "master")
        
        print(f"    Checking: {repo_name}")
        
        # Fetch .metadata file
        metadata = self.fetch_metadata(repo_name, default_branch)
        
        # Check production_code value
        if not metadata:
            print(f"      SKIP: No .metadata file found")
            return None
        
        production_code = str(metadata.get("production_code", "no")).lower()
        
        if production_code != "yes":
            print(f"      SKIP: production_code = '{production_code}' (not 'yes')")
            return None
        
        # Get production_branches from metadata
        production_branches = metadata.get("production_branches", [])
        
        # Ensure it's a list
        if isinstance(production_branches, str):
            production_branches = [production_branches] if production_branches.strip() else []
        
        # Skip if production_branches is empty
        if not production_branches or production_branches == [""] or production_branches == [" "]:
            print(f"      SKIP: production_code = 'yes' but production_branches is empty")
            return None
        
        print(f"      production_code: yes")
        print(f"      production_branches: {production_branches}")
        
        # Check ONLY the production branches
        branch_results = []
        for branch in production_branches:
            branch = branch.strip()
            if not branch:
                continue
            print(f"      Branch: {branch}")
            result = self.check_branch(repo_name, branch, default_branch)
            branch_results.append(result)
        
        if not branch_results:
            print(f"      SKIP: No valid production branches to check")
            return None
        
        return {
            "repository": repo_name,
            "default_branch": default_branch,
            "total_branches": len(branch_results),
            "production_branches": production_branches,
            "branches": branch_results
        }
    
    def run_all_checks(self):
        """
        Execute all branch protection compliance checks.
        
        Returns:
            list: All repository/branch check results
        """
        print("\n" + "=" * 60)
        print("BRANCH PROTECTION COMPLIANCE CHECKS")
        print("=" * 60)
        
        repos = self.get_repositories()
        
        print(f"\n  Scanning {len(repos)} repositories...")
        print(f"  (Only checking repos with production_code='yes' and production_branches defined)")
        
        skipped_count = 0
        for repo in repos:
            result = self.check_repository(repo)
            if result:
                self.results.append(result)
            else:
                skipped_count += 1
        
        total_branches = sum(r["total_branches"] for r in self.results)
        print(f"\n  Checked {total_branches} production branches across {len(self.results)} production repositories")
        print(f"  Skipped {skipped_count} repositories (no .metadata, production_code!=yes, or empty production_branches)")
        
        return self.results


# =============================================================================
# REPORT GENERATOR
# =============================================================================

class ReportGenerator:
    """
    Generates compliance reports in JSON, Markdown, and Excel formats.
    """
    
    def __init__(self, org_name, results):
        self.org = org_name
        self.results = results
        self.timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    def _calculate_summary(self):
        """Calculate summary statistics."""
        total_repos = len(self.results)
        total_branches = sum(len(r["branches"]) for r in self.results)
        
        total_rules = 0
        passed_rules = 0
        required_failed = 0
        
        for repo in self.results:
            for branch in repo["branches"]:
                for rule in branch["rules"]:
                    total_rules += 1
                    if rule["passed"]:
                        passed_rules += 1
                    elif rule["enforcement"] == "Required":
                        required_failed += 1
        
        return {
            "total_repositories": total_repos,
            "total_branches": total_branches,
            "total_rules_checked": total_rules,
            "total_passed": passed_rules,
            "total_failed": total_rules - passed_rules,
            "required_failed": required_failed
        }
    
    def generate_json_report(self, filepath="branch_compliance_report.json"):
        """Generate JSON report."""
        summary = self._calculate_summary()
        
        report = {
            "report_type": "Branch Protection Compliance",
            "organization": self.org,
            "generated_at": self.timestamp,
            "filtering_logic": {
                "description": "Only repos with production_code=yes and non-empty production_branches",
                "metadata_field_checked": "production_code",
                "branches_field": "production_branches"
            },
            "summary": summary,
            "repositories": self.results
        }
        
        with open(filepath, "w", encoding="utf-8") as f:
            json.dump(report, f, indent=2, default=str)
        
        print(f"  JSON report saved: {filepath}")
        return filepath
    
    def generate_markdown_report(self, filepath="branch_compliance_report.md"):
        """Generate Markdown report."""
        summary = self._calculate_summary()
        
        lines = [
            "# Branch Protection Compliance Report",
            "",
            f"**Organization:** {self.org}",
            f"**Generated:** {self.timestamp}",
            "",
            "## Filtering Logic",
            "",
            "This report only includes repositories where:",
            "- `.metadata` file exists",
            "- `production_code` = \"yes\"",
            "- `production_branches` array is not empty",
            "",
            "Only the branches listed in `production_branches` are checked.",
            "",
            "## Summary",
            "",
            f"- **Production Repositories Checked:** {summary['total_repositories']}",
            f"- **Production Branches Checked:** {summary['total_branches']}",
            f"- **Total Rules Checked:** {summary['total_rules_checked']}",
            f"- **Passed:** {summary['total_passed']}",
            f"- **Failed:** {summary['total_failed']}",
            f"- **Required Rules Failed:** {summary['required_failed']}",
            "",
            "## Repository Details",
            ""
        ]
        
        for repo in self.results:
            prod_branches = repo.get('production_branches', [])
            lines.append(f"### {repo['repository']}")
            lines.append("")
            lines.append(f"- **production_code:** yes")
            lines.append(f"- **production_branches:** {prod_branches}")
            lines.append(f"- **Default branch:** `{repo['default_branch']}`")
            
            for branch_result in repo["branches"]:
                branch = branch_result["branch"]
                rules = branch_result["rules"]
                passed = sum(1 for r in rules if r["passed"])
                failed = len(rules) - passed
                req_failed = sum(1 for r in rules if not r["passed"] and r["enforcement"] == "Required")
                
                status = "✅" if req_failed == 0 else "❌"
                lines.append(f"#### {status} Branch: `{branch}`")
                lines.append("")
                lines.append(f"- Protection configured: {'Yes' if branch_result['has_protection'] else 'No'}")
                lines.append(f"- Rules: {passed}/{len(rules)} passed")
                lines.append("")
                
                if failed > 0:
                    lines.append("**Failed Rules:**")
                    lines.append("")
                    lines.append("| Rule | Enforcement | Current | Expected |")
                    lines.append("|------|-------------|---------|----------|")
                    for rule in rules:
                        if not rule["passed"]:
                            lines.append(f"| {rule['rule']} | {rule['enforcement']} | {rule['current_value']} | {rule['expected_value']} |")
                    lines.append("")
        
        with open(filepath, "w", encoding="utf-8") as f:
            f.write("\n".join(lines))
        
        print(f"  Markdown report saved: {filepath}")
        return filepath
    
    def generate_excel_report(self, filepath="branch_compliance_report.xlsx"):
        """Generate Excel report."""
        wb = openpyxl.Workbook()
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        warn_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        # Summary Sheet
        ws = wb.active
        ws.title = "Summary"
        summary = self._calculate_summary()
        
        summary_data = [
            ["Branch Protection Compliance Report", ""],
            ["", ""],
            ["Organization", self.org],
            ["Generated", self.timestamp],
            ["", ""],
            ["Filtering Logic:", ""],
            ["  - Only repos with .metadata file", ""],
            ["  - production_code = 'yes'", ""],
            ["  - production_branches not empty", ""],
            ["", ""],
            ["Production Repositories", summary["total_repositories"]],
            ["Production Branches", summary["total_branches"]],
            ["Total Rules Checked", summary["total_rules_checked"]],
            ["Passed", summary["total_passed"]],
            ["Failed", summary["total_failed"]],
            ["Required Failed", summary["required_failed"]]
        ]
        
        for row_idx, row in enumerate(summary_data, 1):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 1:
                    cell.font = Font(bold=True, size=14)
        
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
        
        # Results Sheet
        ws2 = wb.create_sheet("Rule Results")
        headers = ["Repository", "Production Branches", "Branch", "Rule", "Status", "Enforcement", "Current Value", "Expected Value", "Reason"]
        
        for col, header in enumerate(headers, 1):
            cell = ws2.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        
        row_idx = 2
        for repo in self.results:
            repo_name = repo["repository"]
            prod_branches = ", ".join(repo.get("production_branches", []))
            for branch_result in repo["branches"]:
                branch_name = branch_result["branch"]
                for rule in branch_result["rules"]:
                    values = [
                        repo_name,
                        prod_branches,
                        branch_name,
                        rule["rule"],
                        "PASS" if rule["passed"] else "FAIL",
                        rule["enforcement"],
                        rule["current_value"],
                        rule["expected_value"],
                        rule["reason"]
                    ]
                    for col_idx, value in enumerate(values, 1):
                        cell = ws2.cell(row=row_idx, column=col_idx, value=value)
                        cell.border = thin_border
                        if col_idx == 5:  # Status column (shifted by 1)
                            if rule["passed"]:
                                cell.fill = pass_fill
                            elif rule["enforcement"] == "Required":
                                cell.fill = fail_fill
                            else:
                                cell.fill = warn_fill
                    row_idx += 1
        
        # Adjust widths
        ws2.column_dimensions['A'].width = 25  # Repository
        ws2.column_dimensions['B'].width = 20  # Production Branches
        ws2.column_dimensions['C'].width = 15  # Branch
        ws2.column_dimensions['D'].width = 25  # Rule
        ws2.column_dimensions['E'].width = 10  # Status
        ws2.column_dimensions['F'].width = 12  # Enforcement
        ws2.column_dimensions['G'].width = 30  # Current Value
        ws2.column_dimensions['H'].width = 25  # Expected Value
        ws2.column_dimensions['I'].width = 60  # Reason
        
        wb.save(filepath)
        print(f"  Excel report saved: {filepath}")
        return filepath
    
    def generate_all_reports(self):
        """Generate all report formats."""
        print("\nGenerating Reports...")
        self.generate_json_report()
        self.generate_markdown_report()
        self.generate_excel_report()


# =============================================================================
# BRANCH PROTECTION APPLIER
# =============================================================================

class BranchProtectionApplier:
    """
    Applies compliant branch protection settings to repositories.
    Supports backup before changes and rollback if needed.
    
    APPLY MODE:
    -----------
    1. Fetch current protection settings for all branches
    2. Save backup to backup_TIMESTAMP.json
    3. Apply compliant settings to non-compliant branches
    4. Generate report of changes made
    
    ROLLBACK MODE:
    --------------
    1. Read backup file
    2. Restore original protection settings
    3. Report what was restored
    """
    
    def __init__(self, api_client, org_name, dry_run=False):
        self.api = api_client
        self.org = org_name
        self.dry_run = dry_run
        self.changes_made = []
        self.errors = []
    
    def check_codeowners_exists(self, repo_name, default_branch="master"):
        """Check if CODEOWNERS file exists in the repository."""
        locations = [
            f"/repos/{self.org}/{repo_name}/contents/CODEOWNERS?ref={default_branch}",
            f"/repos/{self.org}/{repo_name}/contents/docs/CODEOWNERS?ref={default_branch}",
            f"/repos/{self.org}/{repo_name}/contents/.github/CODEOWNERS?ref={default_branch}",
        ]
        for url in locations:
            result = self.api.get(url, allow_404=True)
            if result:
                return True
            time.sleep(SLEEP_INTERVAL)
        return False
    
    def get_compliant_protection_payload(self, existing_protection=None, has_codeowners=True):
        """
        Build the API payload for compliant branch protection settings.
        
        Preserves existing settings where appropriate (like status checks)
        while ensuring required settings are compliant.
        
        NOTE: require_code_owner_reviews can only be True if CODEOWNERS file exists.
              GitHub API returns 422 error otherwise.
        
        API: PUT /repos/{org}/{repo}/branches/{branch}/protection
        
        Returns:
            dict: Protection settings payload for GitHub API
        """
        # Start with existing status checks if any (preserve CI/CD settings)
        existing_status_checks = None
        if existing_protection:
            existing_status_checks = existing_protection.get("required_status_checks")
        
        # Convert existing status checks to proper format if they exist
        status_checks_payload = None
        if existing_status_checks:
            checks = existing_status_checks.get("checks", []) or existing_status_checks.get("contexts", [])
            # Convert contexts (strings) to checks format if needed
            if checks and isinstance(checks[0], str):
                checks = [{"context": c} for c in checks]
            status_checks_payload = {
                "strict": existing_status_checks.get("strict", True),
                "checks": checks
            }
        
        payload = {
            # REQUIRED: Pull Request Reviews
            "required_pull_request_reviews": {
                "dismiss_stale_reviews": True,                    # Rule: dismiss_stale
                "require_code_owner_reviews": has_codeowners,     # Rule: code_owners_review (only if CODEOWNERS exists!)
                "required_approving_review_count": 1,             # Rule: approvers_count
                "require_last_push_approval": True                # Rule: require_last_push_approval
            },
            
            # REQUIRED: Enforce for administrators (no bypass)
            "enforce_admins": True,                               # Rule: not_bypass
            
            # RECOMMENDED: Conversation resolution
            "required_conversation_resolution": True,             # Rule: conversation_resolution
            
            # Preserve existing status checks (don't set if none exist to avoid issues)
            "required_status_checks": status_checks_payload,
            
            # Allow force pushes and deletions - typically should be disabled
            "allow_force_pushes": False,
            "allow_deletions": False,
            
            # Required for API - restrictions on who can push
            "restrictions": None
        }
        
        return payload
    
    def backup_current_settings(self, checker_results):
        """
        Save current branch protection settings to backup file.
        
        Backup format:
        {
            "timestamp": "2024-01-15T12:00:00",
            "organization": "org-name",
            "branches": [
                {
                    "repository": "repo-name",
                    "branch": "main",
                    "had_protection": true,
                    "protection_settings": {...}  # Raw API response or null
                }
            ]
        }
        
        Returns:
            str: Path to backup file
        """
        timestamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
        backup_file = f"backup_{timestamp}.json"
        
        backup_data = {
            "timestamp": datetime.now().isoformat(),
            "organization": self.org,
            "branches": []
        }
        
        print("\n  Creating backup of current settings...")
        
        for repo_result in checker_results:
            repo_name = repo_result["repository"]
            
            for branch_result in repo_result["branches"]:
                branch_name = branch_result["branch"]
                
                # Fetch current protection settings
                protection = self.api.get(
                    f"/repos/{self.org}/{repo_name}/branches/{branch_name}/protection",
                    allow_404=True
                )
                time.sleep(SLEEP_INTERVAL)
                
                backup_data["branches"].append({
                    "repository": repo_name,
                    "branch": branch_name,
                    "had_protection": protection is not None,
                    "protection_settings": protection
                })
        
        with open(backup_file, "w", encoding="utf-8") as f:
            json.dump(backup_data, f, indent=2, default=str)
        
        print(f"    Backup saved: {backup_file}")
        print(f"    Backed up {len(backup_data['branches'])} branch configurations")
        
        return backup_file
    
    def apply_protection(self, repo_name, branch_name, existing_protection=None, default_branch="master"):
        """
        Apply compliant branch protection settings to a single branch.
        
        API: PUT /repos/{org}/{repo}/branches/{branch}/protection
        
        Args:
            repo_name: Repository name
            branch_name: Branch name
            existing_protection: Current protection settings (to preserve some values)
            default_branch: Default branch for checking CODEOWNERS
        
        Returns:
            dict: Result with success status and details
        """
        # Check if CODEOWNERS exists (required for require_code_owner_reviews)
        has_codeowners = self.check_codeowners_exists(repo_name, default_branch)
        
        endpoint = f"/repos/{self.org}/{repo_name}/branches/{branch_name}/protection"
        payload = self.get_compliant_protection_payload(existing_protection, has_codeowners=has_codeowners)
        
        if self.dry_run:
            return {
                "success": True,
                "dry_run": True,
                "repository": repo_name,
                "branch": branch_name,
                "has_codeowners": has_codeowners,
                "action": f"Would apply compliant settings (code_owner_reviews={has_codeowners})"
            }
        
        try:
            self.api.put(endpoint, payload)
            time.sleep(SLEEP_INTERVAL)
            
            return {
                "success": True,
                "repository": repo_name,
                "branch": branch_name,
                "has_codeowners": has_codeowners,
                "action": f"Applied compliant settings (code_owner_reviews={has_codeowners})"
            }
        except requests.exceptions.HTTPError as e:
            return {
                "success": False,
                "repository": repo_name,
                "branch": branch_name,
                "has_codeowners": has_codeowners,
                "error": str(e)
            }
    
    def apply_all(self, checker_results):
        """
        Apply compliant settings to all non-compliant branches.
        
        Process:
        1. Create backup of current settings
        2. Identify branches with failed REQUIRED rules
        3. Apply compliant settings to those branches
        4. Report results
        
        Args:
            checker_results: Results from BranchComplianceChecker.run_all_checks()
        
        Returns:
            dict: Summary of changes made
        """
        print("\n" + "=" * 60)
        print("APPLYING BRANCH PROTECTION RULES")
        print("=" * 60)
        
        if self.dry_run:
            print("\n  *** DRY RUN MODE - No changes will be made ***\n")
        
        # Step 1: Create backup
        backup_file = self.backup_current_settings(checker_results)
        
        # Step 2: Find non-compliant branches
        print("\n  Identifying non-compliant branches...")
        non_compliant = []
        
        for repo_result in checker_results:
            repo_name = repo_result["repository"]
            default_branch = repo_result.get("default_branch", "master")
            
            for branch_result in repo_result["branches"]:
                branch_name = branch_result["branch"]
                
                # Check if any REQUIRED rules failed
                required_failed = any(
                    not rule["passed"] and rule["enforcement"] == "Required"
                    for rule in branch_result["rules"]
                )
                
                if required_failed:
                    non_compliant.append({
                        "repository": repo_name,
                        "branch": branch_name,
                        "default_branch": default_branch,
                        "has_protection": branch_result["has_protection"],
                        "failed_rules": [
                            rule["rule"] for rule in branch_result["rules"]
                            if not rule["passed"] and rule["enforcement"] == "Required"
                        ]
                    })
        
        print(f"    Found {len(non_compliant)} non-compliant branches")
        
        if not non_compliant:
            print("\n  All branches are compliant! No changes needed.")
            return {
                "backup_file": backup_file,
                "total_checked": sum(len(r["branches"]) for r in checker_results),
                "changes_made": 0,
                "errors": 0
            }
        
        # Step 3: Apply compliant settings
        print("\n  Applying compliant settings...")
        
        for item in non_compliant:
            repo_name = item["repository"]
            branch_name = item["branch"]
            default_branch = item.get("default_branch", "master")
            
            print(f"    {repo_name}/{branch_name}: ", end="")
            
            # Get existing protection to preserve some settings
            existing = self.api.get(
                f"/repos/{self.org}/{repo_name}/branches/{branch_name}/protection",
                allow_404=True
            )
            
            result = self.apply_protection(repo_name, branch_name, existing, default_branch)
            
            if result["success"]:
                self.changes_made.append(result)
                if self.dry_run:
                    print("Would apply")
                else:
                    print("Applied ✓")
            else:
                self.errors.append(result)
                print(f"ERROR: {result.get('error', 'Unknown error')}")
            
            time.sleep(SLEEP_INTERVAL)
        
        # Summary
        print("\n" + "-" * 40)
        print("APPLY SUMMARY")
        print("-" * 40)
        print(f"  Backup file: {backup_file}")
        print(f"  Branches processed: {len(non_compliant)}")
        print(f"  Successfully applied: {len(self.changes_made)}")
        print(f"  Errors: {len(self.errors)}")
        
        if self.dry_run:
            print("\n  *** DRY RUN - No actual changes were made ***")
            print(f"  Run without --dry-run to apply {len(self.changes_made)} changes")
        
        return {
            "backup_file": backup_file,
            "total_checked": sum(len(r["branches"]) for r in checker_results),
            "branches_processed": len(non_compliant),
            "changes_made": len(self.changes_made),
            "errors": len(self.errors),
            "dry_run": self.dry_run
        }


# =============================================================================
# ROLLBACK FUNCTIONALITY
# =============================================================================

def rollback_from_backup(api_client, backup_file):
    """
    Restore branch protection settings from a backup file.
    
    Process:
    1. Read backup file
    2. For each branch in backup:
       - If had_protection was True: restore those settings
       - If had_protection was False: remove protection
    
    Args:
        api_client: GitHubAPIClient instance
        backup_file: Path to backup JSON file
    
    Returns:
        dict: Summary of rollback results
    """
    print("\n" + "=" * 60)
    print("ROLLING BACK FROM BACKUP")
    print("=" * 60)
    
    # Read backup file
    with open(backup_file, "r", encoding="utf-8") as f:
        backup_data = json.load(f)
    
    org = backup_data["organization"]
    branches = backup_data["branches"]
    
    print(f"\n  Backup from: {backup_data['timestamp']}")
    print(f"  Organization: {org}")
    print(f"  Branches to restore: {len(branches)}")
    
    restored = 0
    removed = 0
    errors = []
    
    print("\n  Restoring settings...")
    
    for item in branches:
        repo_name = item["repository"]
        branch_name = item["branch"]
        had_protection = item["had_protection"]
        protection_settings = item["protection_settings"]
        
        print(f"    {repo_name}/{branch_name}: ", end="")
        
        try:
            if had_protection and protection_settings:
                # Restore original protection settings
                # Need to convert API response format to PUT request format
                payload = convert_protection_response_to_payload(protection_settings)
                api_client.put(
                    f"/repos/{org}/{repo_name}/branches/{branch_name}/protection",
                    payload
                )
                print("Restored ✓")
                restored += 1
            else:
                # Remove protection entirely
                api_client.delete(
                    f"/repos/{org}/{repo_name}/branches/{branch_name}/protection"
                )
                print("Removed protection ✓")
                removed += 1
            
            time.sleep(SLEEP_INTERVAL)
            
        except requests.exceptions.HTTPError as e:
            print(f"ERROR: {e}")
            errors.append({
                "repository": repo_name,
                "branch": branch_name,
                "error": str(e)
            })
    
    # Summary
    print("\n" + "-" * 40)
    print("ROLLBACK SUMMARY")
    print("-" * 40)
    print(f"  Settings restored: {restored}")
    print(f"  Protection removed: {removed}")
    print(f"  Errors: {len(errors)}")
    
    return {
        "restored": restored,
        "removed": removed,
        "errors": len(errors),
        "error_details": errors
    }


def convert_protection_response_to_payload(protection):
    """
    Convert GitHub API GET response format to PUT request format.
    
    The API response includes nested objects with URLs and metadata,
    but the PUT request expects a simpler format.
    
    Args:
        protection: Branch protection response from GET API
    
    Returns:
        dict: Payload suitable for PUT request
    """
    pr_reviews = protection.get("required_pull_request_reviews") or {}
    status_checks = protection.get("required_status_checks")
    enforce_admins = protection.get("enforce_admins") or {}
    conversation = protection.get("required_conversation_resolution") or {}
    
    payload = {
        "required_pull_request_reviews": {
            "dismiss_stale_reviews": pr_reviews.get("dismiss_stale_reviews", False),
            "require_code_owner_reviews": pr_reviews.get("require_code_owner_reviews", False),
            "required_approving_review_count": pr_reviews.get("required_approving_review_count", 0),
            "require_last_push_approval": pr_reviews.get("require_last_push_approval", False)
        } if pr_reviews else None,
        
        "enforce_admins": enforce_admins.get("enabled", False),
        
        "required_conversation_resolution": conversation.get("enabled", False),
        
        "required_status_checks": {
            "strict": status_checks.get("strict", False) if status_checks else False,
            "checks": status_checks.get("checks", []) if status_checks else []
        } if status_checks else None,
        
        "allow_force_pushes": protection.get("allow_force_pushes", {}).get("enabled", False),
        "allow_deletions": protection.get("allow_deletions", {}).get("enabled", False),
        
        "restrictions": None
    }
    
    return payload


# =============================================================================
# COMMAND LINE INTERFACE
# =============================================================================

def parse_arguments():
    """Parse command line arguments."""
    parser = argparse.ArgumentParser(
        description="GitHub Branch Protection Compliance Checker & Enforcer",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  %(prog)s --check              Check compliance (default, report only)
  %(prog)s --repo my-repo       Check only a specific repository
  %(prog)s --apply              Apply compliant settings to non-compliant branches
  %(prog)s --apply --dry-run    Preview changes without applying
  %(prog)s --repo my-repo --apply --dry-run   Test apply on one repo first
  %(prog)s --rollback backup.json  Restore settings from backup file
  %(prog)s --qualification-only   Only check if org requires compliance

Organization Qualification:
  Compliance checks only apply to organizations that contain at least
  one repository with production_code, ip_sensitive, or security_sensitive = yes.
  Branch protection is checked on production_branches of repos with production_code=yes.
        """
    )
    
    mode_group = parser.add_mutually_exclusive_group()
    mode_group.add_argument(
        "--check", "-c",
        action="store_true",
        default=True,
        help="Check compliance and generate reports (default)"
    )
    mode_group.add_argument(
        "--apply", "-a",
        action="store_true",
        help="Apply compliant settings to non-compliant branches"
    )
    mode_group.add_argument(
        "--rollback", "-r",
        metavar="BACKUP_FILE",
        help="Rollback to settings from backup file"
    )
    
    parser.add_argument(
        "--dry-run", "-n",
        action="store_true",
        help="Preview changes without applying (use with --apply)"
    )
    
    parser.add_argument(
        "--repo",
        metavar="REPO_NAME",
        help="Target a specific repository (for testing before org-wide apply)"
    )
    
    parser.add_argument(
        "--skip-qualification",
        action="store_true",
        help="Skip qualification check and run compliance checks regardless"
    )
    
    parser.add_argument(
        "--qualification-only",
        action="store_true",
        help="Only run the qualification check, don't run compliance checks"
    )
    
    return parser.parse_args()


# =============================================================================
# MAIN ENTRY POINT
# =============================================================================

def main():
    """Main entry point."""
    args = parse_arguments()
    
    print("\n" + "=" * 60)
    print("GHE BRANCH PROTECTION COMPLIANCE CHECKER")
    print("=" * 60)
    
    # Validate configuration
    if not GITHUB_TOKEN:
        print("ERROR: GITHUB_TOKEN environment variable is not set.")
        sys.exit(1)
    if not GITHUB_ORG:
        print("ERROR: GITHUB_ORG environment variable is not set.")
        sys.exit(1)
    
    print(f"\nConfiguration:")
    print(f"  Organization: {GITHUB_ORG}")
    print(f"  API Base URL: {GITHUB_BASE}")
    if args.repo:
        print(f"  Target Repo: {args.repo}")
    
    # Initialize API client
    api_client = GitHubAPIClient(GITHUB_BASE, GITHUB_TOKEN)
    
    # Handle ROLLBACK mode
    if args.rollback:
        print(f"  Mode: ROLLBACK from {args.rollback}")
        rollback_from_backup(api_client, args.rollback)
        print("\n" + "=" * 60)
        return
    
    # =========================================================================
    # STEP 1: QUALIFICATION CHECK
    # =========================================================================
    # Check if this organization requires compliance checks based on whether
    # it contains any repositories with:
    # - production_code = yes, OR
    # - ip_sensitive = yes, OR
    # - security_sensitive = yes
    #
    # Branch protection checks only apply to production_branches of repos
    # with production_code=yes.
    # =========================================================================
    
    if not args.skip_qualification:
        qual_checker = OrgQualificationChecker(api_client, GITHUB_ORG)
        qual_result = qual_checker.check_qualification()
        
        # Save qualification result
        qual_report_file = f"branch_qualification_{datetime.now().strftime('%Y-%m-%d_%H%M%S')}.json"
        with open(qual_report_file, "w", encoding="utf-8") as f:
            json.dump({
                "timestamp": datetime.now().isoformat(),
                "organization": GITHUB_ORG,
                "qualification": qual_result
            }, f, indent=2, default=str)
        print(f"\n  Qualification report saved: {qual_report_file}")
        
        if args.qualification_only:
            print("\n" + "=" * 60)
            return
        
        if not qual_result["qualified"]:
            print("\n" + "=" * 60)
            print("SKIPPING COMPLIANCE CHECKS")
            print("=" * 60)
            print("  Organization does not require compliance checks.")
            print("  Use --skip-qualification to force compliance checks anyway.")
            print("\n" + "=" * 60)
            return
    else:
        print("\n  Skipping qualification check (--skip-qualification)")
    
    # =========================================================================
    # STEP 2: COMPLIANCE CHECKS
    # =========================================================================
    
    # Handle CHECK and APPLY modes (both need to run checks first)
    if args.apply:
        print(f"  Mode: APPLY {'(DRY RUN)' if args.dry_run else ''}")
    else:
        print(f"  Mode: CHECK (report only)")
    
    # Initialize checker and run checks
    checker = BranchComplianceChecker(api_client, GITHUB_ORG, target_repo=args.repo)
    results = checker.run_all_checks()
    
    if not results:
        print("\n  No repositories with production branches found.")
        print("  (Branch protection is only checked on production_branches of repos with production_code=yes)")
        return
    
    # Generate reports
    report_gen = ReportGenerator(GITHUB_ORG, results)
    report_gen.generate_all_reports()
    
    # Print summary
    summary = report_gen._calculate_summary()
    
    print("\n" + "=" * 60)
    print("CHECK SUMMARY")
    print("=" * 60)
    print(f"  Repositories Checked: {summary['total_repositories']}")
    print(f"  Total Branches Checked: {summary['total_branches']}")
    print(f"  Rules Checked: {summary['total_rules_checked']}")
    print(f"  Passed: {summary['total_passed']}")
    print(f"  Failed: {summary['total_failed']}")
    print(f"  Required Failed: {summary['required_failed']}")
    
    if summary['required_failed'] > 0:
        print("\n  ⚠️  COMPLIANCE ISSUES DETECTED - Review required rules!")
    else:
        print("\n  ✅ All required branch protection rules passed!")
    
    # Handle APPLY mode
    if args.apply:
        if summary['required_failed'] == 0:
            print("\n  No compliance issues to fix.")
        else:
            applier = BranchProtectionApplier(api_client, GITHUB_ORG, dry_run=args.dry_run)
            apply_result = applier.apply_all(results)
            
            # Save apply results to file
            apply_log_file = f"branch_apply_log_{datetime.now().strftime('%Y-%m-%d_%H%M%S')}.json"
            with open(apply_log_file, "w", encoding="utf-8") as f:
                json.dump({
                    "timestamp": datetime.now().isoformat(),
                    "organization": GITHUB_ORG,
                    "dry_run": args.dry_run,
                    "summary": apply_result,
                    "changes": applier.changes_made,
                    "errors": applier.errors
                }, f, indent=2, default=str)
            print(f"\n  Apply log saved: {apply_log_file}")
    
    print("\n" + "=" * 60)


if __name__ == "__main__":
    main()
