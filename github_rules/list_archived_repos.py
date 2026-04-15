"""
================================================================================
ARCHIVED REPOSITORY LISTER
================================================================================

Lists all archived repositories for specified GitHub organizations.

HOW TO RUN:
    1. Set environment variables:
       - GITHUB_TOKEN: Your GitHub personal access token
       - GITHUB_BASE: GitHub API base URL (e.g., https://api.github.example.com)

    2. Run:
       python list_archived_repos.py

    3. Output files will be generated:
       - archived_repos_report.json
       - archived_repos_report.xlsx

Author: GitHub Compliance Team
Last Updated: 2026
================================================================================
"""

import os
import json
import time
import requests
import urllib3
from datetime import datetime

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side

# =============================================================================
# CONFIGURATION
# =============================================================================

GITHUB_TOKEN = os.environ.get("GITHUB_TOKEN")
GITHUB_BASE = os.environ.get("GITHUB_BASE", "https://api.github.com")

# Organizations to scan
TARGET_ORGS = ["tornado", "vmwsolution"]

SLEEP_INTERVAL = 0.3


# =============================================================================
# GITHUB API CLIENT
# =============================================================================

class GitHubAPIClient:
    def __init__(self, base_url, token):
        self.base_url = base_url.rstrip("/")
        self.headers = {
            "Authorization": f"token {token}",
            "Accept": "application/vnd.github.v3+json"
        }

    def paginate(self, endpoint):
        results = []
        url = f"{self.base_url}{endpoint}"
        while url:
            response = requests.get(url, headers=self.headers, verify=False)
            response.raise_for_status()
            data = response.json()
            if isinstance(data, list):
                results.extend(data)
            else:
                results.append(data)
            url = None
            link_header = response.headers.get("Link", "")
            for link in link_header.split(","):
                if 'rel="next"' in link:
                    url = link.split(";")[0].strip()[1:-1]
                    break
            time.sleep(SLEEP_INTERVAL)
        return results


# =============================================================================
# ARCHIVED REPO LISTER
# =============================================================================

def list_archived_repos(api_client, org_name):
    print(f"\n  Fetching repositories for '{org_name}'...")
    repos = api_client.paginate(f"/orgs/{org_name}/repos?per_page=100")
    archived = [r for r in repos if r.get("archived", False)]
    print(f"    Total repos: {len(repos)}, Archived: {len(archived)}")
    return [
        {
            "organization": org_name,
            "repository": r["name"],
            "full_name": r["full_name"],
            "archived": True,
            "default_branch": r.get("default_branch", ""),
            "updated_at": r.get("updated_at", ""),
            "html_url": r.get("html_url", "")
        }
        for r in archived
    ]


# =============================================================================
# REPORT GENERATORS
# =============================================================================

def generate_json_report(all_archived, filepath="archived_repos_report.json"):
    report = {
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "organizations": TARGET_ORGS,
        "summary": {
            org: sum(1 for r in all_archived if r["organization"] == org)
            for org in TARGET_ORGS
        },
        "total_archived": len(all_archived),
        "archived_repos": all_archived
    }
    with open(filepath, "w", encoding="utf-8") as f:
        json.dump(report, f, indent=2, default=str)
    print(f"  JSON report saved: {filepath}")


def generate_excel_report(all_archived, filepath="archived_repos_report.xlsx"):
    wb = openpyxl.Workbook()

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    archived_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Summary Sheet
    ws_summary = wb.active
    ws_summary.title = "Summary"
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    summary_data = [
        ["Archived Repository Report", ""],
        ["", ""],
        ["Generated", timestamp],
        ["", ""],
        ["Organization", "Archived Repos"],
    ]
    for org in TARGET_ORGS:
        count = sum(1 for r in all_archived if r["organization"] == org)
        summary_data.append([org, count])
    summary_data.append(["", ""])
    summary_data.append(["Total Archived", len(all_archived)])

    for row_idx, row in enumerate(summary_data, 1):
        for col_idx, value in enumerate(row, 1):
            cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                cell.font = Font(bold=True, size=14)
            if row_idx == 5 and col_idx in (1, 2):
                cell.font = Font(bold=True)

    ws_summary.column_dimensions['A'].width = 25
    ws_summary.column_dimensions['B'].width = 20

    # Per-org sheets
    for org in TARGET_ORGS:
        org_repos = [r for r in all_archived if r["organization"] == org]
        ws = wb.create_sheet(org)

        headers = ["Repository", "Full Name", "Default Branch", "Last Updated", "URL"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border

        for row_idx, r in enumerate(org_repos, 2):
            values = [
                r["repository"],
                r["full_name"],
                r["default_branch"],
                r["updated_at"],
                r["html_url"]
            ]
            for col_idx, value in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.fill = archived_fill
                cell.border = thin_border

        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 45
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 22
        ws.column_dimensions['E'].width = 60

    # All archived sheet
    ws_all = wb.create_sheet("All Archived")
    headers = ["Organization", "Repository", "Full Name", "Default Branch", "Last Updated", "URL"]
    for col, header in enumerate(headers, 1):
        cell = ws_all.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    for row_idx, r in enumerate(all_archived, 2):
        values = [
            r["organization"],
            r["repository"],
            r["full_name"],
            r["default_branch"],
            r["updated_at"],
            r["html_url"]
        ]
        for col_idx, value in enumerate(values, 1):
            cell = ws_all.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = archived_fill
            cell.border = thin_border

    ws_all.column_dimensions['A'].width = 18
    ws_all.column_dimensions['B'].width = 35
    ws_all.column_dimensions['C'].width = 45
    ws_all.column_dimensions['D'].width = 18
    ws_all.column_dimensions['E'].width = 22
    ws_all.column_dimensions['F'].width = 60

    wb.save(filepath)
    print(f"  Excel report saved: {filepath}")


# =============================================================================
# MAIN
# =============================================================================

def main():
    print("\n" + "=" * 60)
    print("ARCHIVED REPOSITORY LISTER")
    print("=" * 60)

    if not GITHUB_TOKEN:
        print("ERROR: GITHUB_TOKEN environment variable is not set.")
        return
    if not GITHUB_BASE:
        print("ERROR: GITHUB_BASE environment variable is not set.")
        return

    print(f"\n  API Base URL : {GITHUB_BASE}")
    print(f"  Organizations: {', '.join(TARGET_ORGS)}")

    api_client = GitHubAPIClient(GITHUB_BASE, GITHUB_TOKEN)

    all_archived = []
    for org in TARGET_ORGS:
        archived = list_archived_repos(api_client, org)
        all_archived.extend(archived)

    print(f"\n  Total archived repos found: {len(all_archived)}")

    print("\nGenerating Reports...")
    generate_json_report(all_archived)
    generate_excel_report(all_archived)

    print("\n" + "=" * 60)
    print("SUMMARY")
    print("=" * 60)
    for org in TARGET_ORGS:
        count = sum(1 for r in all_archived if r["organization"] == org)
        print(f"  {org}: {count} archived repo(s)")
    print(f"\n  Total: {len(all_archived)} archived repo(s)")
    print("\n" + "=" * 60)


if __name__ == "__main__":
    main()
