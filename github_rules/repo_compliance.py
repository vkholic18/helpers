"""
================================================================================
REPOSITORY-LEVEL COMPLIANCE CHECKER
================================================================================

This script checks repository-level settings against IBM CISO policy requirements.

HOW TO RUN:
    1. Set environment variables:
       - GITHUB_TOKEN: Your GitHub personal access token
       - GITHUB_ORG: Organization name to check
       - GITHUB_BASE: GitHub API base URL (e.g., https://api.github.example.com)
    
    2. Run: python repo_compliance.py
    
    3. Output files will be generated:
       - repo_compliance_report.json
       - repo_compliance_report.md
       - repo_compliance_report.xlsx

RULES CHECKED (Reference: IBM Cloud Policy 3.1.3, 3.1.4, ITSS Chapter 2):

REQUIRED RULES:
---------------
1. unsecure_hooks
   - Setting: SSL verification on repository webhooks
   - Required Value: Enable SSL verification
   - How we check: Fetch /repos/{owner}/{repo}/hooks, check config.insecure_ssl = 0/False
   - Why: Webhooks without SSL can leak sensitive data.

2. collaborators_in_org
   - Setting: No outside collaborators
   - Required Value: None
   - How we check: Fetch /repos/{owner}/{repo}/collaborators?affiliation=outside
   - Why: All access must be through AccessHub teams, not individual outside users.

3. collaborators_in_team
   - Setting: No direct collaborators (individual users)
   - Required Value: None
   - How we check: Fetch /repos/{owner}/{repo}/collaborators?affiliation=direct
   - Why: All access must be through teams, not individual org members added directly.

4. shared_repo_readers
   - Setting: Cloud_Readers team access
   - Required Value: Not granted for public/IP-sensitive/security-sensitive repos
   - How we check: Fetch /repos/{owner}/{repo}/teams, check for Cloud_Readers
                   Also check .metadata for ip_sensitive/security_sensitive flags
   - Why: Cloud_Readers is meaningless for public repos; sensitive repos shouldn't be shared.

5. metadata_existing
   - Setting: .metadata file
   - Required Value: Exists in root of default branch
   - How we check: Fetch /repos/{owner}/{repo}/contents/.metadata?ref={default_branch}
   - Why: Required for repository tracking and compliance verification.

6. private_if_sensitive
   - Setting: Repository visibility
   - Required Value: Private if contains production/IP-sensitive/security-sensitive code
   - How we check: Check repo.private against .metadata flags
   - Why: Sensitive code must not be exposed publicly.

7. archived_status
   - Setting: Repository archived state
   - Required Value: Not archived if production code is active
   - How we check: Check repo.archived against .metadata production_code flags
   - Why: Active production repos must not be archived.

Author: GitHub Compliance Team
Last Updated: 2024
================================================================================
"""

import os
import sys
import json
import time
import base64
import argparse
import requests
import urllib3
from datetime import datetime, timezone

# Disable SSL warnings for GHE with self-signed certificates
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# Excel support
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# YAML support (optional)
try:
    import yaml
    YAML_AVAILABLE = True
except ImportError:
    YAML_AVAILABLE = False


# =============================================================================
# CONFIGURATION
# =============================================================================

GITHUB_TOKEN = os.environ.get("GITHUB_TOKEN")
GITHUB_ORG = os.environ.get("GITHUB_ORG")
GITHUB_BASE = os.environ.get("GITHUB_BASE", "https://api.github.com")

# API settings
SLEEP_INTERVAL = 0.3  # Delay between API calls


# =============================================================================
# GITHUB API CLIENT
# =============================================================================

class GitHubAPIClient:
    """
    Simple GitHub API client with authentication and pagination support.
    """
    
    def __init__(self, base_url, token):
        self.base_url = base_url.rstrip("/")
        self.headers = {
            "Authorization": f"token {token}",
            "Accept": "application/vnd.github.v3+json"
        }
    
    def get(self, endpoint, allow_404=False):
        """Make a GET request to the GitHub API."""
        url = f"{self.base_url}{endpoint}"
        response = requests.get(url, headers=self.headers, verify=False)
        
        if response.status_code == 404 and allow_404:
            return None
        
        response.raise_for_status()
        return response.json()
    
    def paginate(self, endpoint):
        """Fetch all pages of a paginated API endpoint."""
        results = []
        url = f"{self.base_url}{endpoint}"
        
        while url:
            response = requests.get(url, headers=self.headers, verify=False)
            response.raise_for_status()
            data = response.json()
            
            if isinstance(data, list):
                results.extend(data)
            else:
                results.append(data)
            
            url = None
            link_header = response.headers.get("Link", "")
            for link in link_header.split(","):
                if 'rel="next"' in link:
                    url = link.split(";")[0].strip()[1:-1]
                    break
            
            time.sleep(SLEEP_INTERVAL)
        
        return results
    
    def put(self, endpoint, data):
        """
        Make a PUT request to the GitHub API.
        
        Args:
            endpoint: API endpoint path
            data: JSON data to send
        
        Returns:
            dict: Response JSON
        """
        url = f"{self.base_url}{endpoint}"
        response = requests.put(url, headers=self.headers, json=data, verify=False)
        response.raise_for_status()
        return response.json() if response.text else {}
    
    def patch(self, endpoint, data):
        """
        Make a PATCH request to the GitHub API.
        
        Used for updating repository settings.
        
        Args:
            endpoint: API endpoint path
            data: JSON data to send
        
        Returns:
            dict: Response JSON
        """
        url = f"{self.base_url}{endpoint}"
        response = requests.patch(url, headers=self.headers, json=data, verify=False)
        response.raise_for_status()
        return response.json() if response.text else {}
    
    def delete(self, endpoint):
        """
        Make a DELETE request to the GitHub API.
        
        Args:
            endpoint: API endpoint path
        
        Returns:
            bool: True if successful
        """
        url = f"{self.base_url}{endpoint}"
        response = requests.delete(url, headers=self.headers, verify=False)
        response.raise_for_status()
        return True


# =============================================================================
# REPOSITORY COMPLIANCE CHECKER
# =============================================================================

class RepoComplianceChecker:
    """
    Checks repository-level settings against IBM CISO policy requirements.
    """
    
    def __init__(self, api_client, org_name):
        self.api = api_client
        self.org = org_name
        self.results = []
    
    def get_repositories(self, include_archived=True):
        """
        Fetch all repositories in the organization.
        
        API Call: GET /orgs/{org}/repos?per_page=100
        
        We check ALL repositories including archived ones because archived
        repos might still need compliance verification.
        """
        print(f"  Fetching repositories for '{self.org}'...")
        repos = self.api.paginate(f"/orgs/{self.org}/repos?per_page=100")
        print(f"    Found {len(repos)} repositories")
        return repos
    
    def fetch_metadata(self, repo_name, default_branch):
        """
        Fetch and parse .metadata file from repository.
        
        HOW WE FETCH METADATA:
        ----------------------
        1. Call GET /repos/{org}/{repo}/contents/.metadata?ref={default_branch}
        2. Response contains base64-encoded content
        3. Decode the content
        4. Try to parse as YAML first (if pyyaml is available)
        5. Fall back to JSON parsing
        6. Return parsed dict or None if not found/invalid
        
        The .metadata file contains:
        - production_code: yes/no
        - production_branches: list of branches with production code
        - production_code_end: date when production code is no longer active
        - ip_sensitive: yes/no (contains intellectual property)
        - security_sensitive: yes/no (contains security-sensitive info)
        - public_override: true/false (exempt from private requirement)
        """
        url = f"/repos/{self.org}/{repo_name}/contents/.metadata?ref={default_branch}"
        response = self.api.get(url, allow_404=True)
        time.sleep(SLEEP_INTERVAL)
        
        if not response:
            return None
        
        try:
            # Content is base64 encoded
            content = base64.b64decode(response.get("content", "")).decode("utf-8")
            
            # Try YAML first if available (more common format for .metadata)
            if YAML_AVAILABLE:
                try:
                    return yaml.safe_load(content)
                except:
                    pass
            
            # Try JSON
            try:
                return json.loads(content)
            except:
                pass
            
            return None
        except Exception:
            return None
    
    # =========================================================================
    # REQUIRED RULES
    # =========================================================================
    
    def check_unsecure_hooks(self, repo_name):
        """
        REQUIRED RULE: unsecure_hooks
        
        HOW WE APPLY THIS RULE:
        -----------------------
        1. Fetch webhooks via GET /repos/{org}/{repo}/hooks
        2. For each webhook, examine config.insecure_ssl field
        3. insecure_ssl values:
           - "0", 0, False, None = SSL ENABLED (good)
           - "1", 1, True = SSL DISABLED (bad)
        4. Collect all hooks with SSL disabled
        5. Rule passes if no hooks have SSL disabled
        
        API Endpoint: GET /repos/{org}/{repo}/hooks
        Check Field: config.insecure_ssl
        Expected: "0" or 0 or False (meaning SSL is enabled)
        
        Why this matters:
        - Webhooks send repository data to external URLs
        - Without SSL, this data could be intercepted
        - Confidentiality of code and events would be compromised
        """
        hooks = self.api.get(f"/repos/{self.org}/{repo_name}/hooks", allow_404=True) or []
        time.sleep(SLEEP_INTERVAL)
        
        # Find hooks where SSL verification is disabled
        # config.insecure_ssl = "1" or 1 or True means SSL is DISABLED
        insecure_hooks = []
        for hook in hooks:
            config = hook.get("config", {})
            insecure_ssl = config.get("insecure_ssl")
            
            if insecure_ssl not in (None, "0", 0, False):
                insecure_hooks.append({
                    "id": hook["id"],
                    "name": hook.get("name", "unknown"),
                    "url": config.get("url", "N/A")
                })
        
        passed = len(insecure_hooks) == 0
        
        return {
            "rule": "unsecure_hooks",
            "passed": passed,
            "current_value": f"{len(insecure_hooks)} hooks with SSL disabled",
            "expected_value": "0 hooks with SSL disabled",
            "enforcement": "Required",
            "insecure_hooks": insecure_hooks,
            "reason": (
                f"Found {len(insecure_hooks)} webhook(s) without SSL: "
                f"{[h['url'] for h in insecure_hooks]}. Enable SSL verification."
                if not passed else
                "All webhooks have SSL verification enabled."
            )
        }
    
    def check_collaborators_in_org(self, repo_name):
        """
        REQUIRED RULE: collaborators_in_org
        
        HOW WE APPLY THIS RULE:
        -----------------------
        1. Fetch outside collaborators via:
           GET /repos/{org}/{repo}/collaborators?affiliation=outside
        2. "Outside collaborators" are users who have access but are NOT
           members of the organization
        3. Collect all outside collaborator logins
        4. Rule passes if zero outside collaborators exist
        
        API Endpoint: GET /repos/{org}/{repo}/collaborators?affiliation=outside
        Expected: Empty list (no outside collaborators)
        
        Why this matters:
        - Outside collaborators bypass AccessHub team management
        - Access should only be granted through properly managed teams
        - Individual outside access is harder to track and audit
        """
        outside_collabs = self.api.paginate(
            f"/repos/{self.org}/{repo_name}/collaborators?affiliation=outside&per_page=100"
        )
        time.sleep(SLEEP_INTERVAL)
        
        collab_logins = [c["login"] for c in outside_collabs]
        passed = len(collab_logins) == 0
        
        return {
            "rule": "collaborators_in_org",
            "passed": passed,
            "current_value": f"{len(collab_logins)} outside collaborators",
            "expected_value": "0 outside collaborators",
            "enforcement": "Required",
            "outside_collaborators": collab_logins,
            "reason": (
                f"Outside collaborators found: {collab_logins}. "
                "Remove these users and grant access through AccessHub teams instead."
                if not passed else
                "No outside collaborators. All access is through org membership."
            )
        }
    
    def check_collaborators_in_team(self, repo_name):
        """
        REQUIRED RULE: collaborators_in_team
        
        HOW WE APPLY THIS RULE:
        -----------------------
        1. Fetch direct collaborators via:
           GET /repos/{org}/{repo}/collaborators?affiliation=direct
        2. "Direct collaborators" are org members added individually
           (not through team membership)
        3. Collect all direct collaborator logins
        4. Rule passes if zero direct collaborators exist
        
        API Endpoint: GET /repos/{org}/{repo}/collaborators?affiliation=direct
        Expected: Empty list (no direct collaborators)
        
        Why this matters:
        - Direct collaborators bypass team-based access management
        - Team access is managed through AccessHub
        - Individual access is harder to audit and manage at scale
        """
        direct_collabs = self.api.paginate(
            f"/repos/{self.org}/{repo_name}/collaborators?affiliation=direct&per_page=100"
        )
        time.sleep(SLEEP_INTERVAL)
        
        collab_logins = [c["login"] for c in direct_collabs]
        passed = len(collab_logins) == 0
        
        return {
            "rule": "collaborators_in_team",
            "passed": passed,
            "current_value": f"{len(collab_logins)} direct collaborators",
            "expected_value": "0 direct collaborators",
            "enforcement": "Required",
            "direct_collaborators": collab_logins,
            "reason": (
                f"Direct collaborators found: {collab_logins}. "
                "Remove direct access and add users to AccessHub-managed teams."
                if not passed else
                "No direct collaborators. All access is through teams."
            )
        }
    
    def check_shared_repo_readers(self, repo_name, repo_data, metadata):
        """
        REQUIRED RULE: shared_repo_readers
        
        HOW WE APPLY THIS RULE:
        -----------------------
        1. Determine if repo should NOT have Cloud_Readers access:
           a. Is repo public? (Cloud_Readers is meaningless for public repos)
           b. Is repo IP-sensitive? (from .metadata ip_sensitive field)
           c. Is repo security-sensitive? (from .metadata security_sensitive field)
        2. Fetch teams with access via GET /repos/{org}/{repo}/teams
        3. Check if "Cloud_Readers" team is in the list
        4. Rule fails if Cloud_Readers has access AND repo is public/sensitive
        
        API Endpoint: GET /repos/{org}/{repo}/teams
        Check: Team named "Cloud_Readers" (case-insensitive)
        
        Why this matters:
        - Cloud_Readers provides org-wide read access
        - Public repos don't need it (anyone can read)
        - Sensitive repos shouldn't have broad read access
        """
        is_public = not repo_data.get("private", False)
        is_ip_sensitive = False
        is_security_sensitive = False
        
        if metadata:
            is_ip_sensitive = str(metadata.get("ip_sensitive", "no")).lower() == "yes"
            is_security_sensitive = str(metadata.get("security_sensitive", "no")).lower() == "yes"
        
        # Fetch teams with repository access
        teams = self.api.paginate(f"/repos/{self.org}/{repo_name}/teams?per_page=100")
        time.sleep(SLEEP_INTERVAL)
        
        # Check for Cloud_Readers team (case-insensitive)
        has_cloud_readers = any(
            t.get("name", "").lower() == "cloud_readers" 
            for t in teams
        )
        
        # Determine if Cloud_Readers should NOT have access
        should_not_have = is_public or is_ip_sensitive or is_security_sensitive
        passed = not (should_not_have and has_cloud_readers)
        
        reason_parts = []
        if is_public:
            reason_parts.append("public (Cloud_Readers is meaningless)")
        if is_ip_sensitive:
            reason_parts.append("IP-sensitive")
        if is_security_sensitive:
            reason_parts.append("security-sensitive")
        
        return {
            "rule": "shared_repo_readers",
            "passed": passed,
            "current_value": f"Cloud_Readers: {'Yes' if has_cloud_readers else 'No'}, "
                           f"Public: {'Yes' if is_public else 'No'}, "
                           f"IP-sensitive: {'Yes' if is_ip_sensitive else 'No'}, "
                           f"Security-sensitive: {'Yes' if is_security_sensitive else 'No'}",
            "expected_value": "No Cloud_Readers for public/sensitive repos",
            "enforcement": "Required",
            "reason": (
                f"Cloud_Readers has access but repo is {', '.join(reason_parts)}. "
                "Remove Cloud_Readers team access."
                if not passed else
                "Cloud_Readers access is appropriate for this repository."
            )
        }
    
    def check_metadata_existing(self, repo_name, default_branch):
        """
        REQUIRED RULE: metadata_existing
        
        HOW WE APPLY THIS RULE:
        -----------------------
        1. Attempt to fetch .metadata file from repository root:
           GET /repos/{org}/{repo}/contents/.metadata?ref={default_branch}
        2. If API returns 404, file doesn't exist - FAIL
        3. If API returns content, try to parse it (YAML or JSON)
        4. Rule passes if file exists and can be parsed
        
        API Endpoint: GET /repos/{org}/{repo}/contents/.metadata?ref={default_branch}
        Expected: File exists and contains valid YAML/JSON
        
        Why this matters:
        - .metadata file identifies repository classification
        - Required for automated compliance checking
        - Contains production_code, sensitivity flags, etc.
        """
        metadata = self.fetch_metadata(repo_name, default_branch)
        passed = metadata is not None
        
        return {
            "rule": "metadata_existing",
            "passed": passed,
            "current_value": "Exists" if passed else "Missing",
            "expected_value": "Exists",
            "enforcement": "Required",
            "metadata": metadata,
            "reason": (
                f"Missing .metadata file in repository root on branch '{default_branch}'. "
                "Create .metadata with production_code, ip_sensitive, security_sensitive fields."
                if not passed else
                ".metadata file exists and is valid."
            )
        }
    
    def check_private_if_sensitive(self, repo_name, repo_data, metadata):
        """
        REQUIRED RULE: private_if_sensitive
        
        HOW WE APPLY THIS RULE:
        -----------------------
        1. Get repository visibility from repo_data.private field
        2. Parse .metadata for sensitivity flags:
           - production_code: yes/no
           - ip_sensitive: yes/no  
           - security_sensitive: yes/no
           - public_override: true/false (exemption)
        3. If any sensitivity flag is "yes" AND public_override is not true:
           Repository MUST be private
        4. Rule fails if sensitive repo is public without override
        
        Why this matters:
        - Production code may contain secrets or proprietary logic
        - IP-sensitive code is company intellectual property
        - Security-sensitive code could reveal vulnerabilities
        - Public exposure of sensitive code is a major security risk
        """
        is_private = repo_data.get("private", False)
        
        if not metadata:
            return {
                "rule": "private_if_sensitive",
                "passed": False,
                "current_value": f"Private: {'Yes' if is_private else 'No'}, Metadata: Missing",
                "expected_value": "Cannot verify - metadata missing",
                "enforcement": "Required",
                "reason": "Cannot verify sensitivity - .metadata file is missing."
            }
        
        # Parse sensitivity flags from metadata
        has_production = str(metadata.get("production_code", "no")).lower() == "yes"
        is_ip_sensitive = str(metadata.get("ip_sensitive", "no")).lower() == "yes"
        is_security_sensitive = str(metadata.get("security_sensitive", "no")).lower() == "yes"
        has_override = metadata.get("public_override", False)
        
        # Check if override is truthy
        if isinstance(has_override, str):
            has_override = has_override.lower() in ("true", "yes", "1")
        
        # Repo must be private if sensitive (unless override granted)
        is_sensitive = has_production or is_ip_sensitive or is_security_sensitive
        must_be_private = is_sensitive and not has_override
        
        passed = not (must_be_private and not is_private)
        
        sensitivity_reasons = []
        if has_production:
            sensitivity_reasons.append("production_code")
        if is_ip_sensitive:
            sensitivity_reasons.append("ip_sensitive")
        if is_security_sensitive:
            sensitivity_reasons.append("security_sensitive")
        
        return {
            "rule": "private_if_sensitive",
            "passed": passed,
            "current_value": f"Private: {'Yes' if is_private else 'No'}, "
                           f"Sensitive: {', '.join(sensitivity_reasons) or 'No'}, "
                           f"Override: {'Yes' if has_override else 'No'}",
            "expected_value": "Private if sensitive (unless override)",
            "enforcement": "Required",
            "reason": (
                f"Repository is public but marked as {', '.join(sensitivity_reasons)}. "
                "Must be made private or add public_override with documented justification."
                if not passed else
                "Repository visibility is appropriate for its sensitivity level."
            )
        }
    
    def check_archived_status(self, repo_name, repo_data, metadata):
        """
        REQUIRED RULE: archived_status
        
        HOW WE APPLY THIS RULE:
        -----------------------
        1. Get archived status from repo_data.archived field
        2. Parse .metadata for production status:
           - production_code: yes/no
           - production_code_end: YYYY-MM-DD date (when production usage ends)
        3. Logic:
           - If production_code=yes AND no end date: MUST NOT be archived
           - If production_code=yes AND end date in future: MUST NOT be archived
           - If production_code=yes AND end date passed: Can be archived
           - If production_code=no: Archive status doesn't matter
        
        Why this matters:
        - Archived repos cannot receive updates
        - Active production code needs to be maintainable
        - Archiving active production code creates security risk
        """
        is_archived = repo_data.get("archived", False)
        
        if not metadata:
            # If no metadata, can't verify - but archived without metadata is suspicious
            return {
                "rule": "archived_status",
                "passed": True,  # Can't verify, don't fail
                "current_value": f"Archived: {'Yes' if is_archived else 'No'}, Metadata: Missing",
                "expected_value": "Cannot verify without metadata",
                "enforcement": "Required",
                "reason": "Cannot verify production status - .metadata file missing."
            }
        
        has_production = str(metadata.get("production_code", "no")).lower() == "yes"
        production_end = metadata.get("production_code_end")
        
        # Check if production end date has passed
        end_date_passed = False
        if production_end:
            try:
                end_date = datetime.strptime(str(production_end), "%Y-%m-%d").replace(tzinfo=timezone.utc)
                today = datetime.now(timezone.utc)
                end_date_passed = end_date < today
            except ValueError:
                pass
        
        # Repo should NOT be archived if actively used in production
        is_active_production = has_production and not end_date_passed
        passed = not (is_active_production and is_archived)
        
        return {
            "rule": "archived_status",
            "passed": passed,
            "current_value": f"Archived: {'Yes' if is_archived else 'No'}, "
                           f"Production: {'Yes' if has_production else 'No'}, "
                           f"End Date: {production_end or 'None'}, "
                           f"End Passed: {'Yes' if end_date_passed else 'No'}",
            "expected_value": "Not archived if active production code",
            "enforcement": "Required",
            "reason": (
                "Repository is archived but contains active production code. "
                "Unarchive the repository or update .metadata with production_code_end date."
                if not passed else
                "Archive status is appropriate for production code status."
            )
        }
    
    # =========================================================================
    # CHECK SINGLE REPOSITORY
    # =========================================================================
    
    def check_repository(self, repo_data):
        """
        Run all compliance checks on a single repository.
        
        Returns dict with all rule results for this repository.
        """
        repo_name = repo_data["name"]
        default_branch = repo_data.get("default_branch", "main")
        
        print(f"    Checking: {repo_name}")
        
        # Fetch metadata first (used by multiple rules)
        metadata = self.fetch_metadata(repo_name, default_branch)
        
        # Run all checks
        results = {
            "repository": repo_name,
            "default_branch": default_branch,
            "private": repo_data.get("private", False),
            "archived": repo_data.get("archived", False),
            "rules": []
        }
        
        # Required Rules
        results["rules"].append(self.check_unsecure_hooks(repo_name))
        results["rules"].append(self.check_collaborators_in_org(repo_name))
        results["rules"].append(self.check_collaborators_in_team(repo_name))
        results["rules"].append(self.check_shared_repo_readers(repo_name, repo_data, metadata))
        results["rules"].append(self.check_metadata_existing(repo_name, default_branch))
        results["rules"].append(self.check_private_if_sensitive(repo_name, repo_data, metadata))
        results["rules"].append(self.check_archived_status(repo_name, repo_data, metadata))
        
        return results
    
    def run_all_checks(self):
        """
        Execute all repository compliance checks.
        
        Returns:
            list: All repository check results
        """
        print("\n" + "=" * 60)
        print("REPOSITORY-LEVEL COMPLIANCE CHECKS")
        print("=" * 60)
        
        repos = self.get_repositories()
        
        print(f"\n  Checking {len(repos)} repositories...")
        for repo in repos:
            result = self.check_repository(repo)
            self.results.append(result)
        
        return self.results


# =============================================================================
# REPORT GENERATOR
# =============================================================================

class ReportGenerator:
    """
    Generates compliance reports in JSON, Markdown, and Excel formats.
    """
    
    def __init__(self, org_name, results):
        self.org = org_name
        self.results = results
        self.timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    def _calculate_summary(self):
        """Calculate summary statistics."""
        total_repos = len(self.results)
        total_rules = sum(len(r["rules"]) for r in self.results)
        total_passed = sum(1 for r in self.results for rule in r["rules"] if rule["passed"])
        total_failed = total_rules - total_passed
        repos_with_issues = sum(1 for r in self.results if any(not rule["passed"] for rule in r["rules"]))
        
        return {
            "total_repositories": total_repos,
            "total_rules_checked": total_rules,
            "total_passed": total_passed,
            "total_failed": total_failed,
            "repos_with_issues": repos_with_issues,
            "repos_compliant": total_repos - repos_with_issues
        }
    
    def generate_json_report(self, filepath="repo_compliance_report.json"):
        """Generate JSON report."""
        summary = self._calculate_summary()
        
        report = {
            "report_type": "Repository Compliance",
            "organization": self.org,
            "generated_at": self.timestamp,
            "summary": summary,
            "repositories": self.results
        }
        
        with open(filepath, "w", encoding="utf-8") as f:
            json.dump(report, f, indent=2, default=str)
        
        print(f"  JSON report saved: {filepath}")
        return filepath
    
    def generate_markdown_report(self, filepath="repo_compliance_report.md"):
        """Generate Markdown report."""
        summary = self._calculate_summary()
        
        lines = [
            "# Repository Compliance Report",
            "",
            f"**Organization:** {self.org}",
            f"**Generated:** {self.timestamp}",
            "",
            "## Summary",
            "",
            f"- **Total Repositories:** {summary['total_repositories']}",
            f"- **Repositories with Issues:** {summary['repos_with_issues']}",
            f"- **Compliant Repositories:** {summary['repos_compliant']}",
            f"- **Total Rules Checked:** {summary['total_rules_checked']}",
            f"- **Passed:** {summary['total_passed']}",
            f"- **Failed:** {summary['total_failed']}",
            "",
            "## Repository Details",
            ""
        ]
        
        for repo_result in self.results:
            repo_name = repo_result["repository"]
            rules = repo_result["rules"]
            passed = sum(1 for r in rules if r["passed"])
            failed = len(rules) - passed
            
            status = "✅" if failed == 0 else "❌"
            lines.append(f"### {status} {repo_name}")
            lines.append("")
            lines.append(f"- Branch: `{repo_result['default_branch']}`")
            lines.append(f"- Private: {'Yes' if repo_result['private'] else 'No'}")
            lines.append(f"- Archived: {'Yes' if repo_result['archived'] else 'No'}")
            lines.append(f"- Rules: {passed}/{len(rules)} passed")
            lines.append("")
            
            if failed > 0:
                lines.append("**Failed Rules:**")
                lines.append("")
                lines.append("| Rule | Current | Expected |")
                lines.append("|------|---------|----------|")
                for rule in rules:
                    if not rule["passed"]:
                        lines.append(f"| {rule['rule']} | {rule['current_value']} | {rule['expected_value']} |")
                lines.append("")
        
        with open(filepath, "w", encoding="utf-8") as f:
            f.write("\n".join(lines))
        
        print(f"  Markdown report saved: {filepath}")
        return filepath
    
    def generate_excel_report(self, filepath="repo_compliance_report.xlsx"):
        """Generate Excel report."""
        wb = openpyxl.Workbook()
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        # Summary Sheet
        ws = wb.active
        ws.title = "Summary"
        summary = self._calculate_summary()
        
        summary_data = [
            ["Repository Compliance Report", ""],
            ["", ""],
            ["Organization", self.org],
            ["Generated", self.timestamp],
            ["", ""],
            ["Total Repositories", summary["total_repositories"]],
            ["Repos with Issues", summary["repos_with_issues"]],
            ["Compliant Repos", summary["repos_compliant"]],
            ["Total Rules Checked", summary["total_rules_checked"]],
            ["Passed", summary["total_passed"]],
            ["Failed", summary["total_failed"]]
        ]
        
        for row_idx, row in enumerate(summary_data, 1):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 1:
                    cell.font = Font(bold=True, size=14)
        
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
        
        # Results Sheet
        ws2 = wb.create_sheet("Rule Results")
        headers = ["Repository", "Rule", "Status", "Enforcement", "Current Value", "Expected Value", "Reason"]
        
        for col, header in enumerate(headers, 1):
            cell = ws2.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        
        row_idx = 2
        for repo_result in self.results:
            repo_name = repo_result["repository"]
            for rule in repo_result["rules"]:
                values = [
                    repo_name,
                    rule["rule"],
                    "PASS" if rule["passed"] else "FAIL",
                    rule["enforcement"],
                    rule["current_value"],
                    rule["expected_value"],
                    rule["reason"]
                ]
                for col_idx, value in enumerate(values, 1):
                    cell = ws2.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = thin_border
                    if col_idx == 3:  # Status column
                        cell.fill = pass_fill if rule["passed"] else fail_fill
                row_idx += 1
        
        # Adjust widths
        ws2.column_dimensions['A'].width = 30
        ws2.column_dimensions['B'].width = 25
        ws2.column_dimensions['C'].width = 10
        ws2.column_dimensions['D'].width = 12
        ws2.column_dimensions['E'].width = 45
        ws2.column_dimensions['F'].width = 30
        ws2.column_dimensions['G'].width = 60
        
        wb.save(filepath)
        print(f"  Excel report saved: {filepath}")
        return filepath
    
    def generate_all_reports(self):
        """Generate all report formats."""
        print("\nGenerating Reports...")
        self.generate_json_report()
        self.generate_markdown_report()
        self.generate_excel_report()


# =============================================================================
# REPOSITORY COMPLIANCE APPLIER
# =============================================================================

class RepoComplianceApplier:
    """
    Applies compliant repository settings.
    Supports backup before changes and rollback if needed.
    
    WHAT CAN BE APPLIED AUTOMATICALLY:
    ----------------------------------
    1. unsecure_hooks: Enable SSL verification on webhooks
       API: PATCH /repos/{org}/{repo}/hooks/{hook_id} with {"config": {"insecure_ssl": "0"}}
       
    2. collaborators_in_org: Remove outside collaborators
       API: DELETE /repos/{org}/{repo}/collaborators/{username}
       
    3. collaborators_in_team: Remove direct collaborators (move to teams)
       API: DELETE /repos/{org}/{repo}/collaborators/{username}
       Note: User should be added to appropriate teams via AccessHub
       
    4. shared_repo_readers: Remove Cloud_Readers team access
       API: DELETE /orgs/{org}/teams/{team_slug}/repos/{org}/{repo}
       
    5. private_if_sensitive: Make repository private
       API: PATCH /repos/{org}/{repo} with {"private": true}
       
    6. archived_status: Unarchive repository
       API: PATCH /repos/{org}/{repo} with {"archived": false}
    
    WHAT CANNOT BE APPLIED AUTOMATICALLY:
    -------------------------------------
    - metadata_existing: Requires creating .metadata file (content unknown)
    """
    
    def __init__(self, api_client, org_name, dry_run=False):
        self.api = api_client
        self.org = org_name
        self.dry_run = dry_run
        self.changes_made = []
        self.errors = []
        self.skipped = []
    
    def backup_current_settings(self, check_results):
        """
        Save current repository settings to backup file.
        
        Args:
            check_results: Results from compliance check
        
        Returns:
            str: Path to backup file
        """
        timestamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
        backup_file = f"repo_backup_{timestamp}.json"
        
        backup_data = {
            "timestamp": datetime.now().isoformat(),
            "organization": self.org,
            "repositories": []
        }
        
        print("\n  Creating backup of current settings...")
        
        for repo_result in check_results:
            repo_name = repo_result["repository"]
            
            # Fetch current repo settings
            repo_data = self.api.get(f"/repos/{self.org}/{repo_name}", allow_404=True)
            time.sleep(SLEEP_INTERVAL)
            
            # Fetch hooks
            hooks = self.api.get(f"/repos/{self.org}/{repo_name}/hooks", allow_404=True) or []
            time.sleep(SLEEP_INTERVAL)
            
            # Fetch collaborators
            outside_collabs = self.api.paginate(
                f"/repos/{self.org}/{repo_name}/collaborators?affiliation=outside&per_page=100"
            )
            direct_collabs = self.api.paginate(
                f"/repos/{self.org}/{repo_name}/collaborators?affiliation=direct&per_page=100"
            )
            
            # Fetch teams
            teams = self.api.paginate(f"/repos/{self.org}/{repo_name}/teams?per_page=100")
            
            backup_data["repositories"].append({
                "repository": repo_name,
                "private": repo_data.get("private") if repo_data else None,
                "archived": repo_data.get("archived") if repo_data else None,
                "hooks": hooks,
                "outside_collaborators": [c["login"] for c in outside_collabs],
                "direct_collaborators": [c["login"] for c in direct_collabs],
                "teams": [{"name": t["name"], "slug": t["slug"], "permission": t.get("permission")} for t in teams],
                "rules_checked": repo_result["rules"]
            })
        
        with open(backup_file, "w", encoding="utf-8") as f:
            json.dump(backup_data, f, indent=2, default=str)
        
        print(f"    Backup saved: {backup_file}")
        print(f"    Backed up {len(backup_data['repositories'])} repository configurations")
        
        return backup_file
    
    def fix_unsecure_hooks(self, repo_name, insecure_hooks):
        """
        Enable SSL verification on insecure webhooks.
        
        Args:
            repo_name: Repository name
            insecure_hooks: List of hook info dicts with id
        
        Returns:
            list: Results for each hook
        """
        results = []
        
        for hook in insecure_hooks:
            hook_id = hook.get("id")
            if not hook_id:
                continue
            
            if self.dry_run:
                results.append({
                    "success": True,
                    "dry_run": True,
                    "hook_id": hook_id,
                    "action": f"Would enable SSL verification on hook {hook_id}"
                })
                continue
            
            try:
                # Enable SSL verification
                self.api.patch(
                    f"/repos/{self.org}/{repo_name}/hooks/{hook_id}",
                    {"config": {"insecure_ssl": "0"}}
                )
                time.sleep(SLEEP_INTERVAL)
                results.append({
                    "success": True,
                    "hook_id": hook_id,
                    "action": f"Enabled SSL verification on hook {hook_id}"
                })
            except requests.exceptions.HTTPError as e:
                results.append({
                    "success": False,
                    "hook_id": hook_id,
                    "error": str(e)
                })
        
        return results
    
    def remove_collaborator(self, repo_name, username, collab_type):
        """
        Remove a collaborator from repository.
        
        Args:
            repo_name: Repository name
            username: Collaborator username
            collab_type: "outside" or "direct"
        
        Returns:
            dict: Result
        """
        if self.dry_run:
            return {
                "success": True,
                "dry_run": True,
                "username": username,
                "type": collab_type,
                "action": f"Would remove {collab_type} collaborator: {username}"
            }
        
        try:
            self.api.delete(f"/repos/{self.org}/{repo_name}/collaborators/{username}")
            time.sleep(SLEEP_INTERVAL)
            return {
                "success": True,
                "username": username,
                "type": collab_type,
                "action": f"Removed {collab_type} collaborator: {username}"
            }
        except requests.exceptions.HTTPError as e:
            return {
                "success": False,
                "username": username,
                "type": collab_type,
                "error": str(e)
            }
    
    def remove_team_access(self, repo_name, team_slug):
        """
        Remove team access from repository.
        
        Args:
            repo_name: Repository name
            team_slug: Team slug
        
        Returns:
            dict: Result
        """
        if self.dry_run:
            return {
                "success": True,
                "dry_run": True,
                "team": team_slug,
                "action": f"Would remove team access: {team_slug}"
            }
        
        try:
            self.api.delete(f"/orgs/{self.org}/teams/{team_slug}/repos/{self.org}/{repo_name}")
            time.sleep(SLEEP_INTERVAL)
            return {
                "success": True,
                "team": team_slug,
                "action": f"Removed team access: {team_slug}"
            }
        except requests.exceptions.HTTPError as e:
            return {
                "success": False,
                "team": team_slug,
                "error": str(e)
            }
    
    def make_private(self, repo_name):
        """
        Make repository private.
        
        Args:
            repo_name: Repository name
        
        Returns:
            dict: Result
        """
        if self.dry_run:
            return {
                "success": True,
                "dry_run": True,
                "action": "Would make repository private"
            }
        
        try:
            self.api.patch(f"/repos/{self.org}/{repo_name}", {"private": True})
            time.sleep(SLEEP_INTERVAL)
            return {
                "success": True,
                "action": "Made repository private"
            }
        except requests.exceptions.HTTPError as e:
            return {
                "success": False,
                "error": str(e)
            }
    
    def unarchive(self, repo_name):
        """
        Unarchive repository.
        
        Args:
            repo_name: Repository name
        
        Returns:
            dict: Result
        """
        if self.dry_run:
            return {
                "success": True,
                "dry_run": True,
                "action": "Would unarchive repository"
            }
        
        try:
            self.api.patch(f"/repos/{self.org}/{repo_name}", {"archived": False})
            time.sleep(SLEEP_INTERVAL)
            return {
                "success": True,
                "action": "Unarchived repository"
            }
        except requests.exceptions.HTTPError as e:
            return {
                "success": False,
                "error": str(e)
            }
    
    def apply_repo_fixes(self, repo_result):
        """
        Apply fixes for a single repository.
        
        Args:
            repo_result: Repository check results
        
        Returns:
            dict: Summary of changes for this repo
        """
        repo_name = repo_result["repository"]
        repo_changes = {
            "repository": repo_name,
            "changes": [],
            "errors": [],
            "skipped": []
        }
        
        failed_rules = [r for r in repo_result["rules"] if not r["passed"]]
        
        for rule in failed_rules:
            rule_name = rule["rule"]
            
            if rule_name == "unsecure_hooks":
                insecure_hooks = rule.get("insecure_hooks", [])
                if insecure_hooks:
                    results = self.fix_unsecure_hooks(repo_name, insecure_hooks)
                    for r in results:
                        if r.get("success"):
                            repo_changes["changes"].append(r)
                        else:
                            repo_changes["errors"].append(r)
            
            elif rule_name == "collaborators_in_org":
                outside_collabs = rule.get("outside_collaborators", [])
                for username in outside_collabs:
                    result = self.remove_collaborator(repo_name, username, "outside")
                    if result.get("success"):
                        repo_changes["changes"].append(result)
                    else:
                        repo_changes["errors"].append(result)
            
            elif rule_name == "collaborators_in_team":
                direct_collabs = rule.get("direct_collaborators", [])
                for username in direct_collabs:
                    result = self.remove_collaborator(repo_name, username, "direct")
                    if result.get("success"):
                        repo_changes["changes"].append(result)
                    else:
                        repo_changes["errors"].append(result)
            
            elif rule_name == "shared_repo_readers":
                # Check if Cloud_Readers should be removed
                current_val = rule.get("current_value", "")
                if "Cloud_Readers: Yes" in current_val:
                    result = self.remove_team_access(repo_name, "cloud_readers")
                    if result.get("success"):
                        repo_changes["changes"].append(result)
                    else:
                        repo_changes["errors"].append(result)
            
            elif rule_name == "private_if_sensitive":
                # Check if repo should be made private
                current_val = rule.get("current_value", "")
                if "Private: No" in current_val:
                    result = self.make_private(repo_name)
                    result["rule"] = rule_name
                    if result.get("success"):
                        repo_changes["changes"].append(result)
                    else:
                        repo_changes["errors"].append(result)
            
            elif rule_name == "archived_status":
                # Check if repo needs to be unarchived
                current_val = rule.get("current_value", "")
                if "Archived: Yes" in current_val:
                    result = self.unarchive(repo_name)
                    result["rule"] = rule_name
                    if result.get("success"):
                        repo_changes["changes"].append(result)
                    else:
                        repo_changes["errors"].append(result)
            
            elif rule_name == "metadata_existing":
                repo_changes["skipped"].append({
                    "rule": rule_name,
                    "reason": "Cannot create .metadata automatically - content must be defined by repository owner"
                })
            
            else:
                repo_changes["skipped"].append({
                    "rule": rule_name,
                    "reason": f"Unknown rule: {rule_name}"
                })
        
        return repo_changes
    
    def apply_all(self, check_results, target_repo=None):
        """
        Apply compliant settings to all non-compliant repositories.
        
        Args:
            check_results: Results from RepoComplianceChecker
            target_repo: Optional - only apply to this repository
        
        Returns:
            dict: Summary of changes made
        """
        print("\n" + "=" * 60)
        print("APPLYING REPOSITORY COMPLIANCE RULES")
        print("=" * 60)
        
        if self.dry_run:
            print("\n  *** DRY RUN MODE - No changes will be made ***\n")
        
        # Filter to target repo if specified
        if target_repo:
            check_results = [r for r in check_results if r["repository"] == target_repo]
            if not check_results:
                print(f"\n  Repository '{target_repo}' not found in results.")
                return {"changes_made": 0, "errors": 0, "skipped": 0}
        
        # Step 1: Create backup
        backup_file = self.backup_current_settings(check_results)
        
        # Step 2: Find repositories with non-compliant rules
        print("\n  Identifying non-compliant repositories...")
        non_compliant = [r for r in check_results if any(not rule["passed"] for rule in r["rules"])]
        
        if not non_compliant:
            print("    All repositories are compliant! No changes needed.")
            return {
                "backup_file": backup_file,
                "changes_made": 0,
                "errors": 0,
                "skipped": 0
            }
        
        print(f"    Found {len(non_compliant)} non-compliant repositories")
        
        # Step 3: Apply fixes
        print("\n  Applying fixes...")
        
        all_repo_changes = []
        
        for repo_result in non_compliant:
            repo_name = repo_result["repository"]
            print(f"\n    {repo_name}:")
            
            repo_changes = self.apply_repo_fixes(repo_result)
            all_repo_changes.append(repo_changes)
            
            for change in repo_changes["changes"]:
                action = change.get("action", "Applied")
                if self.dry_run:
                    print(f"      - {action}")
                else:
                    print(f"      - {action} ✓")
                self.changes_made.append(change)
            
            for error in repo_changes["errors"]:
                print(f"      - ERROR: {error.get('error', 'Unknown error')}")
                self.errors.append(error)
            
            for skip in repo_changes["skipped"]:
                print(f"      - SKIPPED: {skip['rule']} ({skip['reason']})")
                self.skipped.append(skip)
        
        # Summary
        print("\n" + "-" * 40)
        print("APPLY SUMMARY")
        print("-" * 40)
        print(f"  Backup file: {backup_file}")
        print(f"  Repositories processed: {len(non_compliant)}")
        print(f"  Changes made: {len(self.changes_made)}")
        print(f"  Skipped (requires manual action): {len(self.skipped)}")
        print(f"  Errors: {len(self.errors)}")
        
        if self.dry_run:
            print("\n  *** DRY RUN - No actual changes were made ***")
            print(f"  Run without --dry-run to apply {len(self.changes_made)} changes")
        
        return {
            "backup_file": backup_file,
            "repositories_processed": len(non_compliant),
            "changes_made": len(self.changes_made),
            "skipped": len(self.skipped),
            "errors": len(self.errors),
            "dry_run": self.dry_run
        }


# =============================================================================
# ROLLBACK FUNCTIONALITY
# =============================================================================

def rollback_from_backup(api_client, backup_file, org_name):
    """
    Restore repository settings from a backup file.
    
    Note: This can restore:
    - Repository visibility (private/public)
    - Archived status
    - Webhook SSL settings
    - (Re-adding collaborators and teams would require storing their permissions)
    
    Args:
        api_client: GitHubAPIClient instance
        backup_file: Path to backup JSON file
        org_name: Organization name
    
    Returns:
        dict: Summary of rollback results
    """
    print("\n" + "=" * 60)
    print("ROLLING BACK FROM BACKUP")
    print("=" * 60)
    
    # Read backup file
    with open(backup_file, "r", encoding="utf-8") as f:
        backup_data = json.load(f)
    
    org = backup_data["organization"]
    repos = backup_data["repositories"]
    
    print(f"\n  Backup from: {backup_data['timestamp']}")
    print(f"  Organization: {org}")
    print(f"  Repositories in backup: {len(repos)}")
    
    restored = 0
    errors = []
    
    print("\n  Restoring settings...")
    
    for repo_backup in repos:
        repo_name = repo_backup["repository"]
        print(f"\n    {repo_name}:")
        
        # Restore visibility
        original_private = repo_backup.get("private")
        if original_private is not None:
            try:
                api_client.patch(f"/repos/{org}/{repo_name}", {"private": original_private})
                print(f"      - Visibility: {'private' if original_private else 'public'} ✓")
                restored += 1
                time.sleep(SLEEP_INTERVAL)
            except requests.exceptions.HTTPError as e:
                print(f"      - Visibility: ERROR: {e}")
                errors.append({"repo": repo_name, "field": "visibility", "error": str(e)})
        
        # Restore archived status
        original_archived = repo_backup.get("archived")
        if original_archived is not None:
            try:
                api_client.patch(f"/repos/{org}/{repo_name}", {"archived": original_archived})
                print(f"      - Archived: {original_archived} ✓")
                restored += 1
                time.sleep(SLEEP_INTERVAL)
            except requests.exceptions.HTTPError as e:
                print(f"      - Archived: ERROR: {e}")
                errors.append({"repo": repo_name, "field": "archived", "error": str(e)})
        
        # Restore webhook SSL settings
        for hook in repo_backup.get("hooks", []):
            hook_id = hook.get("id")
            original_ssl = hook.get("config", {}).get("insecure_ssl")
            if hook_id and original_ssl is not None:
                try:
                    api_client.patch(
                        f"/repos/{org}/{repo_name}/hooks/{hook_id}",
                        {"config": {"insecure_ssl": str(original_ssl)}}
                    )
                    print(f"      - Hook {hook_id} SSL: {original_ssl} ✓")
                    restored += 1
                    time.sleep(SLEEP_INTERVAL)
                except requests.exceptions.HTTPError as e:
                    print(f"      - Hook {hook_id}: ERROR: {e}")
                    errors.append({"repo": repo_name, "hook": hook_id, "error": str(e)})
        
        # Note: Re-adding collaborators and teams would require storing their
        # original permission levels, which is more complex. Just log info.
        if repo_backup.get("outside_collaborators"):
            print(f"      - Note: {len(repo_backup['outside_collaborators'])} outside collaborators were removed")
            print(f"        Re-add manually: {repo_backup['outside_collaborators']}")
        
        if repo_backup.get("direct_collaborators"):
            print(f"      - Note: {len(repo_backup['direct_collaborators'])} direct collaborators were removed")
            print(f"        Re-add manually: {repo_backup['direct_collaborators']}")
    
    # Summary
    print("\n" + "-" * 40)
    print("ROLLBACK SUMMARY")
    print("-" * 40)
    print(f"  Settings restored: {restored}")
    print(f"  Errors: {len(errors)}")
    
    return {
        "restored": restored,
        "errors": len(errors),
        "error_details": errors
    }


# =============================================================================
# COMMAND LINE INTERFACE
# =============================================================================

def parse_arguments():
    """Parse command line arguments."""
    parser = argparse.ArgumentParser(
        description="GitHub Repository Compliance Checker & Enforcer",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  %(prog)s --check                    Check compliance (default, report only)
  %(prog)s --repo my-repo             Check only a specific repository
  %(prog)s --apply                    Apply compliant settings to all repos
  %(prog)s --apply --dry-run          Preview changes without applying
  %(prog)s --repo my-repo --apply     Apply fixes to one repo only
  %(prog)s --rollback backup.json     Restore settings from backup file

Settings that can be applied automatically:
  - unsecure_hooks (enable SSL verification on webhooks)
  - collaborators_in_org (remove outside collaborators)
  - collaborators_in_team (remove direct collaborators)
  - shared_repo_readers (remove Cloud_Readers team access)
  - private_if_sensitive (make repository private)
  - archived_status (unarchive repository)

Settings that require manual action:
  - metadata_existing (.metadata file must be created manually)
        """
    )
    
    mode_group = parser.add_mutually_exclusive_group()
    mode_group.add_argument(
        "--check", "-c",
        action="store_true",
        default=True,
        help="Check compliance and generate reports (default)"
    )
    mode_group.add_argument(
        "--apply", "-a",
        action="store_true",
        help="Apply compliant settings to non-compliant repositories"
    )
    mode_group.add_argument(
        "--rollback", "-r",
        metavar="BACKUP_FILE",
        help="Rollback to settings from backup file"
    )
    
    parser.add_argument(
        "--dry-run", "-n",
        action="store_true",
        help="Preview changes without applying (use with --apply)"
    )
    
    parser.add_argument(
        "--repo",
        metavar="REPO_NAME",
        help="Target a specific repository (for testing before org-wide apply)"
    )
    
    return parser.parse_args()


# =============================================================================
# MAIN ENTRY POINT
# =============================================================================

def main():
    """Main entry point."""
    args = parse_arguments()
    
    print("\n" + "=" * 60)
    print("GHE REPOSITORY COMPLIANCE CHECKER")
    print("=" * 60)
    
    # Validate configuration
    if not GITHUB_TOKEN:
        print("ERROR: GITHUB_TOKEN environment variable is not set.")
        sys.exit(1)
    if not GITHUB_ORG:
        print("ERROR: GITHUB_ORG environment variable is not set.")
        sys.exit(1)
    
    print(f"\nConfiguration:")
    print(f"  Organization: {GITHUB_ORG}")
    print(f"  API Base URL: {GITHUB_BASE}")
    if args.repo:
        print(f"  Target Repo: {args.repo}")
    
    # Initialize API client
    api_client = GitHubAPIClient(GITHUB_BASE, GITHUB_TOKEN)
    
    # Handle ROLLBACK mode
    if args.rollback:
        print(f"  Mode: ROLLBACK from {args.rollback}")
        rollback_from_backup(api_client, args.rollback, GITHUB_ORG)
        print("\n" + "=" * 60)
        return
    
    # Handle CHECK and APPLY modes
    if args.apply:
        print(f"  Mode: APPLY {'(DRY RUN)' if args.dry_run else ''}")
    else:
        print(f"  Mode: CHECK (report only)")
    
    # Initialize checker and run checks
    checker = RepoComplianceChecker(api_client, GITHUB_ORG)
    results = checker.run_all_checks()
    
    # Filter results if target repo specified
    if args.repo:
        results = [r for r in results if r["repository"] == args.repo]
        if not results:
            print(f"\n  Repository '{args.repo}' not found.")
            return
    
    if not results:
        print("\n  No repositories found in organization.")
        return
    
    # Generate reports
    report_gen = ReportGenerator(GITHUB_ORG, results)
    report_gen.generate_all_reports()
    
    # Print summary
    total_repos = len(results)
    repos_with_issues = sum(1 for r in results if any(not rule["passed"] for rule in r["rules"]))
    total_rules = sum(len(r["rules"]) for r in results)
    total_failed = sum(1 for r in results for rule in r["rules"] if not rule["passed"])
    
    print("\n" + "=" * 60)
    print("CHECK SUMMARY")
    print("=" * 60)
    print(f"  Repositories Checked: {total_repos}")
    print(f"  Repositories with Issues: {repos_with_issues}")
    print(f"  Compliant Repositories: {total_repos - repos_with_issues}")
    print(f"  Total Rules Checked: {total_rules}")
    print(f"  Failed Rules: {total_failed}")
    
    if repos_with_issues > 0:
        print("\n  ⚠️  COMPLIANCE ISSUES DETECTED - Review failed repositories!")
    else:
        print("\n  ✅ All repositories are compliant!")
    
    # Handle APPLY mode
    if args.apply:
        if repos_with_issues == 0:
            print("\n  No compliance issues to fix.")
        else:
            applier = RepoComplianceApplier(api_client, GITHUB_ORG, dry_run=args.dry_run)
            apply_result = applier.apply_all(results, target_repo=args.repo)
            
            # Save apply results to file
            apply_log_file = f"repo_apply_log_{datetime.now().strftime('%Y-%m-%d_%H%M%S')}.json"
            with open(apply_log_file, "w", encoding="utf-8") as f:
                json.dump({
                    "timestamp": datetime.now().isoformat(),
                    "organization": GITHUB_ORG,
                    "target_repo": args.repo,
                    "dry_run": args.dry_run,
                    "summary": apply_result,
                    "changes": applier.changes_made,
                    "skipped": applier.skipped,
                    "errors": applier.errors
                }, f, indent=2, default=str)
            print(f"\n  Apply log saved: {apply_log_file}")
    
    print("\n" + "=" * 60)


if __name__ == "__main__":
    main()
