"""
================================================================================
ORGANIZATION-LEVEL COMPLIANCE CHECKER
================================================================================

This script checks organization-level settings against IBM CISO policy requirements.

HOW TO RUN:
    1. Set environment variables:
       - GITHUB_TOKEN: Your GitHub personal access token
       - GITHUB_ORG: Organization name to check
       - GITHUB_BASE: GitHub API base URL (e.g., https://api.github.example.com)
    
    2. Run: python org_compliance.py
    
    3. Output files will be generated:
       - org_compliance_report.json
       - org_compliance_report.md
       - org_compliance_report.xlsx

RULES CHECKED (Reference: IBM Cloud Policy 3.1.3, 3.1.4, ITSS Chapter 2):

REQUIRED RULES:
---------------
1. default_repository_permission
   - Setting: Base permissions (Member privileges section)
   - Required Value: "No permission"
   - How we check: API field "default_repository_permission" must equal "none"
   - Why: Other values allow members to read all private repositories by default.

2. org_outside_collaborators  
   - Setting: Allow repository administrators to add outside collaborators
   - Required Value: Disabled
   - How we check: API field "members_can_invite_outside_collaborators" must be False
   - Why: All access must be through teams managed in AccessHub.

3. unsecure_org_hooks
   - Setting: SSL verification on organization webhooks
   - Required Value: Enable SSL verification
   - How we check: Fetch /orgs/{org}/hooks and check each hook's config.insecure_ssl must be 0/False/None
   - Why: Needed to maintain confidentiality of webhook data in transit.

RECOMMENDED RULES:
------------------
4. members_can_create_public_repositories
   - Setting: Repository Creation (Member privileges section)
   - Recommended Value: Private (only)
   - How we check: API field "members_can_create_public_repositories" should be False
   - Why: Prevents accidental public exposure of sensitive code.

5. visibility_change_disabled
   - Setting: Allow members to change repository visibilities
   - Recommended Value: Disabled
   - How we check: API field "members_can_change_repo_visibility" should be False
   - Why: Prevents accidental visibility changes that could expose private code.

6. delete_transfer_disabled
   - Setting: Allow members to delete or transfer repositories
   - Recommended Value: Disabled
   - How we check: API field "members_can_delete_repositories" should be False
   - Why: Limits accidental or malicious deletion/removal of repositories.

7. team_creation_disabled
   - Setting: Allow members to create teams
   - Recommended Value: Disabled
   - How we check: API field "members_can_create_teams" should be False
   - Why: Ensures access management through AccessHub is not subverted.

8. admin_activity_6_months
   - Setting: Organization admins should have recent activity
   - Recommended Value: Activity within last 6 months
   - How we check: Fetch /orgs/{org}/members?role=admin and check each admin's
     /users/{login}/events for recent activity
   - Why: Inactive admins indicate poor access governance and revalidation.

Author: GitHub Compliance Team
Last Updated: 2024
================================================================================
"""

import os
import sys
import json
import time
import argparse
import requests
import urllib3
from datetime import datetime, timedelta, timezone

# Disable SSL warnings for GHE with self-signed certificates
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# Excel support (required for report generation)
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


# =============================================================================
# CONFIGURATION
# =============================================================================

GITHUB_TOKEN = os.environ.get("GITHUB_TOKEN")
GITHUB_ORG = os.environ.get("GITHUB_ORG")
GITHUB_BASE = os.environ.get("GITHUB_BASE", "https://api.github.com")

# API settings
SLEEP_INTERVAL = 0.3  # Delay between API calls to avoid rate limiting


# =============================================================================
# GITHUB API CLIENT
# =============================================================================

class GitHubAPIClient:
    """
    Simple GitHub API client with authentication and pagination support.
    """
    
    def __init__(self, base_url, token):
        self.base_url = base_url.rstrip("/")
        self.headers = {
            "Authorization": f"token {token}",
            "Accept": "application/vnd.github.v3+json"
        }
    
    def get(self, endpoint, allow_404=False):
        """
        Make a GET request to the GitHub API.
        
        Args:
            endpoint: API endpoint (e.g., /orgs/myorg)
            allow_404: If True, return None for 404 errors instead of raising
        
        Returns:
            dict or list: JSON response data
        """
        url = f"{self.base_url}{endpoint}"
        response = requests.get(url, headers=self.headers, verify=False)
        
        if response.status_code == 404 and allow_404:
            return None
        
        response.raise_for_status()
        return response.json()
    
    def paginate(self, endpoint):
        """
        Fetch all pages of a paginated API endpoint.
        
        Args:
            endpoint: API endpoint with pagination support
        
        Returns:
            list: All items from all pages
        """
        results = []
        url = f"{self.base_url}{endpoint}"
        
        while url:
            response = requests.get(url, headers=self.headers, verify=False)
            response.raise_for_status()
            data = response.json()
            
            if isinstance(data, list):
                results.extend(data)
            else:
                results.append(data)
            
            # Get next page URL from Link header
            url = None
            link_header = response.headers.get("Link", "")
            for link in link_header.split(","):
                if 'rel="next"' in link:
                    url = link.split(";")[0].strip()[1:-1]
                    break
            
            time.sleep(SLEEP_INTERVAL)
        
        return results
    
    def put(self, endpoint, data):
        """
        Make a PUT request to the GitHub API.
        
        Args:
            endpoint: API endpoint path
            data: JSON data to send
        
        Returns:
            dict: Response JSON
        """
        url = f"{self.base_url}{endpoint}"
        response = requests.put(url, headers=self.headers, json=data, verify=False)
        response.raise_for_status()
        return response.json() if response.text else {}
    
    def patch(self, endpoint, data):
        """
        Make a PATCH request to the GitHub API.
        
        Used for updating organization settings.
        
        Args:
            endpoint: API endpoint path
            data: JSON data to send
        
        Returns:
            dict: Response JSON
        """
        url = f"{self.base_url}{endpoint}"
        response = requests.patch(url, headers=self.headers, json=data, verify=False)
        response.raise_for_status()
        return response.json() if response.text else {}
    
    def delete(self, endpoint):
        """
        Make a DELETE request to the GitHub API.
        
        Args:
            endpoint: API endpoint path
        
        Returns:
            bool: True if successful
        """
        url = f"{self.base_url}{endpoint}"
        response = requests.delete(url, headers=self.headers, verify=False)
        response.raise_for_status()
        return True


# =============================================================================
# ORGANIZATION COMPLIANCE CHECKER
# =============================================================================

class OrgComplianceChecker:
    """
    Checks organization-level settings against IBM CISO policy requirements.
    """
    
    def __init__(self, api_client, org_name):
        self.api = api_client
        self.org = org_name
        self.org_data = None
        self.results = []
    
    def fetch_org_settings(self):
        """
        Fetch organization settings from GitHub API.
        
        API Call: GET /orgs/{org}
        Returns organization details including all settings we need to check.
        """
        print(f"  Fetching organization settings for '{self.org}'...")
        self.org_data = self.api.get(f"/orgs/{self.org}")
        return self.org_data
    
    # =========================================================================
    # REQUIRED RULES
    # =========================================================================
    
    def check_default_repository_permission(self):
        """
        REQUIRED RULE: default_repository_permission
        
        HOW WE APPLY THIS RULE:
        -----------------------
        1. We read the "default_repository_permission" field from org settings
        2. This field can be: "none", "read", "write", or "admin"
        3. We check if it equals "none"
        4. If not "none", members can read all private repos by default - VIOLATION
        
        Setting Location in GitHub UI: 
            Organization Settings > Member privileges > Base permissions
        
        API Field: default_repository_permission
        Expected Value: "none"
        """
        print("    Checking: default_repository_permission (REQUIRED)")
        
        # Get current value from organization settings
        current = self.org_data.get("default_repository_permission", "read")
        expected = "none"
        passed = current == expected
        
        result = {
            "rule": "default_repository_permission",
            "passed": passed,
            "current_value": current,
            "expected_value": expected,
            "enforcement": "Required",
            "reason": (
                f"Base permissions is set to '{current}'. Must be 'No permission' (none). "
                "Other values allow all organization members to read private repositories."
                if not passed else
                "Base permissions correctly set to 'No permission'."
            )
        }
        self.results.append(result)
        return result
    
    def check_org_outside_collaborators(self):
        """
        REQUIRED RULE: org_outside_collaborators
        
        HOW WE APPLY THIS RULE:
        -----------------------
        1. We read "members_can_invite_outside_collaborators" from org settings
        2. This boolean indicates if repo admins can add outside collaborators
        3. We check if it is False (disabled)
        4. If True, repo admins can add non-org members directly - VIOLATION
        
        Setting Location in GitHub UI:
            Organization Settings > Member privileges > 
            "Allow repository administrators to add outside collaborators"
        
        API Field: members_can_invite_outside_collaborators
        Expected Value: False
        """
        print("    Checking: org_outside_collaborators (REQUIRED)")
        
        # Get current value - True means admins CAN add outside collaborators
        can_invite = self.org_data.get("members_can_invite_outside_collaborators", True)
        passed = not can_invite
        
        result = {
            "rule": "org_outside_collaborators",
            "passed": passed,
            "current_value": "Enabled" if can_invite else "Disabled",
            "expected_value": "Disabled",
            "enforcement": "Required",
            "reason": (
                "Outside collaborators can be added by repository administrators. "
                "This must be disabled - all access should be through AccessHub teams."
                if not passed else
                "Outside collaborators cannot be added by repository administrators."
            )
        }
        self.results.append(result)
        return result
    
    def check_unsecure_org_hooks(self):
        """
        REQUIRED RULE: unsecure_org_hooks
        
        HOW WE APPLY THIS RULE:
        -----------------------
        1. We fetch all organization webhooks via GET /orgs/{org}/hooks
        2. For each hook, we examine config.insecure_ssl field
        3. insecure_ssl should be "0", 0, False, or None (meaning SSL IS enabled)
        4. If insecure_ssl is "1", 1, or True - SSL is DISABLED - VIOLATION
        5. We collect all hooks with SSL disabled and report them
        
        Setting Location in GitHub UI:
            Organization Settings > Webhooks > [each webhook] > SSL verification
        
        API Endpoint: GET /orgs/{org}/hooks
        Check Field: config.insecure_ssl
        Expected Value: "0" or 0 or False (meaning SSL verification is enabled)
        
        Note: The field name is confusing - "insecure_ssl=1" means SSL is DISABLED
        """
        print("    Checking: unsecure_org_hooks (REQUIRED)")
        
        # Fetch all organization webhooks
        hooks = self.api.get(f"/orgs/{self.org}/hooks", allow_404=True) or []
        time.sleep(SLEEP_INTERVAL)
        
        # Find hooks with SSL verification disabled
        # insecure_ssl = "1" or 1 or True means SSL is DISABLED (bad)
        # insecure_ssl = "0" or 0 or False or None means SSL is ENABLED (good)
        insecure_hooks = []
        for hook in hooks:
            config = hook.get("config", {})
            insecure_ssl = config.get("insecure_ssl")
            
            # Check if SSL is disabled
            if insecure_ssl not in (None, "0", 0, False):
                insecure_hooks.append({
                    "id": hook["id"],
                    "name": hook.get("name", "unknown"),
                    "url": config.get("url", "N/A"),
                    "insecure_ssl": insecure_ssl
                })
        
        passed = len(insecure_hooks) == 0
        
        result = {
            "rule": "unsecure_org_hooks",
            "passed": passed,
            "current_value": f"{len(insecure_hooks)} hooks with SSL disabled",
            "expected_value": "0 hooks with SSL disabled",
            "enforcement": "Required",
            "insecure_hooks": insecure_hooks,
            "reason": (
                f"Found {len(insecure_hooks)} webhook(s) without SSL verification. "
                f"Hook IDs: {[h['id'] for h in insecure_hooks]}. "
                "All webhooks must have SSL verification enabled for confidentiality."
                if not passed else
                "All organization webhooks have SSL verification enabled."
            )
        }
        self.results.append(result)
        return result
    
    # =========================================================================
    # RECOMMENDED RULES
    # =========================================================================
    
    def check_members_can_create_public_repositories(self):
        """
        RECOMMENDED RULE: members_can_create_public_repositories
        
        HOW WE APPLY THIS RULE:
        -----------------------
        1. We read "members_can_create_public_repositories" from org settings
        2. This boolean indicates if members can create public repos
        3. We check if it is False (disabled)
        4. If True, members might accidentally create public repos - WARNING
        
        Setting Location in GitHub UI:
            Organization Settings > Member privileges > Repository creation
        
        API Field: members_can_create_public_repositories
        Expected Value: False
        """
        print("    Checking: members_can_create_public_repositories (RECOMMENDED)")
        
        can_create_public = self.org_data.get("members_can_create_public_repositories", True)
        can_create_internal = self.org_data.get("members_can_create_internal_repositories", True)
        passed = not can_create_public
        
        result = {
            "rule": "members_can_create_public_repositories",
            "passed": passed,
            "current_value": f"Public: {'Yes' if can_create_public else 'No'}, Internal: {'Yes' if can_create_internal else 'No'}",
            "expected_value": "Public: No",
            "enforcement": "Recommended",
            "reason": (
                "Members can create public repositories. This should be restricted to "
                "private only to prevent accidental public exposure of sensitive code."
                if not passed else
                "Public repository creation is appropriately restricted."
            )
        }
        self.results.append(result)
        return result
    
    def check_visibility_change_disabled(self):
        """
        RECOMMENDED RULE: visibility_change_disabled
        
        HOW WE APPLY THIS RULE:
        -----------------------
        1. We read "members_can_change_repo_visibility" from org settings
        2. This boolean indicates if members can change repo visibility
        3. We check if it is False (disabled)
        4. If True, members could make private repos public - WARNING
        
        Setting Location in GitHub UI:
            Organization Settings > Member privileges > 
            "Allow members to change repository visibilities"
        
        API Field: members_can_change_repo_visibility
        Expected Value: False
        """
        print("    Checking: visibility_change_disabled (RECOMMENDED)")
        
        can_change = self.org_data.get("members_can_change_repo_visibility", True)
        passed = not can_change
        
        result = {
            "rule": "visibility_change_disabled",
            "passed": passed,
            "current_value": "Enabled" if can_change else "Disabled",
            "expected_value": "Disabled",
            "enforcement": "Recommended",
            "reason": (
                "Members can change repository visibility. This could allow "
                "accidental exposure of private repositories to the public."
                if not passed else
                "Repository visibility changes are appropriately restricted."
            )
        }
        self.results.append(result)
        return result
    
    def check_delete_transfer_disabled(self):
        """
        RECOMMENDED RULE: delete_transfer_disabled
        
        HOW WE APPLY THIS RULE:
        -----------------------
        1. We read "members_can_delete_repositories" from org settings
        2. This boolean indicates if members can delete/transfer repositories
        3. We check if it is False (disabled)
        4. If True, members could delete repos - WARNING
        
        Setting Location in GitHub UI:
            Organization Settings > Member privileges > 
            "Allow members to delete or transfer repositories"
        
        API Field: members_can_delete_repositories
        Expected Value: False
        """
        print("    Checking: delete_transfer_disabled (RECOMMENDED)")
        
        can_delete = self.org_data.get("members_can_delete_repositories", True)
        passed = not can_delete
        
        result = {
            "rule": "delete_transfer_disabled",
            "passed": passed,
            "current_value": "Enabled" if can_delete else "Disabled",
            "expected_value": "Disabled",
            "enforcement": "Recommended",
            "reason": (
                "Members can delete or transfer repositories. This could lead to "
                "accidental or malicious loss of code and history."
                if not passed else
                "Repository deletion/transfer is appropriately restricted."
            )
        }
        self.results.append(result)
        return result
    
    def check_team_creation_disabled(self):
        """
        RECOMMENDED RULE: team_creation_disabled
        
        HOW WE APPLY THIS RULE:
        -----------------------
        1. We read "members_can_create_teams" from org settings
        2. This boolean indicates if members can create teams
        3. We check if it is False (disabled)
        4. If True, members could create teams outside AccessHub - WARNING
        
        Setting Location in GitHub UI:
            Organization Settings > Member privileges > 
            "Allow members to create teams"
        
        API Field: members_can_create_teams
        Expected Value: False
        """
        print("    Checking: team_creation_disabled (RECOMMENDED)")
        
        can_create = self.org_data.get("members_can_create_teams", True)
        passed = not can_create
        
        result = {
            "rule": "team_creation_disabled",
            "passed": passed,
            "current_value": "Enabled" if can_create else "Disabled",
            "expected_value": "Disabled",
            "enforcement": "Recommended",
            "reason": (
                "Members can create teams. This could subvert access management "
                "through AccessHub and create unauthorized access patterns."
                if not passed else
                "Team creation is appropriately restricted to admins."
            )
        }
        self.results.append(result)
        return result
    
    def check_admin_activity_6_months(self):
        """
        RECOMMENDED RULE: admin_activity_6_months
        
        HOW WE APPLY THIS RULE:
        -----------------------
        1. Fetch all organization admins via GET /orgs/{org}/members?role=admin
        2. For each admin, fetch their public events via GET /users/{login}/events
        3. Check if any event occurred within the last 6 months
        4. Report admins with no recent activity - WARNING
        
        Note: This is not a GitHub setting but a governance check.
        Inactive admins indicate poor access revalidation practices.
        
        API Endpoints:
            GET /orgs/{org}/members?role=admin - List org admins
            GET /users/{login}/events - Get user's recent activity
        
        Expected: All admins should have activity within 6 months
        """
        print("    Checking: admin_activity_6_months (RECOMMENDED)")
        
        # Fetch organization admins
        admins = self.api.paginate(f"/orgs/{self.org}/members?role=admin&per_page=100")
        time.sleep(SLEEP_INTERVAL)
        
        six_months_ago = datetime.now(timezone.utc) - timedelta(days=180)
        inactive_admins = []
        
        for admin in admins:
            login = admin["login"]
            
            # Fetch user's recent events
            events = self.api.get(f"/users/{login}/events?per_page=10", allow_404=True) or []
            time.sleep(SLEEP_INTERVAL)
            
            # Check if any event is within 6 months
            has_recent_activity = False
            for event in events:
                created_at = event.get("created_at", "")
                if created_at:
                    try:
                        event_date = datetime.strptime(created_at, "%Y-%m-%dT%H:%M:%SZ").replace(tzinfo=timezone.utc)
                        if event_date > six_months_ago:
                            has_recent_activity = True
                            break
                    except ValueError:
                        pass
            
            if not has_recent_activity:
                inactive_admins.append(login)
        
        passed = len(inactive_admins) == 0
        
        result = {
            "rule": "admin_activity_6_months",
            "passed": passed,
            "current_value": f"{len(inactive_admins)} inactive admins",
            "expected_value": "0 inactive admins",
            "enforcement": "Recommended",
            "inactive_admins": inactive_admins,
            "total_admins": len(admins),
            "reason": (
                f"Found {len(inactive_admins)} admin(s) without activity in 6 months: "
                f"{inactive_admins}. Inactive admins should be reviewed for access removal."
                if not passed else
                f"All {len(admins)} organization admins have recent activity."
            )
        }
        self.results.append(result)
        return result
    
    # =========================================================================
    # RUN ALL CHECKS
    # =========================================================================
    
    def run_all_checks(self):
        """
        Execute all organization-level compliance checks.
        
        Returns:
            list: All check results
        """
        print("\n" + "=" * 60)
        print("ORGANIZATION-LEVEL COMPLIANCE CHECKS")
        print("=" * 60)
        
        # Fetch organization settings first
        self.fetch_org_settings()
        
        print("\n  Running Required Rules...")
        self.check_default_repository_permission()
        self.check_org_outside_collaborators()
        self.check_unsecure_org_hooks()
        
        print("\n  Running Recommended Rules...")
        self.check_members_can_create_public_repositories()
        self.check_visibility_change_disabled()
        self.check_delete_transfer_disabled()
        self.check_team_creation_disabled()
        self.check_admin_activity_6_months()
        
        return self.results


# =============================================================================
# REPORT GENERATOR
# =============================================================================

class ReportGenerator:
    """
    Generates compliance reports in JSON, Markdown, and Excel formats.
    """
    
    def __init__(self, org_name, results):
        self.org = org_name
        self.results = results
        self.timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    def generate_json_report(self, filepath="org_compliance_report.json"):
        """Generate JSON report."""
        report = {
            "report_type": "Organization Compliance",
            "organization": self.org,
            "generated_at": self.timestamp,
            "summary": {
                "total_rules": len(self.results),
                "passed": sum(1 for r in self.results if r["passed"]),
                "failed": sum(1 for r in self.results if not r["passed"]),
                "required_failed": sum(1 for r in self.results if not r["passed"] and r["enforcement"] == "Required"),
                "recommended_failed": sum(1 for r in self.results if not r["passed"] and r["enforcement"] == "Recommended")
            },
            "results": self.results
        }
        
        with open(filepath, "w", encoding="utf-8") as f:
            json.dump(report, f, indent=2, default=str)
        
        print(f"  JSON report saved: {filepath}")
        return filepath
    
    def generate_markdown_report(self, filepath="org_compliance_report.md"):
        """Generate Markdown report."""
        passed = sum(1 for r in self.results if r["passed"])
        failed = sum(1 for r in self.results if not r["passed"])
        required_failed = sum(1 for r in self.results if not r["passed"] and r["enforcement"] == "Required")
        
        lines = [
            "# Organization Compliance Report",
            "",
            f"**Organization:** {self.org}",
            f"**Generated:** {self.timestamp}",
            "",
            "## Summary",
            "",
            f"- **Total Rules Checked:** {len(self.results)}",
            f"- **Passed:** {passed}",
            f"- **Failed:** {failed}",
            f"- **Required Rules Failed:** {required_failed}",
            "",
        ]
        
        # Required rules
        lines.extend(["## Required Rules", ""])
        lines.append("| Rule | Status | Current Value | Expected Value |")
        lines.append("|------|--------|---------------|----------------|")
        for r in self.results:
            if r["enforcement"] == "Required":
                status = "✅ PASS" if r["passed"] else "❌ FAIL"
                lines.append(f"| {r['rule']} | {status} | {r['current_value']} | {r['expected_value']} |")
        lines.append("")
        
        # Recommended rules
        lines.extend(["## Recommended Rules", ""])
        lines.append("| Rule | Status | Current Value | Expected Value |")
        lines.append("|------|--------|---------------|----------------|")
        for r in self.results:
            if r["enforcement"] == "Recommended":
                status = "✅ PASS" if r["passed"] else "⚠️ WARN"
                lines.append(f"| {r['rule']} | {status} | {r['current_value']} | {r['expected_value']} |")
        lines.append("")
        
        # Details for failed rules
        failed_results = [r for r in self.results if not r["passed"]]
        if failed_results:
            lines.extend(["## Failed Rule Details", ""])
            for r in failed_results:
                lines.append(f"### {r['rule']} ({r['enforcement']})")
                lines.append(f"**Reason:** {r['reason']}")
                lines.append("")
        
        with open(filepath, "w", encoding="utf-8") as f:
            f.write("\n".join(lines))
        
        print(f"  Markdown report saved: {filepath}")
        return filepath
    
    def generate_excel_report(self, filepath="org_compliance_report.xlsx"):
        """Generate Excel report."""
        wb = openpyxl.Workbook()
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        warn_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        # Summary Sheet
        ws = wb.active
        ws.title = "Summary"
        
        passed = sum(1 for r in self.results if r["passed"])
        failed = sum(1 for r in self.results if not r["passed"])
        
        summary_data = [
            ["Organization Compliance Report", ""],
            ["", ""],
            ["Organization", self.org],
            ["Generated", self.timestamp],
            ["", ""],
            ["Total Rules", len(self.results)],
            ["Passed", passed],
            ["Failed", failed],
            ["Compliance %", f"{(passed/len(self.results)*100):.1f}%" if self.results else "N/A"]
        ]
        
        for row_idx, row in enumerate(summary_data, 1):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 1:
                    cell.font = Font(bold=True, size=14)
        
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
        
        # Results Sheet
        ws2 = wb.create_sheet("Rule Results")
        headers = ["Rule", "Status", "Enforcement", "Current Value", "Expected Value", "Reason"]
        
        for col, header in enumerate(headers, 1):
            cell = ws2.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        
        for row_idx, r in enumerate(self.results, 2):
            values = [
                r["rule"],
                "PASS" if r["passed"] else "FAIL",
                r["enforcement"],
                r["current_value"],
                r["expected_value"],
                r["reason"]
            ]
            for col_idx, value in enumerate(values, 1):
                cell = ws2.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                if col_idx == 2:  # Status column
                    if r["passed"]:
                        cell.fill = pass_fill
                    elif r["enforcement"] == "Required":
                        cell.fill = fail_fill
                    else:
                        cell.fill = warn_fill
        
        # Adjust column widths
        ws2.column_dimensions['A'].width = 35
        ws2.column_dimensions['B'].width = 10
        ws2.column_dimensions['C'].width = 15
        ws2.column_dimensions['D'].width = 40
        ws2.column_dimensions['E'].width = 25
        ws2.column_dimensions['F'].width = 60
        
        wb.save(filepath)
        print(f"  Excel report saved: {filepath}")
        return filepath
    
    def generate_all_reports(self):
        """Generate all report formats."""
        print("\nGenerating Reports...")
        self.generate_json_report()
        self.generate_markdown_report()
        self.generate_excel_report()


# =============================================================================
# ORGANIZATION COMPLIANCE APPLIER
# =============================================================================

class OrgComplianceApplier:
    """
    Applies compliant organization settings.
    Supports backup before changes and rollback if needed.
    
    APPLY MODE:
    -----------
    1. Fetch current organization settings
    2. Save backup to org_backup_TIMESTAMP.json
    3. Apply compliant settings for failed rules
    4. Generate report of changes made
    
    ROLLBACK MODE:
    --------------
    1. Read backup file
    2. Restore original organization settings
    3. Report what was restored
    
    WHAT CAN BE APPLIED:
    --------------------
    - default_repository_permission: PATCH /orgs/{org} with {"default_repository_permission": "none"}
    - members_can_invite_outside_collaborators: PATCH /orgs/{org} with {"members_can_invite_outside_collaborators": false}
    - members_can_create_public_repositories: PATCH /orgs/{org}
    - members_can_change_repo_visibility: PATCH /orgs/{org}
    - members_can_delete_repositories: PATCH /orgs/{org}
    - members_can_create_teams: PATCH /orgs/{org}
    
    WHAT CANNOT BE APPLIED AUTOMATICALLY:
    -------------------------------------
    - unsecure_org_hooks: Requires manual webhook reconfiguration or deleting insecure hooks
    - admin_activity_6_months: Governance issue, requires manual admin review
    """
    
    def __init__(self, api_client, org_name, dry_run=False):
        self.api = api_client
        self.org = org_name
        self.dry_run = dry_run
        self.changes_made = []
        self.errors = []
        self.skipped = []
    
    def get_compliant_settings(self):
        """
        Get the compliant settings payload for organization.
        
        API: PATCH /orgs/{org}
        
        Returns:
            dict: Compliant settings payload
        """
        return {
            "default_repository_permission": "none",                    # Required rule
            "members_can_invite_outside_collaborators": False,         # Required rule (via API: members_allowed_repository_creation_type)
            "members_can_create_public_repositories": False,           # Recommended rule
            "members_can_change_repo_visibility": False,               # Recommended rule
            "members_can_delete_repositories": False,                  # Recommended rule
            "members_can_create_teams": False                          # Recommended rule
        }
    
    def backup_current_settings(self, current_org_data, check_results):
        """
        Save current organization settings to backup file.
        
        Args:
            current_org_data: Current organization settings from API
            check_results: Results from compliance check
        
        Returns:
            str: Path to backup file
        """
        timestamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
        backup_file = f"org_backup_{timestamp}.json"
        
        # Extract relevant settings
        backup_data = {
            "timestamp": datetime.now().isoformat(),
            "organization": self.org,
            "settings": {
                "default_repository_permission": current_org_data.get("default_repository_permission"),
                "members_can_invite_outside_collaborators": current_org_data.get("members_can_invite_outside_collaborators"),
                "members_can_create_public_repositories": current_org_data.get("members_can_create_public_repositories"),
                "members_can_change_repo_visibility": current_org_data.get("members_can_change_repo_visibility"),
                "members_can_delete_repositories": current_org_data.get("members_can_delete_repositories"),
                "members_can_create_teams": current_org_data.get("members_can_create_teams"),
                "members_can_create_internal_repositories": current_org_data.get("members_can_create_internal_repositories"),
                "members_can_create_private_repositories": current_org_data.get("members_can_create_private_repositories")
            },
            "check_results": check_results
        }
        
        with open(backup_file, "w", encoding="utf-8") as f:
            json.dump(backup_data, f, indent=2, default=str)
        
        print(f"  Backup saved: {backup_file}")
        return backup_file
    
    def apply_rule(self, rule_name, current_value, compliant_value):
        """
        Apply a single compliant setting.
        
        Args:
            rule_name: Name of the rule/setting
            current_value: Current value
            compliant_value: Value to set for compliance
        
        Returns:
            dict: Result with success status
        """
        # Map rule names to API field names
        field_mapping = {
            "default_repository_permission": "default_repository_permission",
            "org_outside_collaborators": "members_can_invite_outside_collaborators",
            "members_can_create_public_repositories": "members_can_create_public_repositories",
            "visibility_change_disabled": "members_can_change_repo_visibility",
            "delete_transfer_disabled": "members_can_delete_repositories",
            "team_creation_disabled": "members_can_create_teams"
        }
        
        api_field = field_mapping.get(rule_name)
        if not api_field:
            return {
                "success": False,
                "rule": rule_name,
                "error": f"Cannot apply automatically: {rule_name} requires manual intervention"
            }
        
        if self.dry_run:
            return {
                "success": True,
                "dry_run": True,
                "rule": rule_name,
                "field": api_field,
                "old_value": current_value,
                "new_value": compliant_value,
                "action": f"Would set {api_field} to {compliant_value}"
            }
        
        try:
            payload = {api_field: compliant_value}
            self.api.patch(f"/orgs/{self.org}", payload)
            time.sleep(SLEEP_INTERVAL)
            
            return {
                "success": True,
                "rule": rule_name,
                "field": api_field,
                "old_value": current_value,
                "new_value": compliant_value,
                "action": f"Set {api_field} to {compliant_value}"
            }
        except requests.exceptions.HTTPError as e:
            return {
                "success": False,
                "rule": rule_name,
                "field": api_field,
                "error": str(e)
            }
    
    def apply_all(self, check_results, current_org_data):
        """
        Apply compliant settings for all failed rules.
        
        Args:
            check_results: Results from OrgComplianceChecker
            current_org_data: Current organization settings
        
        Returns:
            dict: Summary of changes made
        """
        print("\n" + "=" * 60)
        print("APPLYING ORGANIZATION COMPLIANCE RULES")
        print("=" * 60)
        
        if self.dry_run:
            print("\n  *** DRY RUN MODE - No changes will be made ***\n")
        
        # Step 1: Create backup
        print("\n  Creating backup of current settings...")
        backup_file = self.backup_current_settings(current_org_data, check_results)
        
        # Step 2: Identify failed rules that can be applied
        print("\n  Identifying failed rules...")
        
        # Rules that can be applied automatically
        auto_apply_rules = {
            "default_repository_permission": ("none", lambda x: x != "none"),
            "org_outside_collaborators": (False, lambda x: x == "Enabled"),
            "members_can_create_public_repositories": (False, lambda x: "Public: Yes" in str(x)),
            "visibility_change_disabled": (False, lambda x: x == "Enabled"),
            "delete_transfer_disabled": (False, lambda x: x == "Enabled"),
            "team_creation_disabled": (False, lambda x: x == "Enabled")
        }
        
        # Rules that cannot be applied automatically
        manual_rules = ["unsecure_org_hooks", "admin_activity_6_months"]
        
        failed_rules = [r for r in check_results if not r["passed"]]
        
        if not failed_rules:
            print("    All rules are compliant! No changes needed.")
            return {
                "backup_file": backup_file,
                "changes_made": 0,
                "errors": 0,
                "skipped": 0
            }
        
        print(f"    Found {len(failed_rules)} failed rules")
        
        # Step 3: Apply compliant settings
        print("\n  Applying compliant settings...")
        
        for result in failed_rules:
            rule_name = result["rule"]
            current_value = result["current_value"]
            
            print(f"    {rule_name}: ", end="")
            
            if rule_name in manual_rules:
                print("SKIPPED (requires manual action)")
                self.skipped.append({
                    "rule": rule_name,
                    "reason": "Cannot be applied automatically - requires manual intervention"
                })
                continue
            
            if rule_name in auto_apply_rules:
                compliant_value, _ = auto_apply_rules[rule_name]
                apply_result = self.apply_rule(rule_name, current_value, compliant_value)
                
                if apply_result["success"]:
                    self.changes_made.append(apply_result)
                    if self.dry_run:
                        print(f"Would apply ({compliant_value})")
                    else:
                        print(f"Applied ({compliant_value}) ✓")
                else:
                    self.errors.append(apply_result)
                    print(f"ERROR: {apply_result.get('error', 'Unknown error')}")
            else:
                print("SKIPPED (unknown rule)")
                self.skipped.append({
                    "rule": rule_name,
                    "reason": "Unknown rule - not in auto-apply list"
                })
        
        # Summary
        print("\n" + "-" * 40)
        print("APPLY SUMMARY")
        print("-" * 40)
        print(f"  Backup file: {backup_file}")
        print(f"  Successfully applied: {len(self.changes_made)}")
        print(f"  Skipped (manual action required): {len(self.skipped)}")
        print(f"  Errors: {len(self.errors)}")
        
        if self.dry_run:
            print("\n  *** DRY RUN - No actual changes were made ***")
            print(f"  Run without --dry-run to apply {len(self.changes_made)} changes")
        
        if self.skipped:
            print("\n  Manual action required for:")
            for item in self.skipped:
                print(f"    - {item['rule']}: {item['reason']}")
        
        return {
            "backup_file": backup_file,
            "changes_made": len(self.changes_made),
            "skipped": len(self.skipped),
            "errors": len(self.errors),
            "dry_run": self.dry_run
        }


# =============================================================================
# ROLLBACK FUNCTIONALITY
# =============================================================================

def rollback_from_backup(api_client, backup_file):
    """
    Restore organization settings from a backup file.
    
    Args:
        api_client: GitHubAPIClient instance
        backup_file: Path to backup JSON file
    
    Returns:
        dict: Summary of rollback results
    """
    print("\n" + "=" * 60)
    print("ROLLING BACK FROM BACKUP")
    print("=" * 60)
    
    # Read backup file
    with open(backup_file, "r", encoding="utf-8") as f:
        backup_data = json.load(f)
    
    org = backup_data["organization"]
    settings = backup_data["settings"]
    
    print(f"\n  Backup from: {backup_data['timestamp']}")
    print(f"  Organization: {org}")
    
    # Fields that can be restored
    restorable_fields = [
        "default_repository_permission",
        "members_can_invite_outside_collaborators",
        "members_can_create_public_repositories",
        "members_can_change_repo_visibility",
        "members_can_delete_repositories",
        "members_can_create_teams"
    ]
    
    print("\n  Restoring settings...")
    restored = 0
    errors = []
    
    for field in restorable_fields:
        value = settings.get(field)
        if value is not None:
            print(f"    {field}: ", end="")
            try:
                api_client.patch(f"/orgs/{org}", {field: value})
                print(f"Restored to {value} ✓")
                restored += 1
                time.sleep(SLEEP_INTERVAL)
            except requests.exceptions.HTTPError as e:
                print(f"ERROR: {e}")
                errors.append({"field": field, "error": str(e)})
    
    # Summary
    print("\n" + "-" * 40)
    print("ROLLBACK SUMMARY")
    print("-" * 40)
    print(f"  Settings restored: {restored}")
    print(f"  Errors: {len(errors)}")
    
    return {
        "restored": restored,
        "errors": len(errors),
        "error_details": errors
    }


# =============================================================================
# COMMAND LINE INTERFACE
# =============================================================================

def parse_arguments():
    """Parse command line arguments."""
    parser = argparse.ArgumentParser(
        description="GitHub Organization Compliance Checker & Enforcer",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  %(prog)s --check              Check compliance (default, report only)
  %(prog)s --apply              Apply compliant settings to organization
  %(prog)s --apply --dry-run    Preview changes without applying
  %(prog)s --rollback backup.json  Restore settings from backup file

Settings that can be applied automatically:
  - default_repository_permission (set to 'none')
  - org_outside_collaborators (disable)
  - members_can_create_public_repositories (disable)
  - visibility_change_disabled (disable changes)
  - delete_transfer_disabled (disable deletions)
  - team_creation_disabled (disable team creation)

Settings that require manual action:
  - unsecure_org_hooks (webhook SSL must be enabled manually)
  - admin_activity_6_months (governance review needed)
        """
    )
    
    mode_group = parser.add_mutually_exclusive_group()
    mode_group.add_argument(
        "--check", "-c",
        action="store_true",
        default=True,
        help="Check compliance and generate reports (default)"
    )
    mode_group.add_argument(
        "--apply", "-a",
        action="store_true",
        help="Apply compliant settings to organization"
    )
    mode_group.add_argument(
        "--rollback", "-r",
        metavar="BACKUP_FILE",
        help="Rollback to settings from backup file"
    )
    
    parser.add_argument(
        "--dry-run", "-n",
        action="store_true",
        help="Preview changes without applying (use with --apply)"
    )
    
    return parser.parse_args()


# =============================================================================
# MAIN ENTRY POINT
# =============================================================================

def main():
    """Main entry point."""
    args = parse_arguments()
    
    print("\n" + "=" * 60)
    print("GHE ORGANIZATION COMPLIANCE CHECKER")
    print("=" * 60)
    
    # Validate configuration
    if not GITHUB_TOKEN:
        print("ERROR: GITHUB_TOKEN environment variable is not set.")
        sys.exit(1)
    if not GITHUB_ORG:
        print("ERROR: GITHUB_ORG environment variable is not set.")
        sys.exit(1)
    
    print(f"\nConfiguration:")
    print(f"  Organization: {GITHUB_ORG}")
    print(f"  API Base URL: {GITHUB_BASE}")
    
    # Initialize API client
    api_client = GitHubAPIClient(GITHUB_BASE, GITHUB_TOKEN)
    
    # Handle ROLLBACK mode
    if args.rollback:
        print(f"  Mode: ROLLBACK from {args.rollback}")
        rollback_from_backup(api_client, args.rollback)
        print("\n" + "=" * 60)
        return
    
    # Handle CHECK and APPLY modes
    if args.apply:
        print(f"  Mode: APPLY {'(DRY RUN)' if args.dry_run else ''}")
    else:
        print(f"  Mode: CHECK (report only)")
    
    # Run compliance checks
    checker = OrgComplianceChecker(api_client, GITHUB_ORG)
    results = checker.run_all_checks()
    
    # Generate reports
    report_gen = ReportGenerator(GITHUB_ORG, results)
    report_gen.generate_all_reports()
    
    # Print summary
    passed = sum(1 for r in results if r["passed"])
    failed = sum(1 for r in results if not r["passed"])
    required_failed = sum(1 for r in results if not r["passed"] and r["enforcement"] == "Required")
    
    print("\n" + "=" * 60)
    print("CHECK SUMMARY")
    print("=" * 60)
    print(f"  Total Rules: {len(results)}")
    print(f"  Passed: {passed}")
    print(f"  Failed: {failed}")
    print(f"  Required Failed: {required_failed}")
    
    if required_failed > 0:
        print("\n  ⚠️  COMPLIANCE ISSUES DETECTED - Review required rules!")
    else:
        print("\n  ✅ All required rules passed!")
    
    # Handle APPLY mode
    if args.apply:
        if failed == 0:
            print("\n  No compliance issues to fix.")
        else:
            # Get current org data for backup
            current_org_data = checker.org_data
            
            applier = OrgComplianceApplier(api_client, GITHUB_ORG, dry_run=args.dry_run)
            apply_result = applier.apply_all(results, current_org_data)
            
            # Save apply results to file
            apply_log_file = f"org_apply_log_{datetime.now().strftime('%Y-%m-%d_%H%M%S')}.json"
            with open(apply_log_file, "w", encoding="utf-8") as f:
                json.dump({
                    "timestamp": datetime.now().isoformat(),
                    "organization": GITHUB_ORG,
                    "dry_run": args.dry_run,
                    "summary": apply_result,
                    "changes": applier.changes_made,
                    "skipped": applier.skipped,
                    "errors": applier.errors
                }, f, indent=2, default=str)
            print(f"\n  Apply log saved: {apply_log_file}")
    
    print("\n" + "=" * 60)


if __name__ == "__main__":
    main()
