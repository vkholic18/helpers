import os
import csv
import io
import datetime
import re
from sqlalchemy.orm import Session
from typing import Dict, List, Optional
from common.db import Host
import openpyxl
from register_hosts import register_hosts
# from box_sdk_gen import BoxClient, BoxCCGAuth, CCGConfig

# BOX LOGIC COMMENTED OUT - NOW USING EXCEL INSTEAD
# BOX_CLIENT_ID = os.getenv("BOX_CLIENT_ID")
# BOX_CLIENT_SECRET = os.getenv("BOX_CLIENT_SECRET")
# ENTERPRISE_ID = os.getenv("ENTERPRISE_ID")
# BOX_FOLDER_DALST = os.getenv("BOX_FOLDER_DALST")
# BOX_FOLDER_TOKST = os.getenv("BOX_FOLDER_TOKST")


# class BoxAuthenticationError(Exception):
#     """Raised when Box authentication fails."""
#     pass


class InventoryFileNotFoundError(Exception):
    """Raised when a required file is not found."""

    pass


# def box_auth(client_id: str, client_secret: str, enterprise_id: str) -> BoxClient:
#     if not all([client_id, client_secret, enterprise_id]):
#         raise BoxAuthenticationError(
#             "Missing required Box credentials. Check BOX_CLIENT_ID, BOX_CLIENT_SECRET, and ENTERPRISE_ID environment variables."
#         )
#     try:
#         ccg_config = CCGConfig(
#             client_id=client_id,
#             client_secret=client_secret,
#             enterprise_id=enterprise_id,
#         )
#         auth = BoxCCGAuth(config=ccg_config)
#         client = BoxClient(auth)
#         return client
#     except Exception as e:
#         raise BoxAuthenticationError(f"Box authentication failed: {str(e)}") from e


# def list_files_in_folder(folder_id: str, client: BoxClient) -> List[str]:
#     items = client.folders.get_folder_items(folder_id, limit=1000)
#     file_names = [item.name for item in items.entries if item.type == "file"]
#     return file_names


# def get_latest_inventory_file(file_names: List[str]) -> Optional[str]:
#     inventory_files = [
#         f for f in file_names if re.match(r"\d{2}-\d{2}-\d{2}_vCD_Inventory.*\.csv", f)
#     ]
#     if not inventory_files:
#         return None
#     try:
#         inventory_files.sort(
#             key=lambda x: datetime.datetime.strptime(x[:8], "%m-%d-%y"), reverse=True
#         )
#         return inventory_files[0]
#     except ValueError:
#         return None


# def download_file_from_box(file_name: str, folder_id: str, client: BoxClient) -> str:
#     try:
#         folder = client.folders.get_folder_items(folder_id=folder_id)
#         if folder is None:
#             raise InventoryFileNotFoundError(
#                 f"Folder with ID {folder_id} does not exist"
#             )
#         items = folder.entries
#         for item in items:
#             if item.name == file_name:
#                 file = client.downloads.download_file(item.id)
#                 content = file.read().decode("utf-8")
#                 return content
#         raise InventoryFileNotFoundError(
#             f'File "{file_name}" not found in folder {folder_id}'
#         )
#     except InventoryFileNotFoundError:
#         raise
#     except Exception as e:
#         raise Exception(f"Error downloading file {file_name}: {str(e)}") from e


# def get_vm_inventory_from_box(offering: Optional[str] = None) -> List[Dict]:
#     file_name_today = (
#         f'{datetime.datetime.now().strftime("%m-%d-%y")}_vCD_Inventory.csv'
#     )

#     folders_to_check = [BOX_FOLDER_DALST, BOX_FOLDER_TOKST]

#     if not folders_to_check:
#         raise ValueError(
#             "No Box folder IDs configured. Check BOX_FOLDER_DALST and BOX_FOLDER_TOKST environment variables."
#         )

#     vm_inventory = []
#     client = box_auth(BOX_CLIENT_ID, BOX_CLIENT_SECRET, ENTERPRISE_ID)
#     files_found = []

#     for folder_id in folders_to_check:
#         try:
#             file_names = list_files_in_folder(folder_id, client)
#             target_file = None

#             if file_name_today in file_names:
#                 target_file = file_name_today
#             else:
#                 target_file = get_latest_inventory_file(file_names)

#             if not target_file:
#                 print(f"ERROR: No inventory file found in folder {folder_id}")
#                 raise Exception(f"No inventory file found in Box folder {folder_id}")

#             file_content = download_file_from_box(target_file, folder_id, client)
#             files_found.append((folder_id, target_file, file_content))

#         except Exception as e:
#             print(
#                 f"ERROR: Error retrieving inventory from folder {folder_id}: {str(e)}"
#             )
#             raise Exception(
#                 "Error when trying to retrieve inventory reports. Please retry it later"
#             ) from e

#     for folder_id, target_file, file_content in files_found:
#         csv_file = io.StringIO(file_content)
#         first_line = csv_file.readline()
#         if not first_line.startswith("#TYPE"):
#             csv_file.seek(0)

#         reader = csv.DictReader(csv_file)

#         if not reader.fieldnames:
#             print(f"ERROR: CSV file from folder {folder_id} has no fieldnames/headers")
#             raise Exception(
#                 "Error when trying to retrieve inventory reports. Please retry it later"
#             )

#         reader.fieldnames = [name.strip() if name else "" for name in reader.fieldnames]

#         for row in reader:
#             ips = row.get("IP", "").strip()
#             vcd = row.get("vCD", "").strip()
#             org = row.get("Org", "").strip()
#             name = row.get("Name", "").strip()

#             if not ips:
#                 continue

#             if org.lower() == "public-catalog":
#                 continue

#             for ip in ips.split():
#                 ip_cleaned = ip.strip()
#                 if ip_cleaned:
#                     vm_inventory.append(
#                         {"IP": ip_cleaned, "vCD": vcd, "Org": org, "Name": name}
#                     )

#     return vm_inventory


# NEW FUNCTION: Read VM inventory from Excel file
def get_vm_inventory_from_excel(excel_file_path: str = "inventory_data.xlsx") -> List[Dict]:
    """
    Read VM inventory from Excel file instead of Box.
    Returns list of dictionaries with VM inventory data.
    """
    try:
        if not os.path.exists(excel_file_path):
            raise InventoryFileNotFoundError(f"Excel file not found: {excel_file_path}")
        
        wb = openpyxl.load_workbook(excel_file_path)
        ws = wb.active
        
        # Get headers from first row
        headers = [cell.value for cell in ws[1]]
        
        vm_inventory = []
        
        # Read data rows (skip header row)
        for row in ws.iter_rows(min_row=2, values_only=True):
            # Skip empty rows
            if not any(cell is not None for cell in row):
                continue
            
            # Create dictionary from row data
            row_dict = dict(zip(headers, row))
            
            # Extract IP address
            ip = row_dict.get("ip_address", "")
            if not ip:
                continue
            
            # Map Excel columns to inventory format (similar to Box format)
            # Excel has: data_center, ip_address, environment, platform, host_type, workload_domain, vcd_org, fqdn, category
            # We need to map to: IP, vCD (workload_domain), Org (vcd_org), Name (fqdn)
            vm_inventory.append({
                "IP": str(ip).strip(),
                "vCD": str(row_dict.get("workload_domain", "")).strip(),
                "Org": str(row_dict.get("vcd_org", "")).strip(),
                "Name": str(row_dict.get("fqdn", "")).strip(),
                # Keep extra fields for registration
                "datacenter": str(row_dict.get("data_center", "")).strip(),
                "environment": str(row_dict.get("environment", "")).strip(),
                "platform": str(row_dict.get("platform", "")).strip(),
                "host_type": str(row_dict.get("host_type", "")).strip(),
                "fqdn": str(row_dict.get("fqdn", "")).strip(),
            })
        
        print(f"[INFO] Successfully read {len(vm_inventory)} hosts from Excel file: {excel_file_path}")
        return vm_inventory
        
    except InventoryFileNotFoundError:
        raise
    except Exception as e:
        print(f"ERROR: Error reading Excel file {excel_file_path}: {str(e)}")
        raise Exception(f"Error reading Excel file: {str(e)}") from e


def list_all_hosts_for_reconciliation(
    db_session: Session, offering: Optional[str] = None
) -> List[Dict]:
    query = db_session.query(
        Host.ip_address, Host.hostname, Host.workload_domain, Host.user, Host.vcd_org
    )

    if offering:
        query = query.filter(Host.host_type == offering)
    else:
        query = query.filter(Host.host_type == "VCFaaS")

    hosts = query.all()

    return [
        {
            "ip_address": host.ip_address,
            "hostname": host.hostname,
            "workload_domain": host.workload_domain,
            "user": host.user,
            "vcd_org": host.vcd_org,
        }
        for host in hosts
    ]


def perform_inventory_reconciliation(
    db_session: Session, offering: Optional[str] = None, excel_file_path: str = "inventory_data.xlsx", 
    user_email: str = "system@vmca.com", auto_register: bool = False
) -> Dict:
    try:
        vmca_hosts = list_all_hosts_for_reconciliation(db_session, offering)
    except Exception as e:
        print(f"ERROR: Error retrieving VMCA hosts: {str(e)}")
        return {
            "statusCode": 500,
            "body": {
                "status": "error",
                "message": f"Error retrieving VMCA hosts: {str(e)}",
            },
        }

    # CHANGED: Use Excel instead of Box
    try:
        vm_inventory = get_vm_inventory_from_excel(excel_file_path)
    except InventoryFileNotFoundError as e:
        return {"statusCode": 404, "body": {"status": "error", "message": str(e)}}
    except Exception as e:
        print(f"ERROR: Error reading VM inventory from Excel: {str(e)}")
        return {
            "statusCode": 500,
            "body": {
                "status": "error",
                "message": "Error when trying to retrieve inventory from Excel file. Please retry it later",
            },
        }

    vmca_by_ip = {host.get("ip_address"): host for host in vmca_hosts}

    vm_by_ip = {}
    duplicates = []

    for vm in vm_inventory:
        ip = vm.get("IP")
        if not ip:
            continue

        if ip in vm_by_ip:
            existing_vm = vm_by_ip[ip]
            if existing_vm.get("vCD") != vm.get("vCD") or existing_vm.get(
                "Org"
            ) != vm.get("Org"):
                duplicates.append(
                    {
                        "ip_address": ip,
                        "vCD": vm.get("vCD"),
                        "Org": vm.get("Org"),
                        "Name": vm.get("Name"),
                    }
                )
        else:
            vm_by_ip[ip] = vm

    matched_hosts = []
    missing_in_vmca = []
    not_deployed = []

    for ip, vmca_host in vmca_by_ip.items():
        if ip in vm_by_ip:
            vm = vm_by_ip[ip]
            workload_domain = vmca_host.get("workload_domain", "")
            vcd = vm.get("vCD", "")
            vm_org = vm.get("Org", "")
            vmca_vcd_org = vmca_host.get("vcd_org", "")

            if (
                workload_domain
                and vcd
                and workload_domain.lower() in vcd.lower()
                and vmca_vcd_org
                and vm_org
                and vmca_vcd_org.lower() == vm_org.lower()
            ):
                matched_hosts.append(
                    {
                        "ip_address": ip,
                        "hostname": vmca_host.get("hostname"),
                        "workload_domain": workload_domain,
                        "user": vmca_host.get("user"),
                        "vcd_org": vmca_vcd_org,
                        "vCD": vcd,
                    }
                )
            else:
                duplicates.append(
                    {
                        "ip_address": ip,
                        "hostname": vmca_host.get("hostname"),
                        "workload_domain": workload_domain,
                        "user": vmca_host.get("user"),
                        "vmca_vcd_org": vmca_vcd_org,
                        "vCD": vcd,
                        "vm_org": vm_org,
                    }
                )
        else:
            not_deployed.append(
                {
                    "ip_address": ip,
                    "hostname": vmca_host.get("hostname"),
                    "workload_domain": vmca_host.get("workload_domain"),
                    "user": vmca_host.get("user"),
                    "vcd_org": vmca_host.get("vcd_org"),
                }
            )

    for ip, vm in vm_by_ip.items():
        if ip not in vmca_by_ip:
            missing_in_vmca.append(
                {
                    "ip_address": ip,
                    "vCD": vm.get("vCD"),
                    "Org": vm.get("Org"),
                    "Name": vm.get("Name"),
                }
            )

    # Auto-register missing hosts if enabled
    registration_results = []
    if auto_register and missing_in_vmca:
        for missing_host in missing_in_vmca:
            ip = missing_host.get("ip_address")
            
            # Find the corresponding VM in inventory to get full details
            vm_details = vm_by_ip.get(ip)
            if not vm_details:
                registration_results.append({
                    "ip_address": ip,
                    "status": "error",
                    "message": "VM details not found in inventory",
                    "statusCode": 500,
                })
                continue
            
            # Prepare data for register_hosts API
            host_data = {
                "ip_address": ip,
                "fqdn": vm_details.get("fqdn", vm_details.get("Name", "")),
                "environment": vm_details.get("environment", "Production"),
                "platform": vm_details.get("platform", "Unknown"),
                "datacenter": vm_details.get("datacenter", ""),
                "serial_number": f"AUTO-{ip.replace('.', '-')}",
                "host_type": vm_details.get("host_type", "VCFaaS").replace("VcFaas", "VCFaaS"),
                "workload_domain": vm_details.get("vCD", ""),
                "vcd_org": vm_details.get("Org", ""),
            }
            
            try:
                result = register_hosts([host_data], db_session, user_email)
                registration_results.append({
                    "ip_address": ip,
                    "status": result.get("body", {}).get("status", "unknown"),
                    "message": result.get("body", {}).get("message", ""),
                    "statusCode": result.get("statusCode", 0),
                })
            except Exception as e:
                registration_results.append({
                    "ip_address": ip,
                    "status": "error",
                    "message": f"Exception during registration: {str(e)}",
                    "statusCode": 500,
                })

    return {
        "statusCode": 200,
        "body": {
            "status": "success",
            "reconciliation_summary": {
                "total_vmca_hosts": len(vmca_hosts),
                "total_hosts_in_report": len(vm_inventory),
                "matched_hosts": len(matched_hosts),
                "missing_in_vmca": len(missing_in_vmca),
                "not_deployed": len(not_deployed),
                "duplicates": len(duplicates),
                "registered_hosts": len([r for r in registration_results if r.get("status") == "success"]),
                "failed_registrations": len([r for r in registration_results if r.get("status") != "success"]),
            },
            "offerings": {
                "VCFaaS": {
                    "matched_hosts": {
                        "description": "Hosts that are registered in VMCA and found in VM inventory with matching workload_domain and Org",
                        "hosts": matched_hosts,
                    },
                    "missing_in_vmca": {
                        "description": "Found in VM inventory but not registered in VMCA",
                        "hosts": missing_in_vmca,
                    },
                    "not_deployed": {
                        "description": "Host registered in VMCA but not found in VM inventory report",
                        "hosts": not_deployed,
                    },
                    "duplicates": {
                        "description": "Hosts with matching IP but mismatched workload_domain, vCD, or Org fields",
                        "hosts": duplicates,
                    },
                    "registration_results": {
                        "description": "Results of auto-registering missing hosts",
                        "hosts": registration_results,
                    } if auto_register else None,
                }
            },
        },
    }


def reconciliation_endpoint(
    db_session: Session, offering: Optional[str] = None, excel_file_path: str = "inventory_data.xlsx",
    user_email: str = "system@vmca.com", auto_register: bool = False
) -> Dict:
    """
    Endpoint for inventory reconciliation.
    
    Args:
        db_session: Database session
        offering: Optional offering type filter (e.g., "VCFaaS")
        excel_file_path: Path to Excel file with inventory data (default: "inventory_data.xlsx")
        user_email: Email to use for registering hosts (default: "system@vmca.com")
        auto_register: If True, automatically register missing hosts (default: False)
    
    Returns:
        Dict with reconciliation results and registration status
    """
    return perform_inventory_reconciliation(db_session, offering, excel_file_path, user_email, auto_register)
